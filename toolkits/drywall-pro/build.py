"""Build the Drywall Pro Toolkit workbook.

Run: python build.py
Output: dist/Drywall-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Drywall Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a list of rooms into sheets of "
         "drywall (1/2\" and 5/8\" type X), gallons of joint compound, "
         "rolls of tape, pounds of screws, linear feet of corner bead, "
         "and a print-ready customer quote — all driven by GA-214 finish "
         "level."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs — fill them in. "
         "Set the GA-214 finish level (4 is standard residential).\n"
         "2. SURFACES returns net wall area (gross minus openings) and "
         "ceiling area per room.\n"
         "3. MATERIALS converts surfaces into sheets, mud, tape, screws, "
         "and corner bead with a waste factor.\n"
         "4. LABOR returns hang + tape/mud + sand hours scaled by finish "
         "level.\n"
         "5. QUOTE is print-ready — edit prices on PRICING first."),
        ("DRYWALL SHEET MATH",
         "Standard sheet is 4 ft × 8 ft = 32 sq ft. Sheets needed = "
         "ROUNDUP(area ÷ 32 × waste). 1/2\" reg. is the residential "
         "standard for walls; 5/8\" type X is required on garage walls/"
         "ceilings, attached-garage ceilings to occupied space above, "
         "and most attic / floor-ceiling assemblies."),
        ("GA-214 FINISH LEVELS",
         "Level 1 — tape embedded only (plenums).\n"
         "Level 2 — one coat over tape (garages, tile substrate).\n"
         "Level 3 — two coats (heavy texture).\n"
         "Level 4 — three coats, sanded (standard residential paint).\n"
         "Level 5 — full skim coat over Level 4 (smooth + critical "
         "light). Mud and labor hours scale with the level."),
        ("MATERIAL COVERAGE",
         "Joint compound (ready-mix all-purpose): ~0.053 gal/sq ft "
         "(Level 4) for tape + 3 coats. Paper tape: ~0.5 LF/sq ft. "
         "Screws: 32 per 1/2\" sheet, 36 per 5/8\" sheet. Corner bead "
         "comes from the outside corners LF input."),
        ("PRICING",
         "Sheet + mud + accessory unit costs are 2026 ballpark — edit "
         "to your supplier. Margin compounds on (material + labor)."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Residence Drywall"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "PROJECT-LEVEL DEFAULTS", span=3); r += 1
    proj = [
        ("Ceiling height (ft)", 9, "0.0"),
        ("Finish level (1-5, per GA-214)", 4, "0"),
        ("Sheet size width (ft)", 4, "0.0"),
        ("Sheet size length (ft)", 8, "0.0"),
        ("Waste factor (decimal)", 0.10, "0.0%"),
        ("Outside corners (LF total)", 80, "0"),
        ("Mud coverage (gal / sq ft, L4)", 0.053, "0.000"),
        ("Tape coverage (LF / sq ft)", 0.50, "0.00"),
    ]
    proj_start = r
    for label, default, fmt in proj:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["CeilHeight"] = DefinedName("CeilHeight", attr_text=f"INPUTS!$B${proj_start}")
    wb.defined_names["FinishLevel"] = DefinedName("FinishLevel", attr_text=f"INPUTS!$B${proj_start+1}")
    wb.defined_names["SheetW"] = DefinedName("SheetW", attr_text=f"INPUTS!$B${proj_start+2}")
    wb.defined_names["SheetL"] = DefinedName("SheetL", attr_text=f"INPUTS!$B${proj_start+3}")
    wb.defined_names["Waste"] = DefinedName("Waste", attr_text=f"INPUTS!$B${proj_start+4}")
    wb.defined_names["OutsideCorners"] = DefinedName("OutsideCorners", attr_text=f"INPUTS!$B${proj_start+5}")
    wb.defined_names["MudCov"] = DefinedName("MudCov", attr_text=f"INPUTS!$B${proj_start+6}")
    wb.defined_names["TapeCov"] = DefinedName("TapeCov", attr_text=f"INPUTS!$B${proj_start+7}")

    r += 1
    n = ws.cell(r, 1, "Level 1=tape only · Level 2=1 coat · Level 3=2 coats · Level 4=3+sand (std) · Level 5=skim"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Standard sheet = 4 ft × 8 ft = 32 sq ft. 4 × 12 also common for ceilings (48 sq ft)."); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Per-room geometry goes on the SURFACES tab — separate columns for 1/2\" and 5/8\" surfaces."); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_surfaces(wb):
    ws = wb.create_sheet("SURFACES")
    auto_width(ws, [22, 9, 9, 8, 9, 12, 14, 14, 14, 12])
    banner(ws, "Surfaces by Room", "Enter L × W × openings; net wall + ceiling area per room", width=10)

    r = 4
    section(ws, r, "ROOM TAKEOFF (yellow cells = enter values)", span=10); r += 1
    headers = ["Room", "Length", "Width", "Doors", "Windows", "Ceiling Type",
               "Wall sq ft", "Ceil 1/2\" sq ft", "Ceil 5/8\" sq ft", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    rooms = [
        ("Living Room", 18, 14, 2, 3, "1/2", "main"),
        ("Kitchen", 14, 12, 2, 2, "1/2", ""),
        ("Dining", 12, 12, 1, 2, "1/2", ""),
        ("Master Bedroom", 14, 13, 2, 2, "1/2", ""),
        ("Bedroom 2", 12, 11, 1, 1, "1/2", ""),
        ("Bedroom 3", 11, 10, 1, 1, "1/2", ""),
        ("Bathroom", 9, 7, 1, 1, "1/2", "moisture-resistant"),
        ("Hallway", 12, 4, 2, 0, "1/2", ""),
        ("Garage Ceiling", 22, 20, 0, 0, "5/8", "type X required"),
        ("Garage Walls", 22, 20, 2, 1, "5/8", "type X required, walls=ceil col 5/8"),
    ]
    room_start = r
    for name, length, width, doors, windows, ctype, note in rooms:
        ws.cell(r, 1, name).font = LBL; ws.cell(r, 1).border = BOX
        for col, val, fmt in [(2, length, "0.0"), (3, width, "0.0"),
                              (4, doors, "0"), (5, windows, "0")]:
            c = ws.cell(r, col, val); c.font = MONO; c.alignment = RIGHT
            c.number_format = fmt; c.fill = YELLOW; c.border = BOX
        # Ceiling type dropdown-style
        c = ws.cell(r, 6, ctype); c.font = MONO; c.alignment = CENTER; c.fill = YELLOW; c.border = BOX
        # Walls: always 1/2" for residential. We'll route wall sf into the 1/2" bucket.
        w = ws.cell(r, 7, f"=MAX(0,2*(B{r}+C{r})*CeilHeight-D{r}*21-E{r}*15)")
        w.font = MONO_BOLD; w.alignment = RIGHT; w.number_format = "0"; w.border = BOX
        # Ceiling 1/2" = L*W if type is "1/2", else 0
        c12 = ws.cell(r, 8, f'=IF(F{r}="1/2",B{r}*C{r},0)')
        c12.font = MONO_BOLD; c12.alignment = RIGHT; c12.number_format = "0"; c12.border = BOX
        # Ceiling 5/8" = L*W if type is "5/8", else 0
        c58 = ws.cell(r, 9, f'=IF(F{r}="5/8",B{r}*C{r},0)')
        c58.font = MONO_BOLD; c58.alignment = RIGHT; c58.number_format = "0"; c58.border = BOX
        n = ws.cell(r, 10, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1
    room_end = r - 1

    ws.cell(r, 1, "TOTAL").font = LBL_BOLD; ws.cell(r, 1).fill = PANEL_FILL; ws.cell(r, 1).border = BOX
    for col in range(2, 7):
        ws.cell(r, col, "").fill = PANEL_FILL; ws.cell(r, col).border = BOX
    tw = ws.cell(r, 7, f"=SUM(G{room_start}:G{room_end})")
    tw.font = MONO_BOLD; tw.alignment = RIGHT; tw.number_format = "#,##0"; tw.fill = YELLOW; tw.border = BOX
    t12 = ws.cell(r, 8, f"=SUM(H{room_start}:H{room_end})")
    t12.font = MONO_BOLD; t12.alignment = RIGHT; t12.number_format = "#,##0"; t12.fill = YELLOW; t12.border = BOX
    t58 = ws.cell(r, 9, f"=SUM(I{room_start}:I{room_end})")
    t58.font = MONO_BOLD; t58.alignment = RIGHT; t58.number_format = "#,##0"; t58.fill = YELLOW; t58.border = BOX
    ws.cell(r, 10, "").border = BOX
    total_row = r
    wb.defined_names["WallSF"] = DefinedName("WallSF", attr_text=f"SURFACES!$G${total_row}")
    wb.defined_names["Ceil12SF"] = DefinedName("Ceil12SF", attr_text=f"SURFACES!$H${total_row}")
    wb.defined_names["Ceil58SF"] = DefinedName("Ceil58SF", attr_text=f"SURFACES!$I${total_row}")
    r += 2

    note = ws.cell(r, 1,
        "Wall area assumes 3'×7' doors (21 sf) and 3'×5' windows "
        "(15 sf) subtracted from gross. Walls always count as 1/2\" "
        "(standard residential); set Ceiling Type to 5/8 for garage / "
        "attached-garage ceilings / fire-rated assemblies. The "
        "garage-walls row uses the 5/8 column to represent type-X walls.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=10)
    return ws


def sheet_materials(wb):
    ws = wb.create_sheet("MATERIALS")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Materials Takeoff", "Sheets, mud, tape, screws, corner bead", width=5)

    r = 4
    section(ws, r, "DRYWALL SHEETS (with waste factor)", span=5); r += 1
    label_value(ws, r, "Sheet area (sq ft)",
        "=SheetW*SheetL", fmt="0").font = MONO_BOLD
    sheet_area_row = r; r += 1

    headers = ["Sheet type", "Net sq ft", "÷ sheet × (1+waste)", "Sheets", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # 1/2" sheets — walls + 1/2" ceilings
    ws.cell(r, 1, '1/2" regular').font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=WallSF+Ceil12SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    f = ws.cell(r, 3, f"=ROUNDUP(B{r}*(1+Waste)/$B${sheet_area_row},0)")
    f.font = MONO; f.alignment = RIGHT; f.number_format = "0"; f.border = BOX
    qty = ws.cell(r, 4, f"=C{r}")
    qty.font = MONO_BOLD; qty.alignment = RIGHT; qty.number_format = "0"; qty.border = BOX
    n = ws.cell(r, 5, "Walls + non-rated ceilings"); n.font = SMALL; n.border = BOX
    sheets12_row = r; r += 1

    # 5/8" type X sheets — 5/8" ceilings/walls
    ws.cell(r, 1, '5/8" type X').font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=Ceil58SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    f = ws.cell(r, 3, f"=ROUNDUP(B{r}*(1+Waste)/$B${sheet_area_row},0)")
    f.font = MONO; f.alignment = RIGHT; f.number_format = "0"; f.border = BOX
    qty = ws.cell(r, 4, f"=C{r}")
    qty.font = MONO_BOLD; qty.alignment = RIGHT; qty.number_format = "0"; qty.border = BOX
    n = ws.cell(r, 5, "Garage ceilings, fire-rated assemblies"); n.font = SMALL; n.border = BOX
    sheets58_row = r; r += 1

    wb.defined_names["Sheets12"] = DefinedName("Sheets12", attr_text=f"MATERIALS!$D${sheets12_row}")
    wb.defined_names["Sheets58"] = DefinedName("Sheets58", attr_text=f"MATERIALS!$D${sheets58_row}")

    r += 1
    section(ws, r, "FINISH MATERIALS", span=5); r += 1
    headers2 = ["Item", "Per sf factor", "Total sq ft", "Quantity", "Notes"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Joint compound — finish level multiplier
    ws.cell(r, 1, "Joint compound (gal)").font = LBL; ws.cell(r, 1).border = BOX
    pf = ws.cell(r, 2, "=MudCov*CHOOSE(FinishLevel,0.3,0.6,0.8,1,1.4)")
    pf.font = MONO; pf.alignment = RIGHT; pf.number_format = "0.0000"; pf.border = BOX
    sf = ws.cell(r, 3, "=WallSF+Ceil12SF+Ceil58SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    qty = ws.cell(r, 4, f"=ROUNDUP(B{r}*C{r},0)")
    qty.font = MONO_BOLD; qty.alignment = RIGHT; qty.number_format = "0"; qty.border = BOX
    n = ws.cell(r, 5, "All-purpose ready-mix, 5-gal pail"); n.font = SMALL; n.border = BOX
    mud_row = r; r += 1

    # Tape — paper tape per LF
    ws.cell(r, 1, "Paper tape (LF)").font = LBL; ws.cell(r, 1).border = BOX
    pf = ws.cell(r, 2, "=TapeCov")
    pf.font = MONO; pf.alignment = RIGHT; pf.number_format = "0.00"; pf.border = BOX
    sf = ws.cell(r, 3, "=WallSF+Ceil12SF+Ceil58SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    qty = ws.cell(r, 4, f"=ROUNDUP(B{r}*C{r},0)")
    qty.font = MONO_BOLD; qty.alignment = RIGHT; qty.number_format = "0"; qty.border = BOX
    n = ws.cell(r, 5, "500 LF roll typical"); n.font = SMALL; n.border = BOX
    tape_row = r; r += 1

    # Screws — 32 per 1/2 sheet, 36 per 5/8
    ws.cell(r, 1, "Screws (count)").font = LBL; ws.cell(r, 1).border = BOX
    pf = ws.cell(r, 2, ""); pf.border = BOX
    sf = ws.cell(r, 3, ""); sf.border = BOX
    qty = ws.cell(r, 4, f"=Sheets12*32+Sheets58*36")
    qty.font = MONO_BOLD; qty.alignment = RIGHT; qty.number_format = "#,##0"; qty.border = BOX
    n = ws.cell(r, 5, "1-1/4\" Type W for studs (5lb ≈ 1500 screws)"); n.font = SMALL; n.border = BOX
    screws_row = r; r += 1

    # Corner bead
    ws.cell(r, 1, "Corner bead (LF)").font = LBL; ws.cell(r, 1).border = BOX
    pf = ws.cell(r, 2, ""); pf.border = BOX
    sf = ws.cell(r, 3, ""); sf.border = BOX
    qty = ws.cell(r, 4, "=OutsideCorners")
    qty.font = MONO_BOLD; qty.alignment = RIGHT; qty.number_format = "0"; qty.border = BOX
    n = ws.cell(r, 5, "Metal or paper-faced, 8 ft sticks"); n.font = SMALL; n.border = BOX
    corner_row = r; r += 1

    wb.defined_names["MudGal"] = DefinedName("MudGal", attr_text=f"MATERIALS!$D${mud_row}")
    wb.defined_names["TapeLF"] = DefinedName("TapeLF", attr_text=f"MATERIALS!$D${tape_row}")
    wb.defined_names["Screws"] = DefinedName("Screws", attr_text=f"MATERIALS!$D${screws_row}")
    wb.defined_names["CornerLF"] = DefinedName("CornerLF", attr_text=f"MATERIALS!$D${corner_row}")
    r += 1

    note = ws.cell(r, 1,
        "Mud coverage multiplier per finish level: L1=0.3× · L2=0.6× · "
        "L3=0.8× · L4=1.0× · L5=1.4× (skim coat). Tape is paper-tape "
        "default; mesh is 1:1 in LF but only used on butt joints. "
        "Corner bead LF is a project-level input — measure all outside "
        "corners and arches at the framing stage.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=5)
    return ws


def sheet_labor(wb):
    ws = wb.create_sheet("LABOR")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Labor Hours", "Hang + tape/mud + sand, scaled by finish level", width=5)

    r = 4
    section(ws, r, "FINISH LEVEL MULTIPLIER", span=5); r += 1
    label_value(ws, r, "Finish level (from INPUTS)", "=FinishLevel", fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Tape/mud multiplier",
        '=CHOOSE(FinishLevel,0.4,0.7,1.0,1.3,1.8)', fmt="0.00").font = MONO_BOLD
    finish_mult_row = r; r += 1

    r += 1
    section(ws, r, "BASE HOURS PER SURFACE", span=5); r += 1
    headers = ["Task", "Sq ft / unit", "Hr per unit", "Hours", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Hanging — per sq ft total
    ws.cell(r, 1, "Hang (all sheets)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=WallSF+Ceil12SF+Ceil58SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.012)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Includes lift, screw, cut (RSMeans 0.011)"); n.font = SMALL; n.border = BOX
    hang_row = r; r += 1

    # Tape + mud — scaled by finish level
    ws.cell(r, 1, "Tape + mud (L4 base, finish-scaled)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=WallSF+Ceil12SF+Ceil58SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.014)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}*B{finish_mult_row}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Three coats + skim if L5"); n.font = SMALL; n.border = BOX
    mud_hours_row = r; r += 1

    # Sand
    ws.cell(r, 1, "Sand").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=WallSF+Ceil12SF+Ceil58SF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.005)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Pole sand + touch sand"); n.font = SMALL; n.border = BOX
    sand_row = r; r += 1

    # Corner bead install (per LF)
    ws.cell(r, 1, "Corner bead install (per LF)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=CornerLF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.04)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Cut, set, mud bed"); n.font = SMALL; n.border = BOX
    corner_hours_row = r; r += 1

    # Setup / cleanup
    ws.cell(r, 1, "Setup + cleanup + haul").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, 1); sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 6.0)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.0"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Flat job allowance"); n.font = SMALL; n.border = BOX
    setup_row = r; r += 1

    ws.cell(r, 1, "TOTAL LABOR HOURS").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{hang_row}:D{setup_row})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0.0"; s.fill = YELLOW; s.border = BOX
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    total_hours_row = r
    wb.defined_names["TotalHours"] = DefinedName("TotalHours", attr_text=f"LABOR!$D${total_hours_row}")
    r += 2

    note = ws.cell(r, 1,
        "Tape+mud multiplier: L1=0.4× (tape only) · L2=0.7× · L3=1.0× "
        "(2 coats, no sand) · L4=1.3× (3 coats + sand) · L5=1.8× (full "
        "skim). Hang rate 0.04 hr/sf is residential single-story; add "
        "20% for ceilings over 10 ft or stairwells.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=5)
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "MATERIALS (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ('1/2" regular drywall, 4x8',        16.00, "sheet"),
        ('5/8" type X drywall, 4x8',         22.00, "sheet"),
        ("Joint compound, 5-gal all-purpose", 22.00, "5-gal pail"),
        ("Paper tape, 500 LF roll",            5.00, "roll"),
        ("Drywall screws, 5 lb (1500 ct)",    25.00, "5-lb box"),
        ("Corner bead, 8 ft (metal/paper)",    4.00, "8 ft stick"),
        ("Misc consumables (per job)",        60.00, "job"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceSheet12"] = DefinedName("PriceSheet12", attr_text=f"PRICING!$B${pricing_start}")
    wb.defined_names["PriceSheet58"] = DefinedName("PriceSheet58", attr_text=f"PRICING!$B${pricing_start+1}")
    wb.defined_names["PriceMudPail"] = DefinedName("PriceMudPail", attr_text=f"PRICING!$B${pricing_start+2}")
    wb.defined_names["PriceTape500"] = DefinedName("PriceTape500", attr_text=f"PRICING!$B${pricing_start+3}")
    wb.defined_names["PriceScrew5lb"] = DefinedName("PriceScrew5lb", attr_text=f"PRICING!$B${pricing_start+4}")
    wb.defined_names["PriceCorner8"] = DefinedName("PriceCorner8", attr_text=f"PRICING!$B${pricing_start+5}")
    wb.defined_names["PriceMisc"] = DefinedName("PriceMisc", attr_text=f"PRICING!$B${pricing_start+6}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Labor rate (lead hanger/finisher)",  50.00, "per hour"),
        ("Profit margin (decimal)",             0.25, "0.25 = 25%"),
        ("Sales tax (decimal)",                 0.07, "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit or "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LaborRate"] = DefinedName("LaborRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+2}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Wall sq ft", "=WallSF", fmt="#,##0"); r += 1
    label_value(ws, r, "Ceiling 1/2\" sq ft", "=Ceil12SF", fmt="#,##0"); r += 1
    label_value(ws, r, "Ceiling 5/8\" sq ft", "=Ceil58SF", fmt="#,##0"); r += 1
    label_value(ws, r, "Finish level (GA-214)", "=FinishLevel", fmt="0"); r += 1
    label_value(ws, r, "Total labor hours", "=TotalHours", fmt="0.0"); r += 1

    r += 1
    section(ws, r, "MATERIALS", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    mat_start = r
    # Mud pails: ROUNDUP(MudGal / 5)
    # Tape rolls: ROUNDUP(TapeLF / 500)
    # Screw 5lb boxes: ROUNDUP(Screws / 1500)
    # Corner bead sticks: ROUNDUP(CornerLF / 8)
    mat_items = [
        ('1/2" drywall sheets',    "=Sheets12",                        "=PriceSheet12"),
        ('5/8" type X sheets',     "=Sheets58",                        "=PriceSheet58"),
        ("Joint compound (5-gal pails)", "=ROUNDUP(MudGal/5,0)",       "=PriceMudPail"),
        ("Paper tape (500 LF rolls)", "=ROUNDUP(TapeLF/500,0)",        "=PriceTape500"),
        ("Drywall screws (5 lb boxes)", "=ROUNDUP(Screws/1500,0)",     "=PriceScrew5lb"),
        ("Corner bead (8 ft sticks)", "=ROUNDUP(CornerLF/8,0)",        "=PriceCorner8"),
        ("Misc consumables",       1,                                  "=PriceMisc"),
    ]
    for desc, qty, price in mat_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        p = ws.cell(r, 3, price); p.font = MONO; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    mat_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{mat_start}:D{mat_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    mat_sub_row = r; r += 1

    r += 1
    section(ws, r, "LABOR", span=4); r += 1
    ws.cell(r, 1, "Labor (total hours × rate)").font = LBL; ws.cell(r, 1).border = BOX
    c = ws.cell(r, 2, "=TotalHours"); c.font = MONO; c.alignment = RIGHT; c.number_format = "0.0"; c.border = BOX
    p = ws.cell(r, 3, "=LaborRate"); p.font = MONO; p.alignment = RIGHT
    p.number_format = '"$"#,##0.00'; p.border = BOX
    t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
    t.number_format = '"$"#,##0.00'; t.border = BOX
    labor_total_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{mat_sub_row}+D{labor_total_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3):
        c = ws.cell(r, col, ""); c.fill = YELLOW; c.border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Texture, paint, and finishing beyond "
        "the selected GA-214 finish level are billed separately. "
        "Specialty assemblies (sound, fire-rated multi-layer) priced on "
        "request.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_surfaces(wb)
    sheet_materials(wb)
    sheet_labor(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "SURFACES", "MATERIALS", "LABOR", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Drywall-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
