"""Audit the Drywall Pro Toolkit workbook."""

import math
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"
XLSX = Path(__file__).parent / "dist" / "Drywall-Pro-Toolkit-v1.xlsx"
EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def recalc_via_libreoffice(src):
    tmp = Path(tempfile.mkdtemp(prefix="drywall_audit_"))
    cmd = [SOFFICE, "--headless", "--calc", "--norestore",
           "--convert-to", "xlsx", "--outdir", str(tmp), str(src)]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if res.returncode != 0:
        raise RuntimeError(f"soffice convert failed: {res.stderr or res.stdout}")
    out = tmp / src.name
    if not out.exists():
        raise RuntimeError(f"soffice did not produce {out}")
    return out


def scan_for_errors(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    found = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(e in v for e in EXCEL_ERRORS):
                    found.append((sheet, cell.coordinate, v))
    wb.close()
    return found


def expected_results():
    # INPUTS defaults
    ceil_height = 9.0
    finish_level = 4   # tape/mud mult 1.3; mud cov mult 1.0
    sheet_w = 4
    sheet_l = 8
    waste = 0.10
    outside_corners = 80
    mud_cov = 0.053
    tape_cov = 0.50

    # SURFACES rooms (length, width, doors, windows, ceiling_type)
    rooms = [
        (18, 14, 2, 3, "1/2"),
        (14, 12, 2, 2, "1/2"),
        (12, 12, 1, 2, "1/2"),
        (14, 13, 2, 2, "1/2"),
        (12, 11, 1, 1, "1/2"),
        (11, 10, 1, 1, "1/2"),
        (9, 7, 1, 1, "1/2"),
        (12, 4, 2, 0, "1/2"),
        (22, 20, 0, 0, "5/8"),
        (22, 20, 2, 1, "5/8"),
    ]
    wall_sf = 0
    ceil12_sf = 0
    ceil58_sf = 0
    for length, width, doors, windows, ctype in rooms:
        wall = max(0, 2 * (length + width) * ceil_height - doors * 21 - windows * 15)
        ceil = length * width
        wall_sf += wall
        if ctype == "1/2":
            ceil12_sf += ceil
        else:
            ceil58_sf += ceil

    # MATERIALS
    sheet_area = sheet_w * sheet_l  # 32
    sheets_12 = math.ceil((wall_sf + ceil12_sf) * (1 + waste) / sheet_area)
    sheets_58 = math.ceil(ceil58_sf * (1 + waste) / sheet_area)
    finish_mud_mult = {1: 0.3, 2: 0.6, 3: 0.8, 4: 1.0, 5: 1.4}[finish_level]
    total_sf = wall_sf + ceil12_sf + ceil58_sf
    mud_gal = math.ceil(mud_cov * finish_mud_mult * total_sf)
    tape_lf = math.ceil(tape_cov * total_sf)
    screws = sheets_12 * 32 + sheets_58 * 36
    corner_lf = outside_corners

    # LABOR
    finish_labor_mult = {1: 0.4, 2: 0.7, 3: 1.0, 4: 1.3, 5: 1.8}[finish_level]
    hang_hours = total_sf * 0.012
    mud_hours = total_sf * 0.014 * finish_labor_mult
    sand_hours = total_sf * 0.005
    corner_hours = corner_lf * 0.04
    setup_hours = 6.0
    total_hours = hang_hours + mud_hours + sand_hours + corner_hours + setup_hours

    # PRICING
    price_sheet_12 = 16.0
    price_sheet_58 = 22.0
    price_mud_pail = 22.0
    price_tape_500 = 5.0
    price_screw_5lb = 25.0
    price_corner_8 = 4.0
    price_misc = 60.0
    labor_rate = 50.0
    margin = 0.25
    tax_rate = 0.07

    # QUOTE quantities (ROUNDUP)
    mud_pails = math.ceil(mud_gal / 5)
    tape_rolls = math.ceil(tape_lf / 500)
    screw_boxes = math.ceil(screws / 1500)
    corner_sticks = math.ceil(corner_lf / 8)

    materials = (
        sheets_12 * price_sheet_12 +
        sheets_58 * price_sheet_58 +
        mud_pails * price_mud_pail +
        tape_rolls * price_tape_500 +
        screw_boxes * price_screw_5lb +
        corner_sticks * price_corner_8 +
        1 * price_misc
    )
    labor_cost = total_hours * labor_rate
    cost_sub = materials + labor_cost
    margin_amt = cost_sub * margin
    pretax = cost_sub + margin_amt
    tax = pretax * tax_rate
    total_due = pretax + tax

    return {
        "WallSF": wall_sf,
        "Ceil12SF": ceil12_sf,
        "Ceil58SF": ceil58_sf,
        "Sheets12": sheets_12,
        "Sheets58": sheets_58,
        "MudGal": mud_gal,
        "TapeLF": tape_lf,
        "Screws": screws,
        "CornerLF": corner_lf,
        "TotalHours": total_hours,
        "Materials": materials,
        "Labor": labor_cost,
        "CostSub": cost_sub,
        "Margin": margin_amt,
        "Pretax": pretax,
        "Tax": tax,
        "TotalDue": total_due,
    }


def approx(a, b, tol_pct=0.005, tol_abs=0.01):
    if a is None or b is None:
        return False
    if a == b:
        return True
    if abs(a - b) <= tol_abs:
        return True
    denom = max(abs(a), abs(b))
    return denom > 0 and abs(a - b) / denom <= tol_pct


def main():
    if not XLSX.exists():
        print(f"FAIL: {XLSX} does not exist — run build.py first.")
        sys.exit(1)

    print(f"Source: {XLSX}")
    print("Recalculating via LibreOffice headless...")
    recalced = recalc_via_libreoffice(XLSX)
    print(f"  -> {recalced}")

    errors = scan_for_errors(recalced)
    if errors:
        print(f"\nFAIL: {len(errors)} Excel error cell(s):")
        for sheet, coord, val in errors[:20]:
            print(f"  {sheet}!{coord} = {val}")
        sys.exit(1)
    print("OK: no #REF!/#DIV/0!/#VALUE!/#NAME?/#NULL!/#NUM!/#N/A in any cell")

    wb = load_workbook(recalced, data_only=True)
    wb_f = load_workbook(XLSX, data_only=False)
    exp = expected_results()

    def resolve(name):
        ref = wb_f.defined_names[name].value
        sheet, coord = ref.split("!")
        coord = coord.replace("$", "")
        return wb[sheet][coord].value

    quote = wb["QUOTE"]
    quote_vals = {}
    for row in quote.iter_rows():
        label_cell = row[0]
        total_cell = row[3] if len(row) >= 4 else None
        if isinstance(label_cell.value, str) and total_cell is not None:
            quote_vals[label_cell.value.strip()] = total_cell.value

    checks = [
        ("WallSF",      resolve("WallSF"),      exp["WallSF"]),
        ("Ceil12SF",    resolve("Ceil12SF"),    exp["Ceil12SF"]),
        ("Ceil58SF",    resolve("Ceil58SF"),    exp["Ceil58SF"]),
        ("Sheets12",    resolve("Sheets12"),    exp["Sheets12"]),
        ("Sheets58",    resolve("Sheets58"),    exp["Sheets58"]),
        ("MudGal",      resolve("MudGal"),      exp["MudGal"]),
        ("TapeLF",      resolve("TapeLF"),      exp["TapeLF"]),
        ("Screws",      resolve("Screws"),      exp["Screws"]),
        ("CornerLF",    resolve("CornerLF"),    exp["CornerLF"]),
        ("TotalHours",  resolve("TotalHours"),  exp["TotalHours"]),
        ("Materials sub", quote_vals.get("Materials subtotal"),               exp["Materials"]),
        ("Cost sub",      quote_vals.get("Cost subtotal (materials + labor)"), exp["CostSub"]),
        ("Profit margin", quote_vals.get("Profit margin"),                     exp["Margin"]),
        ("Pre-tax total", quote_vals.get("Pre-tax total"),                     exp["Pretax"]),
        ("Sales tax",     quote_vals.get("Sales tax"),                         exp["Tax"]),
        ("TOTAL DUE",     quote_vals.get("TOTAL DUE"),                         exp["TotalDue"]),
    ]

    print("\nNumeric checks (workbook vs. pure-Python expected):")
    print(f"  {'Label':<22} {'Workbook':>14} {'Expected':>14}   Result")
    fails = 0
    for label, got, expv in checks:
        try:
            got_n = float(got) if got is not None else None
            exp_n = float(expv)
        except (TypeError, ValueError):
            print(f"  {label:<22} {'(not numeric)':>14} {'':>14}   FAIL")
            fails += 1
            continue
        ok = approx(got_n, exp_n)
        flag = "OK  " if ok else "FAIL"
        print(f"  {label:<22} {got_n:>14,.2f} {exp_n:>14,.2f}   {flag}")
        if not ok:
            fails += 1

    wb.close()
    wb_f.close()
    shutil.rmtree(recalced.parent, ignore_errors=True)

    if fails:
        print(f"\nFAIL: {fails} of {len(checks)} numeric checks failed")
        sys.exit(1)
    print(f"\nPASS: {len(checks)} checks, 0 Excel errors, {XLSX.name} is clean")


if __name__ == "__main__":
    main()
