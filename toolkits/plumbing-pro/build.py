"""Build the Plumbing Pro Toolkit workbook.

Run: python build.py
Output: dist/Plumbing-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

# Brand
HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Plumbing Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a fixture list into pipe sizes for "
         "water supply, drain (DFU), and vent — plus a print-ready customer "
         "quote with materials, labor, margin, and tax."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs.\n"
         "2. Open FIXTURE-UNITS — type quantities for each fixture; the "
         "sheet returns total WSFU and DFU.\n"
         "3. PIPE-SIZING auto-picks supply, drain, and trap-arm sizes from "
         "IPC tables.\n"
         "4. VENT confirms vent stack and branch sizing.\n"
         "5. PRICING — set unit costs once.\n"
         "6. Print QUOTE."),
        ("WSFU LOOKUP (IPC E103)",
         "Water Supply Fixture Units (WSFU) drive the supply pipe size. The "
         "toolkit uses the standard private-dwelling table (toilet 2.2, "
         "lavatory 0.7, kitchen sink 1.4, shower 1.4, bathtub 4.0, "
         "dishwasher 1.4, washing machine 1.4, hose bibb 2.5)."),
        ("DFU LOOKUP (IPC 709)",
         "Drainage Fixture Units (DFU) drive drain + vent pipe size. Toilet "
         "= 4.0 DFU (1.6 gpf), lavatory 1.0, shower 2.0, bathtub 2.0, "
         "kitchen sink 2.0, dishwasher 2.0, washing machine 3.0."),
        ("PIPE SIZING",
         "Supply: pipes ≤120 ft total developed length, 30–60 psi. Drain: "
         "horizontal branches at minimum slope 1/4\" per ft for 2.5\" or "
         "smaller, 1/8\" per ft for 3\"+."),
        ("VENT SIZING",
         "Vent stack diameter must be ≥ half the diameter of the drain it "
         "serves and never less than 1¼\". The VENT tab includes the IPC "
         "906 length-vs-DFU table."),
        ("PRICING",
         "Unit prices on PRICING are 2026 ballpark — edit to your supplier. "
         "Labor rate is per hour with a separate helper line. Margin "
         "compounds on (material + labor)."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please don't "
         "redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label); a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top"); a.fill = PANEL_FILL
        b = ws.cell(r, 2, body); b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Bathroom Remodel"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "BUILDING + SERVICE", span=3); r += 1
    bldg = [
        ("Number of bathrooms", 2, "0"),
        ("Number of stories", 2, "0"),
        ("Service street pressure (psi)", 60, "0"),
        ("Static head loss (psi, per 10 ft of rise)", 4.3, "0.0"),
        ("Total developed length (ft)", 95, "0"),
        ("Pipe material (supply)", "PEX", None),  # PEX | Copper Type L | CPVC
        ("Pipe material (drain)", "PVC DWV", None),  # PVC DWV | Cast Iron
    ]
    bldg_start = r
    for label, default, fmt in bldg:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["NumBaths"] = DefinedName("NumBaths", attr_text=f"INPUTS!$B${bldg_start}")
    wb.defined_names["Stories"] = DefinedName("Stories", attr_text=f"INPUTS!$B${bldg_start+1}")
    wb.defined_names["ServicePSI"] = DefinedName("ServicePSI", attr_text=f"INPUTS!$B${bldg_start+2}")
    wb.defined_names["StaticPSI"] = DefinedName("StaticPSI", attr_text=f"INPUTS!$B${bldg_start+3}")
    wb.defined_names["DevLength"] = DefinedName("DevLength", attr_text=f"INPUTS!$B${bldg_start+4}")
    wb.defined_names["SupplyMat"] = DefinedName("SupplyMat", attr_text=f"INPUTS!$B${bldg_start+5}")
    wb.defined_names["DrainMat"] = DefinedName("DrainMat", attr_text=f"INPUTS!$B${bldg_start+6}")

    r += 1
    n = ws.cell(r, 1, "Supply pipe options: PEX | Copper Type L | CPVC"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Drain pipe options: PVC DWV | Cast Iron"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1,
        "Static head: 0.433 psi/ft elevation. 10 ft rise ≈ 4.3 psi loss.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_fixture_units(wb):
    ws = wb.create_sheet("FIXTURE-UNITS")
    auto_width(ws, [34, 12, 12, 12, 14, 14])
    banner(ws, "Fixture Units (WSFU + DFU)", "Type quantity in column B — totals roll into PIPE-SIZING", width=6)

    r = 4
    headers = ["Fixture", "Qty", "WSFU ea", "DFU ea", "Total WSFU", "Total DFU"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1

    fixtures = [
        ("Water closet (1.6 gpf, tank)",   2, 2.2, 4.0),
        ("Lavatory",                        4, 0.7, 1.0),
        ("Bathtub",                         1, 4.0, 2.0),
        ("Shower (separate, 2.5 gpm)",      1, 1.4, 2.0),
        ("Kitchen sink",                    1, 1.4, 2.0),
        ("Dishwasher",                      1, 1.4, 2.0),
        ("Clothes washer (residential)",    1, 1.4, 3.0),
        ("Hose bibb (sillcock, ½\")",       2, 2.5, 0.0),
        ("Bidet",                           0, 1.0, 1.0),
        ("Laundry sink (single)",           0, 1.4, 2.0),
        ("Floor drain (3\" trap)",          1, 0.0, 3.0),
    ]

    fix_start = r
    for label, qty, wsfu, dfu in fixtures:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        q = ws.cell(r, 2, qty); q.font = MONO_BOLD; q.alignment = RIGHT; q.fill = YELLOW; q.border = BOX
        w = ws.cell(r, 3, wsfu); w.font = MONO; w.alignment = RIGHT; w.number_format = "0.0"; w.border = BOX
        df = ws.cell(r, 4, dfu); df.font = MONO; df.alignment = RIGHT; df.number_format = "0.0"; df.border = BOX
        tw = ws.cell(r, 5, f"=B{r}*C{r}"); tw.font = MONO_BOLD; tw.alignment = RIGHT; tw.number_format = "0.0"; tw.border = BOX
        td = ws.cell(r, 6, f"=B{r}*D{r}"); td.font = MONO_BOLD; td.alignment = RIGHT; td.number_format = "0.0"; td.border = BOX
        r += 1
    fix_end = r - 1

    # Totals row
    ws.cell(r, 1, "TOTALS").font = H3; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3, 4):
        ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    tot_w = ws.cell(r, 5, f"=SUM(E{fix_start}:E{fix_end})")
    tot_w.font = Font(name="Arial Black", size=12, color=INK, bold=True)
    tot_w.alignment = RIGHT; tot_w.number_format = "0.0"; tot_w.border = BOX; tot_w.fill = YELLOW
    tot_d = ws.cell(r, 6, f"=SUM(F{fix_start}:F{fix_end})")
    tot_d.font = Font(name="Arial Black", size=12, color=INK, bold=True)
    tot_d.alignment = RIGHT; tot_d.number_format = "0.0"; tot_d.border = BOX; tot_d.fill = YELLOW
    ws.row_dimensions[r].height = 24
    totals_row = r; r += 1

    wb.defined_names["TotalWSFU"] = DefinedName("TotalWSFU", attr_text=f"'FIXTURE-UNITS'!$E${totals_row}")
    wb.defined_names["TotalDFU"] = DefinedName("TotalDFU", attr_text=f"'FIXTURE-UNITS'!$F${totals_row}")

    r += 1
    note = ws.cell(r, 1,
        "WSFU/DFU values from IPC Table E103.3 (private dwellings) and IPC "
        "Table 709.1. Hose bibbs carry WSFU but no DFU (they discharge to "
        "grade, not the DWV system).")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    return ws


def sheet_pipe_sizing(wb):
    ws = wb.create_sheet("PIPE-SIZING")
    auto_width(ws, [38, 14, 18, 36])
    banner(ws, "Pipe Sizing (Supply + Drain)", "Auto-pick from IPC tables based on FIXTURE-UNITS totals", width=4)

    r = 4
    section(ws, r, "WATER SUPPLY (IPC E103, ≤60 psi street pressure)", span=4); r += 1
    label_value(ws, r, "Total WSFU (from FIXTURE-UNITS)",
        "=TotalWSFU", fmt="0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Available pressure (psi)",
        "=ServicePSI-StaticPSI*Stories", fmt="0").font = MONO_BOLD
    avail_row = r; r += 1
    label_value(ws, r, "Pressure loss budget (psi)",
        f"=$B${avail_row}-15", fmt="0").font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "SUPPLY MAIN SIZE (residential, IPC table)", span=4); r += 1
    headers = ["WSFU range", "Service main", "Distribution branch", "Note"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    supply_table = [
        ("≤ 6 WSFU",     "3/4\"",  "1/2\"",  "Single bath / small ADU"),
        ("7 - 16 WSFU",  "1\"",    "3/4\"",  "Typical 1-2 bath home"),
        ("17 - 32 WSFU", "1\"",    "3/4\"",  "3 bath / dual-laundry"),
        ("33 - 50 WSFU", "1-1/4\"","1\"",    "Large home / 4+ bath"),
        ("51 - 100 WSFU","1-1/2\"","1-1/4\"","Multi-family duplex"),
    ]
    for rng, main, branch, note in supply_table:
        ws.cell(r, 1, rng).font = MONO; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, main).font = MONO_BOLD; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, branch).font = MONO_BOLD; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    # Auto-pick
    label_value(ws, r, "RECOMMENDED SERVICE MAIN",
        '=IF(TotalWSFU<=6,"3/4 in",IF(TotalWSFU<=32,"1 in",IF(TotalWSFU<=50,"1-1/4 in","1-1/2 in")))',
        fmt=None).font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1
    label_value(ws, r, "RECOMMENDED DISTRIBUTION BRANCH",
        '=IF(TotalWSFU<=6,"1/2 in",IF(TotalWSFU<=32,"3/4 in",IF(TotalWSFU<=50,"1 in","1-1/4 in")))',
        fmt=None).font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1

    r += 1
    section(ws, r, "DRAIN PIPE SIZE (IPC 710, horizontal branches)", span=4); r += 1
    headers2 = ["DFU range", "Pipe size", "Min slope", "Note"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    drain_table = [
        ("1 - 3 DFU",   "1-1/2\"", "1/4\"/ft", "Lavatory branches only"),
        ("4 - 6 DFU",   "2\"",     "1/4\"/ft", "Showers, kitchen sink branches"),
        ("7 - 20 DFU",  "3\"",     "1/8\"/ft", "Includes water closet"),
        ("21 - 160 DFU","4\"",     "1/8\"/ft", "Building drain typical home"),
        ("161 - 620",   "6\"",     "1/16\"/ft","Multi-family / commercial"),
    ]
    for rng, size, slope, note in drain_table:
        ws.cell(r, 1, rng).font = MONO; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, size).font = MONO_BOLD; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, slope).font = MONO; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    label_value(ws, r, "RECOMMENDED BUILDING DRAIN",
        '=IF(TotalDFU<=3,"1-1/2 in",IF(TotalDFU<=6,"2 in",IF(TotalDFU<=20,"3 in",IF(TotalDFU<=160,"4 in","6 in"))))',
        fmt=None).font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1

    r += 1
    section(ws, r, "TRAP & TRAP-ARM SIZE (IPC 1002.1)", span=4); r += 1
    headers3 = ["Fixture", "Trap size", "Trap arm max", "Note"]
    for i, h in enumerate(headers3, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    traps = [
        ("Lavatory",     "1-1/4\"", "30\"", "Distance trap weir to vent"),
        ("Bathtub",      "1-1/2\"", "42\"", ""),
        ("Shower",       "2\"",     "60\"", ""),
        ("Kitchen sink", "1-1/2\"", "42\"", "2\" if disposer present"),
        ("Dishwasher",   "1-1/2\"", "42\"", "Indirect / air gap to sink"),
        ("Washing mach.", "2\"",     "60\"", "Standpipe 18-30\" tall"),
        ("Floor drain",   "2\" or 3\"", "60-72\"", "3\" recommended for laundry"),
    ]
    for fix, ts, ta, note in traps:
        ws.cell(r, 1, fix).font = LBL; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, ts).font = MONO_BOLD; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, ta).font = MONO; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    return ws


def sheet_vent(wb):
    ws = wb.create_sheet("VENT")
    auto_width(ws, [34, 14, 14, 30])
    banner(ws, "Vent Sizing (IPC 906)", "Vent stack diameter vs. drain DFU and developed length", width=4)

    r = 4
    section(ws, r, "VENT STACK SIZE (IPC 906.1)", span=4); r += 1
    label_value(ws, r, "Total drain DFU (from FIXTURE-UNITS)", "=TotalDFU", fmt="0.0").font = MONO_BOLD; r += 1

    r += 1
    headers = ["Drain Size", "Max DFU on drain", "Min Vent Size", "Max Vent Length (ft)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    vent_table = [
        ('1-1/2"',  8,    '1-1/4"', 50),
        ('2"',      24,   '1-1/2"', 80),
        ('3"',      102,  '2"',     200),
        ('4"',      540,  '3"',     400),
        ('6"',      1300, '4"',     600),
    ]
    for ds, dfu, vs, vl in vent_table:
        ws.cell(r, 1, ds).font = MONO_BOLD; ws.cell(r, 1).alignment = CENTER; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, dfu).font = MONO; ws.cell(r, 2).alignment = RIGHT; ws.cell(r, 2).number_format = "#,##0"; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, vs).font = MONO_BOLD; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        ws.cell(r, 4, vl).font = MONO; ws.cell(r, 4).alignment = RIGHT; ws.cell(r, 4).border = BOX
        r += 1

    r += 1
    section(ws, r, "RECOMMENDED VENT (auto)", span=4); r += 1
    label_value(ws, r, "Building drain (from PIPE-SIZING)",
        '=IF(TotalDFU<=3,"1-1/2 in",IF(TotalDFU<=6,"2 in",IF(TotalDFU<=20,"3 in",IF(TotalDFU<=160,"4 in","6 in"))))',
        fmt=None).font = MONO_BOLD; r += 1
    label_value(ws, r, "Min vent stack size",
        '=IF(TotalDFU<=8,"1-1/4 in",IF(TotalDFU<=24,"1-1/2 in",IF(TotalDFU<=102,"2 in",IF(TotalDFU<=540,"3 in","4 in"))))',
        fmt=None).font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1

    r += 2
    section(ws, r, "AIR ADMITTANCE VALVES (AAV) — IPC 918", span=4); r += 1
    note = ws.cell(r, 1,
        "AAVs are permitted as an alternative to vent extensions through the "
        "roof for individual fixture vents. Each AAV must be sized for the "
        "branch DFU it serves: 1.5\" AAV ≤ 8 DFU, 2\" AAV ≤ 16 DFU, 3\" AAV "
        "≤ 100 DFU. AAVs cannot serve as the building's main vent.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4); r += 3

    note2 = ws.cell(r, 1,
        "Vent stack must extend through the roof, terminate ≥6\" above roof "
        "and ≥10 ft from any opening. Stack diameter ≥ ½ × drain diameter, "
        "minimum 1-1/4\".")
    note2.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "MATERIALS (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("PEX-A 1/2\" tubing (100 ft)",        65.00,  "roll"),
        ("PEX-A 3/4\" tubing (100 ft)",        110.00, "roll"),
        ("PEX-A 1\" tubing (100 ft)",          175.00, "roll"),
        ("PEX expansion ring 1/2\"",           0.45,   "ea"),
        ("PEX expansion ring 3/4\"",           0.60,   "ea"),
        ("Copper Type L 1/2\" (10 ft)",        38.00,  "stick"),
        ("Copper Type L 3/4\" (10 ft)",        58.00,  "stick"),
        ("Copper 1/2\" elbow",                 1.85,   "ea"),
        ("Copper 1/2\" tee",                   2.40,   "ea"),
        ("PVC DWV 1-1/2\" (10 ft)",            14.00,  "stick"),
        ("PVC DWV 2\" (10 ft)",                22.00,  "stick"),
        ("PVC DWV 3\" (10 ft)",                42.00,  "stick"),
        ("PVC DWV 4\" (10 ft)",                68.00,  "stick"),
        ("PVC primer + cement (set)",          18.00,  "set"),
        ("P-trap 1-1/2\"",                     8.50,   "ea"),
        ("P-trap 2\"",                         12.00,  "ea"),
        ("Toilet flange (PVC)",                14.00,  "ea"),
        ("Shut-off valve, quarter-turn 1/2\"", 11.00,  "ea"),
        ("Water heater (40 gal gas, 0.62 UEF)",1100.00,"ea"),
        ("Water heater (50 gal electric)",     650.00, "ea"),
        ("Heat pump water heater 50 gal",      1850.00,"ea"),
        ("Toilet (1.28 gpf, two-piece)",       240.00, "ea"),
        ("Lavatory faucet",                    140.00, "ea"),
        ("Kitchen faucet",                     220.00, "ea"),
        ("Tub/shower valve (rough-in + trim)", 320.00, "ea"),
        ("Sillcock (frost-free 1/2\")",        38.00,  "ea"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceTable"] = DefinedName("PriceTable", attr_text=f"PRICING!$A${pricing_start}:$C${r-1}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Lead plumber labor",          110.00, "per hour"),
        ("Helper labor",                65.00,  "per hour"),
        ("Hours per fixture rough-in",  3.0,    "hr/fixture"),
        ("Hours per fixture set/trim",  1.5,    "hr/fixture"),
        ("Permit + inspection",         225.00, "flat"),
        ("Drain cleaning / haul-away",  150.00, "flat"),
        ("Profit margin (decimal)",     0.25,   "0.25 = 25%"),
        ("Sales tax (decimal)",         0.07,   "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit or "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        elif "hr/" in unit or unit == "hr/fixture":
            v.number_format = "0.0"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LeadRate"] = DefinedName("LeadRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["HelperRate"] = DefinedName("HelperRate", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["RoughHrs"] = DefinedName("RoughHrs", attr_text=f"PRICING!$B${labor_start+2}")
    wb.defined_names["SetHrs"] = DefinedName("SetHrs", attr_text=f"PRICING!$B${labor_start+3}")
    wb.defined_names["Permit"] = DefinedName("Permit", attr_text=f"PRICING!$B${labor_start+4}")
    wb.defined_names["DrainHaul"] = DefinedName("DrainHaul", attr_text=f"PRICING!$B${labor_start+5}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+6}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+7}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Total WSFU", "=TotalWSFU", fmt="0.0"); r += 1
    label_value(ws, r, "Total DFU", "=TotalDFU", fmt="0.0"); r += 1
    label_value(ws, r, "Service main",
        '=IF(TotalWSFU<=6,"3/4 in",IF(TotalWSFU<=32,"1 in",IF(TotalWSFU<=50,"1-1/4 in","1-1/2 in")))'); r += 1
    label_value(ws, r, "Building drain",
        '=IF(TotalDFU<=3,"1-1/2 in",IF(TotalDFU<=6,"2 in",IF(TotalDFU<=20,"3 in",IF(TotalDFU<=160,"4 in","6 in"))))'); r += 1

    r += 1
    section(ws, r, "MATERIALS (typical 2-bath remodel)", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1

    eq_start = r
    eq_items = [
        ("PEX-A 1/2\" tubing", 3, "PEX-A 1/2\" tubing (100 ft)"),
        ("PEX-A 3/4\" tubing", 2, "PEX-A 3/4\" tubing (100 ft)"),
        ("PEX expansion rings 1/2\"", 80, "PEX expansion ring 1/2\""),
        ("PEX expansion rings 3/4\"", 40, "PEX expansion ring 3/4\""),
        ("PVC DWV 2\" sticks", 6, "PVC DWV 2\" (10 ft)"),
        ("PVC DWV 3\" sticks", 5, "PVC DWV 3\" (10 ft)"),
        ("PVC primer + cement", 2, "PVC primer + cement (set)"),
        ("P-traps 1-1/2\"", 4, "P-trap 1-1/2\""),
        ("Toilet flanges", 2, "Toilet flange (PVC)"),
        ("Shut-off valves", 12, "Shut-off valve, quarter-turn 1/2\""),
        ("Toilets (1.28 gpf)", 2, "Toilet (1.28 gpf, two-piece)"),
        ("Lavatory faucets", 4, "Lavatory faucet"),
        ("Tub/shower valve set", 1, "Tub/shower valve (rough-in + trim)"),
    ]
    for desc, qty, price_label in eq_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        escaped = price_label.replace('"', '""')
        p = ws.cell(r, 3, f'=VLOOKUP("{escaped}",PriceTable,2,FALSE)')
        p.font = MONO; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    eq_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{eq_start}:D{eq_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    eq_subtotal_row = r; r += 1

    r += 1
    section(ws, r, "LABOR & FEES", span=4); r += 1
    # Total fixtures = sum of fixture qty (B5..B15 in FIXTURE-UNITS), exclude floor drain & hose bibb
    labor_lines = [
        ("Rough-in labor (lead, hr × fixtures)",   "='FIXTURE-UNITS'!B5*RoughHrs*LeadRate"),
        ("Rough-in labor (helper)",                "='FIXTURE-UNITS'!B5*RoughHrs*HelperRate*0.5"),
        ("Set + trim labor",                       "=(SUM('FIXTURE-UNITS'!B5:B15))*SetHrs*LeadRate"),
        ("Permit + inspection",                    "=Permit"),
        ("Drain cleaning / haul-away",             "=DrainHaul"),
    ]
    labor_start = r
    for desc, formula in labor_lines:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        for col in (2, 3): ws.cell(r, col, "").border = BOX
        t = ws.cell(r, 4, formula); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    labor_end = r - 1

    ws.cell(r, 1, "Labor subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{labor_start}:D{labor_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    labor_subtotal_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{eq_subtotal_row}+D{labor_subtotal_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Material prices subject to change. Work "
        "scheduled within 2-3 weeks of signed acceptance.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_fixture_units(wb)
    sheet_pipe_sizing(wb)
    sheet_vent(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "FIXTURE-UNITS", "PIPE-SIZING", "VENT", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Plumbing-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
