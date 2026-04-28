"""Build the Roofing Pro Toolkit workbook.

Run: python build.py
Output: dist/Roofing-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

# Brand
HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None, name=None):
    """Render a 'Label | Value' row in cols A:B, return value cell."""
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD
    a.alignment = LEFT
    a.fill = PANEL_FILL
    a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO
    b.alignment = LEFT
    b.fill = PANEL2_FILL
    b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2
    c.fill = YELLOW
    c.alignment = LEFT
    return c


def line_item(ws, row, label, qty_formula, unit, qty_fmt="0.0", price_col=None, price_default=None):
    a = ws.cell(row, 1, label)
    a.font = LBL
    a.alignment = LEFT
    a.border = BOX
    q = ws.cell(row, 2, qty_formula)
    q.font = MONO
    q.alignment = RIGHT
    q.number_format = qty_fmt
    q.border = BOX
    u = ws.cell(row, 3, unit)
    u.font = LBL
    u.alignment = LEFT
    u.border = BOX
    if price_col is not None:
        p = ws.cell(row, price_col, price_default)
        p.font = MONO
        p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'
        p.fill = PANEL2_FILL
        p.border = BOX
    return q


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Roofing Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Six linked tabs that turn a roof footprint and pitch into a complete "
         "material takeoff, ventilation balance, and customer-ready quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Fill in customer info and roof specs. "
         "Yellow cells are inputs — gray cells are calculated and locked.\n"
         "2. The TAKEOFF tab shows everything to buy.\n"
         "3. The VENTILATION tab confirms intake/exhaust balance per code.\n"
         "4. The QUOTE tab is print-ready — adjust unit prices in PRICING first."),
        ("WHEN TO RE-ENTER VALUES",
         "Per project. Save a copy with the customer name (e.g. "
         "'Roof – Smith 421 Maple.xlsx') so your originals stay clean."),
        ("PITCH NOTES",
         "Pitch is rise-per-12 (e.g. enter 6 for a 6/12 roof). The PITCH "
         "tab has angle, slope multiplier, and percent grade for every "
         "common pitch from 2/12 to 18/12."),
        ("VENTILATION RULE",
         "Code default: 1 sq ft of net free area (NFA) per 150 sq ft of "
         "attic floor, or 1:300 if a vapor retarder is present and intake/"
         "exhaust is balanced 50/50. The toolkit defaults to 1:150."),
        ("PRICING",
         "Unit prices on the PRICING tab are 2026 ballpark — edit to your "
         "supplier. Labor rate per square is also editable. Margin "
         "compounds on (material + labor)."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(30, body.count("\n") * 18 + 36)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [34, 18, 26])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Residence Re-Roof"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "ROOF GEOMETRY", span=3); r += 1
    rows = [
        ("Building length (ft)", 50, "0.0"),
        ("Building width (ft)", 30, "0.0"),
        ("Roof pitch (rise per 12)", 6, "0"),
        ("Overhang per side (ft)", 1.0, "0.00"),
        ("Roof type", "Gable", None),  # cosmetic — used in quote
    ]
    geom_start = r
    for label, default, fmt in rows:
        c = label_value(ws, r, label, default, fmt=fmt)
        c.fill = YELLOW
        r += 1
    # Named ranges for downstream sheets
    wb.defined_names["LengthFt"] = DefinedName("LengthFt", attr_text=f"INPUTS!$B${geom_start}")
    wb.defined_names["WidthFt"]  = DefinedName("WidthFt",  attr_text=f"INPUTS!$B${geom_start+1}")
    wb.defined_names["Pitch"]    = DefinedName("Pitch",    attr_text=f"INPUTS!$B${geom_start+2}")
    wb.defined_names["Overhang"] = DefinedName("Overhang", attr_text=f"INPUTS!$B${geom_start+3}")

    r += 1
    section(ws, r, "ROOF FEATURES (linear feet)", span=3); r += 1
    feat_start = r
    feat_rows = [
        ("Ridge length (ft)", 50, "0.0"),
        ("Hip length (ft)", 0, "0.0"),
        ("Valley length (ft)", 0, "0.0"),
        ("Eave length (drip edge)", 100, "0.0"),
        ("Rake length (drip edge)", 64, "0.0"),
        ("Step flashing run (ft)", 0, "0.0"),
        ("Pipe boots (count)", 2, "0"),
    ]
    for label, default, fmt in feat_rows:
        c = label_value(ws, r, label, default, fmt=fmt)
        c.fill = YELLOW
        r += 1
    wb.defined_names["Ridge"]    = DefinedName("Ridge",    attr_text=f"INPUTS!$B${feat_start}")
    wb.defined_names["Hip"]      = DefinedName("Hip",      attr_text=f"INPUTS!$B${feat_start+1}")
    wb.defined_names["Valley"]   = DefinedName("Valley",   attr_text=f"INPUTS!$B${feat_start+2}")
    wb.defined_names["Eave"]     = DefinedName("Eave",     attr_text=f"INPUTS!$B${feat_start+3}")
    wb.defined_names["Rake"]     = DefinedName("Rake",     attr_text=f"INPUTS!$B${feat_start+4}")
    wb.defined_names["StepFlash"] = DefinedName("StepFlash", attr_text=f"INPUTS!$B${feat_start+5}")
    wb.defined_names["PipeBoots"] = DefinedName("PipeBoots", attr_text=f"INPUTS!$B${feat_start+6}")

    r += 1
    section(ws, r, "MATERIALS & SCOPE", span=3); r += 1
    mat_start = r
    mat_rows = [
        ("Tear-off layers (0/1/2/3)", 1, "0"),
        ("Underlayment type", "Synthetic", None),  # Synthetic / 15# Felt / 30# Felt
        ("Shingle type", "Architectural", None),   # 3-Tab / Architectural / Designer
        ("Ice & water shield (lin ft)", 100, "0.0"),
        ("Attic floor area (sq ft, for ventilation)", 1500, "0"),
        ("Waste factor (%)", 10, "0"),
    ]
    for label, default, fmt in mat_rows:
        c = label_value(ws, r, label, default, fmt=fmt)
        c.fill = YELLOW
        r += 1
    wb.defined_names["TearOff"]      = DefinedName("TearOff",      attr_text=f"INPUTS!$B${mat_start}")
    wb.defined_names["Underlayment"] = DefinedName("Underlayment", attr_text=f"INPUTS!$B${mat_start+1}")
    wb.defined_names["Shingle"]      = DefinedName("Shingle",      attr_text=f"INPUTS!$B${mat_start+2}")
    wb.defined_names["IceWater"]     = DefinedName("IceWater",     attr_text=f"INPUTS!$B${mat_start+3}")
    wb.defined_names["AtticArea"]    = DefinedName("AtticArea",    attr_text=f"INPUTS!$B${mat_start+4}")
    wb.defined_names["WastePct"]     = DefinedName("WastePct",     attr_text=f"INPUTS!$B${mat_start+5}")

    # Helper notes
    r += 1
    n = ws.cell(r, 1, "Underlayment options: Synthetic | 15# Felt | 30# Felt")
    n.font = SMALL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Shingle options: 3-Tab | Architectural | Designer")
    n.font = SMALL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Pitch values 2..18 (rise per 12-inch run). Steeper than 9/12 = pitch surcharge typical.")
    n.font = SMALL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1

    return ws


def sheet_pitch(wb):
    ws = wb.create_sheet("PITCH")
    auto_width(ws, [10, 14, 18, 14])
    banner(ws, "Pitch & Slope Reference", "Multiply roof footprint area by slope multiplier to get true roof area", width=4)

    headers = ["Pitch", "Angle (°)", "Slope Multiplier", "Grade (%)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    ws.row_dimensions[4].height = 22

    for i, pitch in enumerate(range(2, 19)):
        r = 5 + i
        ws.cell(r, 1, f"{pitch}/12").font = MONO_BOLD
        ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 1).border = BOX
        # Angle in degrees
        ws.cell(r, 2, f"=DEGREES(ATAN({pitch}/12))")
        ws.cell(r, 2).number_format = "0.00"
        ws.cell(r, 2).font = MONO; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        # Slope multiplier = sqrt(rise^2 + 144) / 12
        ws.cell(r, 3, f"=SQRT({pitch}^2+144)/12")
        ws.cell(r, 3).number_format = "0.0000"
        ws.cell(r, 3).font = MONO_BOLD; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        # Grade %
        ws.cell(r, 4, f"={pitch}/12*100")
        ws.cell(r, 4).number_format = "0.0"
        ws.cell(r, 4).font = MONO; ws.cell(r, 4).alignment = CENTER; ws.cell(r, 4).border = BOX

    # Named range so other sheets can VLOOKUP slope multiplier from pitch
    last_row = 5 + (19 - 2) - 1
    wb.defined_names["PitchTable"] = DefinedName(
        "PitchTable",
        attr_text=f"PITCH!$A$5:$D${last_row}"
    )

    r = last_row + 2
    note = ws.cell(r, 1,
        "Slope multiplier = √(rise² + 144) ÷ 12. Multiply your footprint area "
        "by this number to get actual roof surface area. Example: 50×30 = 1,500 sq ft "
        "footprint × 1.118 (6/12 pitch) = 1,677 sq ft of roof.")
    note.font = SMALL
    note.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=4)
    return ws


def sheet_takeoff(wb):
    ws = wb.create_sheet("TAKEOFF")
    auto_width(ws, [38, 14, 14, 36])
    banner(ws, "Material Takeoff", "Pulled from INPUTS — bring this list to the supply house", width=4)

    # Geometry block
    r = 4
    section(ws, r, "ROOF GEOMETRY (calculated)", span=4); r += 1
    label_value(ws, r, "Footprint area (sq ft)", "=LengthFt*WidthFt", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Slope multiplier", "=SQRT(Pitch^2+144)/12", fmt="0.0000").font = MONO_BOLD; r += 1
    label_value(ws, r, "Adjusted roof area (sq ft)", "=LengthFt*WidthFt*(SQRT(Pitch^2+144)/12)", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "With overhangs (sq ft)", "=(LengthFt+2*Overhang)*(WidthFt+2*Overhang)*(SQRT(Pitch^2+144)/12)", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Squares (1 sq = 100 sq ft)", "=((LengthFt+2*Overhang)*(WidthFt+2*Overhang)*(SQRT(Pitch^2+144)/12))/100", fmt="0.00").font = MONO_BOLD; r += 1
    sq_row = r - 1  # row containing squares formula

    # We'll reference squares as the cell's formula (Excel will reuse the formula cell)
    SQ = f"$B${sq_row}"
    WASTE = "(1+WastePct/100)"

    r += 1
    section(ws, r, "ROOFING MATERIALS (with waste factor)", span=4); r += 1
    # Header row
    headers = ["Material", "Quantity", "Unit", "Notes / Coverage Assumption"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Bundles per square: 3-Tab=3, Architectural=3, Designer=5
    bundles_per_sq_inner = 'IF(Shingle="Designer",5,IF(Shingle="3-Tab",3,3))'
    items = [
        ("Shingle bundles",
         f"=ROUNDUP({SQ}*{bundles_per_sq_inner}*{WASTE},0)",
         "bundles",
         "3 bundles/sq for 3-Tab & Architectural; 5 for Designer"),
        ("Roofing nails",
         f"=ROUNDUP({SQ}*2.5*{WASTE},0)",
         "lb",
         "≈2.5 lb per square (4 nails per shingle, 80 shingles/sq)"),
        ("Underlayment rolls",
         f'=ROUNDUP({SQ}/IF(Underlayment="Synthetic",10,IF(Underlayment="30# Felt",2,4))*{WASTE},0)',
         "rolls",
         "Synthetic=10 sq/roll · 15# Felt=4 sq/roll · 30# Felt=2 sq/roll"),
        ("Drip edge (eave + rake)",
         "=ROUNDUP((Eave+Rake)/10,0)",
         "10' pieces",
         "Standard drip edge sold in 10-ft sticks"),
        ("Starter strip",
         "=ROUNDUP(Eave/120,0)",
         "bundles",
         "≈120 lin ft per bundle (covers eave only)"),
        ("Ridge cap shingles",
         "=ROUNDUP((Ridge+Hip)/35,0)",
         "bundles",
         "Pre-cut ridge cap covers ≈35 lin ft per bundle"),
        ("Ice & water shield",
         "=ROUNDUP(IceWater/65,0)",
         "rolls",
         "1 roll = 200 sq ft ≈ 65 lin ft × 36\" wide"),
        ("Valley flashing (W-valley)",
         "=ROUNDUP(Valley/10,0)",
         "10' pieces",
         "Open metal valley; closed-cut uses ice & water shield instead"),
        ("Step flashing",
         "=ROUNDUP(StepFlash*1.2,0)",
         "pieces",
         "≈1.2 pieces per linear foot of wall flashing run"),
        ("Pipe boots (rubber)",
         "=PipeBoots",
         "boots",
         "Sized to vent pipe diameter (1.5\" or 3\")"),
        ("Roofing cement (1-gal)",
         f'=ROUNDUP({SQ}/20,0)',
         "gallons",
         "Approx 1 gal per 20 squares for sealing & flashing"),
    ]

    for label, formula, unit, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).alignment = LEFT; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "TEAR-OFF & DISPOSAL", span=4); r += 1
    label_value(ws, r, "Old shingle weight (lb)",
        f'=ROUND({SQ}*100*IF(TearOff>=2,4.5,2.5)*TearOff,0)',
        fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Dumpster size (recommended)",
        f'=IF({SQ}<15,"10 yd",IF({SQ}<25,"20 yd","30 yd"))',
        fmt=None).font = MONO_BOLD; r += 1

    # Print setup
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    return ws


def sheet_ventilation(wb):
    ws = wb.create_sheet("VENTILATION")
    auto_width(ws, [38, 18, 32])
    banner(ws, "Ventilation Balance", "1:150 NFA rule — 50/50 intake/exhaust split", width=3)

    r = 4
    section(ws, r, "REQUIRED NET FREE AREA (NFA)", span=3); r += 1
    label_value(ws, r, "Attic floor area (sq ft)", "=AtticArea", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Total NFA needed (sq in)", "=AtticArea*144/150", fmt="#,##0").font = MONO_BOLD
    nfa_row = r; r += 1
    label_value(ws, r, "Intake required (50%, sq in)", f"=$B${nfa_row}/2", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Exhaust required (50%, sq in)", f"=$B${nfa_row}/2", fmt="#,##0").font = MONO_BOLD
    exh_row = r; r += 1

    r += 1
    section(ws, r, "RIDGE VENT (12 sq in NFA per linear ft)", span=3); r += 1
    label_value(ws, r, "Ridge vent linear ft needed",
        f"=ROUNDUP($B${exh_row}/12,0)", fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Ridge available (lin ft)", "=Ridge", fmt="0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Ridge vent OK?",
        f'=IF(Ridge>=ROUNDUP($B${exh_row}/12,0),"YES — use full ridge","NO — supplement with box vents")',
        fmt=None).font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "SOFFIT INTAKE (varies by vent type)", span=3); r += 1
    label_value(ws, r, "8\" continuous soffit vent NFA",
        "=Eave*9", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Soffit OK at 8\" continuous?",
        f'=IF(Eave*9>=$B${exh_row},"YES","NO — add gable or roof intake")',
        fmt=None).font = MONO_BOLD; r += 1
    label_value(ws, r, "Round soffit vents (4\" @ 28 sq in each)",
        f"=ROUNDUP($B${exh_row}/28,0)", fmt="0").font = MONO_BOLD; r += 1

    r += 2
    note = ws.cell(r, 1,
        "Code default: 1 sq ft of NFA per 150 sq ft of attic. The 1:300 rule "
        "applies only with vapor retarder + balanced 50/50 intake/exhaust. "
        "Always confirm specific product NFA on the manufacturer label — "
        "ridge vent ranges from 9 to 18 sq in/ft depending on brand.")
    note.font = SMALL; note.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=3)
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [34, 14, 14])
    banner(ws, "Pricing", "Edit unit prices to match your supplier — yellow cells", width=3)

    r = 4
    section(ws, r, "MATERIAL UNIT COSTS (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    items = [
        ("Shingle bundle (Architectural)", 38.00, "bundle"),
        ("Shingle bundle (3-Tab)", 30.00, "bundle"),
        ("Shingle bundle (Designer)", 65.00, "bundle"),
        ("Roofing nails (per lb)", 3.50, "lb"),
        ("Underlayment roll (Synthetic)", 110.00, "roll"),
        ("Underlayment roll (15# Felt)", 28.00, "roll"),
        ("Underlayment roll (30# Felt)", 38.00, "roll"),
        ("Drip edge (10-ft piece)", 9.50, "piece"),
        ("Starter strip bundle", 42.00, "bundle"),
        ("Ridge cap bundle", 55.00, "bundle"),
        ("Ice & water shield roll", 95.00, "roll"),
        ("Valley flashing (10-ft)", 22.00, "piece"),
        ("Step flashing piece", 0.85, "piece"),
        ("Pipe boot", 14.00, "boot"),
        ("Roofing cement (1-gal)", 28.00, "gallon"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1

    # Named ranges so QUOTE can VLOOKUP material → price
    wb.defined_names["PriceTable"] = DefinedName(
        "PriceTable",
        attr_text=f"PRICING!$A${pricing_start}:$C${r-1}"
    )

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    inputs = [
        ("Labor rate per square (install)", 75.00, "per sq"),
        ("Tear-off labor per square per layer", 35.00, "per sq/layer"),
        ("Dumpster + disposal (flat)", 450.00, "flat"),
        ("Permit fee (flat)", 175.00, "flat"),
        ("Profit margin (%)", 0.25, "decimal — 0.25 = 25%"),
        ("Sales tax rate (%)", 0.07, "decimal — 0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in inputs:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LaborPerSq"]   = DefinedName("LaborPerSq",   attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["TearOffPerSq"] = DefinedName("TearOffPerSq", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["Dumpster"]     = DefinedName("Dumpster",     attr_text=f"PRICING!$B${labor_start+2}")
    wb.defined_names["Permit"]       = DefinedName("Permit",       attr_text=f"PRICING!$B${labor_start+3}")
    wb.defined_names["Margin"]       = DefinedName("Margin",       attr_text=f"PRICING!$B${labor_start+4}")
    wb.defined_names["TaxRate"]      = DefinedName("TaxRate",      attr_text=f"PRICING!$B${labor_start+5}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [42, 14, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review unit prices before sending", width=4)

    r = 4
    # Customer block
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Project", "=INPUTS!B5"); r += 1

    r += 1
    section(ws, r, "MATERIALS", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    mat_start_quote = r

    # Each row: Description / qty (from TAKEOFF) / unit price (VLOOKUP into PriceTable) / total
    # TAKEOFF material lines occupy rows 13..23.
    quote_items = [
        ("Architectural shingles (bundle)", 13, "Shingle bundle (Architectural)"),
        ("Roofing nails (lb)",              14, "Roofing nails (per lb)"),
        ("Synthetic underlayment (roll)",   15, "Underlayment roll (Synthetic)"),
        ("Drip edge (10' piece)",           16, "Drip edge (10-ft piece)"),
        ("Starter strip (bundle)",          17, "Starter strip bundle"),
        ("Ridge cap (bundle)",              18, "Ridge cap bundle"),
        ("Ice & water shield (roll)",       19, "Ice & water shield roll"),
        ("Valley flashing (10' piece)",     20, "Valley flashing (10-ft)"),
        ("Step flashing (piece)",           21, "Step flashing piece"),
        ("Pipe boots",                      22, "Pipe boot"),
        ("Roofing cement (1-gal)",          23, "Roofing cement (1-gal)"),
    ]
    for desc, takeoff_row, price_label in quote_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        # qty
        c = ws.cell(r, 2, f"=TAKEOFF!$B${takeoff_row}"); c.font = MONO; c.alignment = RIGHT
        c.number_format = "#,##0"; c.border = BOX
        # unit price via VLOOKUP
        p = ws.cell(r, 3, f'=VLOOKUP("{price_label}",PriceTable,2,FALSE)')
        p.font = MONO; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.border = BOX
        # total
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    mat_end_quote = r - 1

    # Material subtotal
    ws.cell(r, 1, "Material subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{mat_start_quote}:D{mat_end_quote})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    mat_subtotal_row = r; r += 1

    r += 1
    section(ws, r, "LABOR & FEES", span=4); r += 1
    # Squares cell from takeoff = $B$9 (row 9 in TAKEOFF)
    SQ_REF = "TAKEOFF!$B$9"

    labor_lines = [
        ("Tear-off labor",        f"=({SQ_REF})*TearOffPerSq*TearOff"),
        ("Install labor",         f"=({SQ_REF})*LaborPerSq"),
        ("Dumpster + disposal",   "=Dumpster"),
        ("Permit fee",            "=Permit"),
    ]
    labor_start_quote = r
    for desc, formula in labor_lines:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        for col in (2, 3): ws.cell(r, col, "").border = BOX
        t = ws.cell(r, 4, formula); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    labor_end_quote = r - 1

    ws.cell(r, 1, "Labor subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{labor_start_quote}:D{labor_end_quote})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    labor_subtotal_row = r; r += 1

    r += 1
    # Margin + tax
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{mat_subtotal_row}+D{labor_subtotal_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    # GRAND TOTAL
    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    # Signature block
    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Material prices subject to change at supplier. "
        "Work scheduled within 2-4 weeks of signed acceptance, weather permitting.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=4)

    # Print setup
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_pitch(wb)
    sheet_pricing(wb)   # build before TAKEOFF? No — TAKEOFF doesn't reference pricing; QUOTE does.
    sheet_takeoff(wb)
    sheet_ventilation(wb)
    sheet_quote(wb)

    # Reorder so user-facing flow makes sense
    order = ["README", "INPUTS", "TAKEOFF", "VENTILATION", "PRICING", "QUOTE", "PITCH"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Roofing-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
