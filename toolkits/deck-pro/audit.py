"""Audit the Deck Pro Toolkit workbook."""

import math
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"
XLSX = Path(__file__).parent / "dist" / "Deck-Pro-Toolkit-v1.xlsx"
EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def recalc_via_libreoffice(src):
    tmp = Path(tempfile.mkdtemp(prefix="deck_audit_"))
    cmd = [SOFFICE, "--headless", "--calc", "--norestore",
           "--convert-to", "xlsx", "--outdir", str(tmp), str(src)]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if res.returncode != 0:
        raise RuntimeError(f"soffice convert failed: {res.stderr or res.stdout}")
    out = tmp / src.name
    if not out.exists():
        raise RuntimeError(f"soffice did not produce {out}")
    return out


def scan_for_errors(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    found = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(e in v for e in EXCEL_ERRORS):
                    found.append((sheet, cell.coordinate, v))
    wb.close()
    return found


def expected_results():
    # INPUTS defaults
    deck_len = 16.0
    deck_width = 12.0
    joist_space = 16
    beam_count = 3
    post_count = 6
    foot_dia = 12
    foot_depth = 48
    post_height = 3.0
    rail_lf = 36
    stair_count = 4
    stair_width = 4.0
    decking_mat = "Composite"
    board_width = 5.5
    board_len = 16
    deck_waste = 0.15

    # STRUCTURE
    joist_count = math.ceil(deck_width * 12 / joist_space) + 1   # 10
    joist_len = deck_len + 1                                      # 17
    joist_lf = joist_count * joist_len                            # 170
    joist_hangers = joist_count                                   # 10
    beam_lf = (beam_count - 1) * deck_len                         # 32
    post_lf = post_count * (post_height + 1)                      # 24
    post_bases = post_count                                       # 6
    footing_count = post_count                                    # 6
    conc_each = math.pi * (foot_dia / 24) ** 2 * (foot_depth / 12) / 27
    conc_total = footing_count * conc_each * 1.10

    # DECKING
    deck_area = deck_len * deck_width                             # 192
    board_count = math.ceil(deck_area * (1 + deck_waste) / (board_width / 12 * board_len))
    deck_screws = 2 * board_count * joist_count
    rail_lf2 = rail_lf * 2                                        # 72
    balusters = math.ceil(rail_lf * 12 / 4)                       # 108
    rail_posts = math.ceil(rail_lf / 6) + 2                       # 8
    stair_stringers = (math.ceil(stair_width / 1.5) + 1) if stair_count > 0 else 0
    tread_lf = stair_count * stair_width * 2                      # 32

    # LABOR
    foot_h = footing_count * 2.5                                  # 15
    frame_h = (joist_lf + beam_lf) * 0.04                         # 8.08
    deck_h = board_count * 0.20
    rail_h = rail_lf * 0.40                                       # 14.4
    stair_h = stair_count * 1.0                                   # 4
    setup_h = 8.0
    total_hours = foot_h + frame_h + deck_h + rail_h + stair_h + setup_h

    # PRICING
    price_deck_pt = 2.50
    price_deck_comp = 4.50
    price_deck_hw = 12.00
    price_joist = 3.20
    price_beam = 3.20
    price_post = 6.50
    price_rail = 1.40
    price_balu = 5.00
    price_hanger = 3.50
    price_base = 18.00
    price_screw_100 = 4.00
    price_conc_cy = 220.00
    price_stringer = 45.00
    price_misc = 85.00
    labor_rate = 65.00
    margin = 0.25
    tax_rate = 0.07

    # decking LF total for pricing
    decking_lf = board_count * board_len
    deck_unit = price_deck_comp  # decking_mat = Composite

    materials = (
        decking_lf * deck_unit +
        joist_lf * price_joist +
        beam_lf * price_beam +
        post_lf * price_post +
        rail_lf2 * price_rail +
        balusters * price_balu +
        joist_hangers * price_hanger +
        post_bases * price_base +
        math.ceil(deck_screws / 100) * price_screw_100 +
        math.ceil(conc_total) * price_conc_cy +
        stair_stringers * price_stringer +
        1 * price_misc
    )
    labor_cost = total_hours * labor_rate
    cost_sub = materials + labor_cost
    margin_amt = cost_sub * margin
    pretax = cost_sub + margin_amt
    tax = pretax * tax_rate
    total_due = pretax + tax

    return {
        "JoistCount": joist_count,
        "JoistLF": joist_lf,
        "JoistHangers": joist_hangers,
        "BeamLF": beam_lf,
        "PostLF": post_lf,
        "PostBases": post_bases,
        "FootingCount": footing_count,
        "ConcreteCY": conc_total,
        "DeckArea": deck_area,
        "BoardCount": board_count,
        "DeckScrews": deck_screws,
        "Balusters": balusters,
        "TotalHours": total_hours,
        "Materials": materials,
        "Labor": labor_cost,
        "CostSub": cost_sub,
        "Margin": margin_amt,
        "Pretax": pretax,
        "Tax": tax,
        "TotalDue": total_due,
    }


def approx(a, b, tol_pct=0.005, tol_abs=0.01):
    if a is None or b is None:
        return False
    if a == b:
        return True
    if abs(a - b) <= tol_abs:
        return True
    denom = max(abs(a), abs(b))
    return denom > 0 and abs(a - b) / denom <= tol_pct


def main():
    if not XLSX.exists():
        print(f"FAIL: {XLSX} does not exist — run build.py first.")
        sys.exit(1)

    print(f"Source: {XLSX}")
    print("Recalculating via LibreOffice headless...")
    recalced = recalc_via_libreoffice(XLSX)
    print(f"  -> {recalced}")

    errors = scan_for_errors(recalced)
    if errors:
        print(f"\nFAIL: {len(errors)} Excel error cell(s):")
        for sheet, coord, val in errors[:20]:
            print(f"  {sheet}!{coord} = {val}")
        sys.exit(1)
    print("OK: no #REF!/#DIV/0!/#VALUE!/#NAME?/#NULL!/#NUM!/#N/A in any cell")

    wb = load_workbook(recalced, data_only=True)
    wb_f = load_workbook(XLSX, data_only=False)
    exp = expected_results()

    def resolve(name):
        ref = wb_f.defined_names[name].value
        sheet, coord = ref.split("!")
        coord = coord.replace("$", "")
        return wb[sheet][coord].value

    quote = wb["QUOTE"]
    quote_vals = {}
    for row in quote.iter_rows():
        label_cell = row[0]
        total_cell = row[3] if len(row) >= 4 else None
        if isinstance(label_cell.value, str) and total_cell is not None:
            quote_vals[label_cell.value.strip()] = total_cell.value

    checks = [
        ("JoistCount",   resolve("JoistCount"),   exp["JoistCount"]),
        ("JoistLF",      resolve("JoistLF"),      exp["JoistLF"]),
        ("JoistHangers", resolve("JoistHangers"), exp["JoistHangers"]),
        ("BeamLF",       resolve("BeamLF"),       exp["BeamLF"]),
        ("PostLF",       resolve("PostLF"),       exp["PostLF"]),
        ("PostBases",    resolve("PostBases"),    exp["PostBases"]),
        ("FootingCount", resolve("FootingCount"), exp["FootingCount"]),
        ("ConcreteCY",   resolve("ConcreteCY"),   exp["ConcreteCY"]),
        ("DeckArea",     resolve("DeckArea"),     exp["DeckArea"]),
        ("BoardCount",   resolve("BoardCount"),   exp["BoardCount"]),
        ("DeckScrews",   resolve("DeckScrews"),   exp["DeckScrews"]),
        ("Balusters",    resolve("Balusters"),    exp["Balusters"]),
        ("TotalHours",   resolve("TotalHours"),   exp["TotalHours"]),
        ("Materials sub", quote_vals.get("Materials subtotal"),               exp["Materials"]),
        ("Cost sub",      quote_vals.get("Cost subtotal (materials + labor)"), exp["CostSub"]),
        ("Profit margin", quote_vals.get("Profit margin"),                     exp["Margin"]),
        ("Pre-tax total", quote_vals.get("Pre-tax total"),                     exp["Pretax"]),
        ("Sales tax",     quote_vals.get("Sales tax"),                         exp["Tax"]),
        ("TOTAL DUE",     quote_vals.get("TOTAL DUE"),                         exp["TotalDue"]),
    ]

    print("\nNumeric checks (workbook vs. pure-Python expected):")
    print(f"  {'Label':<22} {'Workbook':>14} {'Expected':>14}   Result")
    fails = 0
    for label, got, expv in checks:
        try:
            got_n = float(got) if got is not None else None
            exp_n = float(expv)
        except (TypeError, ValueError):
            print(f"  {label:<22} {'(not numeric)':>14} {'':>14}   FAIL")
            fails += 1
            continue
        ok = approx(got_n, exp_n)
        flag = "OK  " if ok else "FAIL"
        print(f"  {label:<22} {got_n:>14,.2f} {exp_n:>14,.2f}   {flag}")
        if not ok:
            fails += 1

    wb.close()
    wb_f.close()
    shutil.rmtree(recalced.parent, ignore_errors=True)

    if fails:
        print(f"\nFAIL: {fails} of {len(checks)} numeric checks failed")
        sys.exit(1)
    print(f"\nPASS: {len(checks)} checks, 0 Excel errors, {XLSX.name} is clean")


if __name__ == "__main__":
    main()
