"""Build the Deck Pro Toolkit workbook.

Run: python build.py
Output: dist/Deck-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)
WARN = Font(name="Arial", size=9, color="FFB30000", italic=True, bold=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Deck Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a deck's footprint and material "
         "choices into a full takeoff: joist count, beam material, post + "
         "footing count, decking boards, railing balusters, fasteners, "
         "labor hours, and a print-ready customer quote."),
        ("ESTIMATE ONLY — STRUCTURAL",
         "This toolkit COUNTS materials and labor from the geometry and "
         "framing inputs you give it. It does NOT size joists, beams, "
         "posts, or footings — those decisions belong to IRC 2018 / DCA-6 "
         "span tables (or a structural engineer) for your species, grade, "
         "exposure, snow load, and tributary widths. Verify all framing "
         "sizes before ordering material or pulling permits."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs — fill in the "
         "deck footprint, joist spacing, the framing sizes you've chosen, "
         "and the decking / railing materials.\n"
         "2. STRUCTURE returns joist count + length, beam material, post + "
         "footing count and concrete volume.\n"
         "3. DECKING returns boards needed (waste-adjusted), balusters, "
         "rail caps, and fasteners.\n"
         "4. LABOR returns hours by major task.\n"
         "5. QUOTE is print-ready — edit prices on PRICING first."),
        ("JOIST + BOARD MATH",
         "Joist count = ROUNDUP(deck width × 12 / spacing) + 1 (joist on "
         "each end + at each spacing interior).\n"
         "Decking board count = ROUNDUP(deck length × 12 / board coverage) "
         "× ROUNDUP(deck width / board length) × (1 + waste). Board cover "
         "for nominal 5/4×6 PT or 1×6 composite is 5.5\" face × 1 ft of run."),
        ("RAILING",
         "Balusters: ROUNDUP(railing LF × 12 / 4) for 4\" max gap (IRC "
         "R312.1.3 — sphere can't pass 4\")."),
        ("PRICING",
         "Material unit costs on PRICING are 2026 ballpark — edit to your "
         "supplier. Composite vs PT is selected on INPUTS; QUOTE pulls "
         "the matching unit price automatically."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Residence Back Deck"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "DECK GEOMETRY", span=3); r += 1
    geom = [
        ("Deck length (ft, parallel to house)", 16, "0.0"),
        ("Deck width (ft, away from house)", 12, "0.0"),
        ("Joist spacing (in OC: 12 or 16)", 16, "0"),
        ("Beam count (incl. ledger as 1)", 3, "0"),
        ("Post count (total)", 6, "0"),
        ("Footing diameter (in)", 12, "0"),
        ("Footing depth (in, frost line)", 48, "0"),
        ("Post height (ft, grade to beam)", 3, "0.0"),
        ("Railing perimeter (LF)", 36, "0"),
        ("Stair count (steps)", 4, "0"),
        ("Stair width (ft)", 4, "0.0"),
    ]
    geom_start = r
    for label, default, fmt in geom:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["DeckLen"] = DefinedName("DeckLen", attr_text=f"INPUTS!$B${geom_start}")
    wb.defined_names["DeckWidth"] = DefinedName("DeckWidth", attr_text=f"INPUTS!$B${geom_start+1}")
    wb.defined_names["JoistSpace"] = DefinedName("JoistSpace", attr_text=f"INPUTS!$B${geom_start+2}")
    wb.defined_names["BeamCount"] = DefinedName("BeamCount", attr_text=f"INPUTS!$B${geom_start+3}")
    wb.defined_names["PostCount"] = DefinedName("PostCount", attr_text=f"INPUTS!$B${geom_start+4}")
    wb.defined_names["FootDia"] = DefinedName("FootDia", attr_text=f"INPUTS!$B${geom_start+5}")
    wb.defined_names["FootDepth"] = DefinedName("FootDepth", attr_text=f"INPUTS!$B${geom_start+6}")
    wb.defined_names["PostHeight"] = DefinedName("PostHeight", attr_text=f"INPUTS!$B${geom_start+7}")
    wb.defined_names["RailLF"] = DefinedName("RailLF", attr_text=f"INPUTS!$B${geom_start+8}")
    wb.defined_names["StairCount"] = DefinedName("StairCount", attr_text=f"INPUTS!$B${geom_start+9}")
    wb.defined_names["StairWidth"] = DefinedName("StairWidth", attr_text=f"INPUTS!$B${geom_start+10}")

    r += 1
    section(ws, r, "MATERIALS", span=3); r += 1
    mat = [
        ("Decking material (PT / Composite / Hardwood)", "Composite", None),
        ("Decking board face width (in)", 5.5, "0.0"),
        ("Decking board length (ft: 8/12/16/20)", 16, "0"),
        ("Decking waste factor (decimal)", 0.15, "0.0%"),
        ("Joist size (2x8 / 2x10 / 2x12)", "2x10", None),
        ("Beam size (e.g. 2-ply 2x10)", "2-ply 2x10", None),
        ("Post size (e.g. 6x6 PT)", "6x6 PT", None),
        ("Concrete strength (psi)", 3000, "0"),
    ]
    mat_start = r
    for label, default, fmt in mat:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["DeckingMat"] = DefinedName("DeckingMat", attr_text=f"INPUTS!$B${mat_start}")
    wb.defined_names["BoardWidth"] = DefinedName("BoardWidth", attr_text=f"INPUTS!$B${mat_start+1}")
    wb.defined_names["BoardLen"] = DefinedName("BoardLen", attr_text=f"INPUTS!$B${mat_start+2}")
    wb.defined_names["DeckWaste"] = DefinedName("DeckWaste", attr_text=f"INPUTS!$B${mat_start+3}")
    wb.defined_names["JoistSize"] = DefinedName("JoistSize", attr_text=f"INPUTS!$B${mat_start+4}")
    wb.defined_names["BeamSize"] = DefinedName("BeamSize", attr_text=f"INPUTS!$B${mat_start+5}")
    wb.defined_names["PostSize"] = DefinedName("PostSize", attr_text=f"INPUTS!$B${mat_start+6}")
    wb.defined_names["ConcPsi"] = DefinedName("ConcPsi", attr_text=f"INPUTS!$B${mat_start+7}")

    r += 1
    n = ws.cell(r, 1, "Decking material options: PT  |  Composite  |  Hardwood — pricing pulls from PRICING tab"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Joist size / beam size / post size: VERIFY against IRC 2018 / DCA-6 span tables for your conditions"); n.font = WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Footing depth must reach below local frost line — check your AHJ; 48\" is common in IECC zone 5"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_structure(wb):
    ws = wb.create_sheet("STRUCTURE")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Framing Material Counts", "Joists, beams, posts, footings — sizes verified by you", width=5)

    r = 4
    # Disclaimer banner
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=5)
    d = ws.cell(r, 1,
        "ESTIMATE ONLY — joist size, beam size, post spacing, and footing "
        "depth must be verified against IRC 2018 / DCA-6 span tables (or a "
        "structural engineer) for your species, grade, exposure, snow load, "
        "and tributary widths. This sheet counts material only.")
    d.font = WARN
    d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    d.fill = PatternFill("solid", fgColor="FFFFF3CC")
    d.border = BOX
    ws.row_dimensions[r].height = 32
    r += 2

    section(ws, r, "JOISTS", span=5); r += 1
    headers = ["Item", "Formula", "Quantity", "Unit", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Joist count
    ws.cell(r, 1, "Joist count").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(DeckWidth × 12 / JoistSpace) + 1")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=ROUNDUP(DeckWidth*12/JoistSpace,0)+1")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Includes both end joists").font = SMALL; ws.cell(r, 5).border = BOX
    joist_count_row = r; r += 1

    # Joist length (deck length, plus 1 ft buffer for cantilever / cut)
    ws.cell(r, 1, "Joist length (each)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "DeckLen + 1 (cut buffer)")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=DeckLen+1")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0.0"; q.border = BOX
    ws.cell(r, 4, "ft").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Round up to stock length (8/10/12/16/20)").font = SMALL; ws.cell(r, 5).border = BOX
    joist_len_row = r; r += 1

    # Joist linear feet total
    ws.cell(r, 1, "Joist total LF").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    f = ws.cell(r, 2, "count × length"); f.font = MONO; f.alignment = LEFT; f.border = BOX; f.fill = PANEL_FILL
    q = ws.cell(r, 3, f"=C{joist_count_row}*C{joist_len_row}")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX; q.fill = YELLOW
    ws.cell(r, 4, "LF").font = LBL; ws.cell(r, 4).border = BOX; ws.cell(r, 4).fill = PANEL_FILL
    ws.cell(r, 5, "Verify joist size with span tables").font = WARN; ws.cell(r, 5).border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    joist_lf_row = r; r += 1

    # Joist hangers (one per joist end, both sides if joists land on beams not just hung from ledger)
    # Assume joists hang from ledger one end + sit on beams other end → 1 hanger per joist
    ws.cell(r, 1, "Joist hangers").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "1 hanger per joist (ledger end)"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f"=C{joist_count_row}")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Simpson LUS or equivalent").font = SMALL; ws.cell(r, 5).border = BOX
    hanger_row = r; r += 1

    r += 1
    section(ws, r, "BEAMS + POSTS", span=5); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Beam LF = beam count × deck length (each beam runs deck length)
    # Subtract ledger from beam material count: ledger isn't a separate beam in lumber terms
    # So beam_count input already includes ledger; beam material = (beam_count - 1) * deck length
    ws.cell(r, 1, "Beam LF (excluding ledger)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "(BeamCount - 1) × DeckLen"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=(BeamCount-1)*DeckLen")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "LF").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Ledger is dimensional but charged with house attach").font = SMALL; ws.cell(r, 5).border = BOX
    beam_lf_row = r; r += 1

    # Posts
    ws.cell(r, 1, "Post count").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "PostCount (your input)"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=PostCount")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Verify spacing per DCA-6 max tributary").font = WARN; ws.cell(r, 5).border = BOX
    post_count_row = r; r += 1

    # Post length (above grade + below grade buffer)
    ws.cell(r, 1, "Post length (each)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "PostHeight + 1 ft (cut buffer)"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=PostHeight+1")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0.0"; q.border = BOX
    ws.cell(r, 4, "ft").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Round up to 8/10/12 stock").font = SMALL; ws.cell(r, 5).border = BOX
    post_len_row = r; r += 1

    # Post total LF
    ws.cell(r, 1, "Post total LF").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    f = ws.cell(r, 2, "count × length"); f.font = MONO; f.alignment = LEFT; f.border = BOX; f.fill = PANEL_FILL
    q = ws.cell(r, 3, f"=C{post_count_row}*C{post_len_row}")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX; q.fill = YELLOW
    ws.cell(r, 4, "LF").font = LBL; ws.cell(r, 4).border = BOX; ws.cell(r, 4).fill = PANEL_FILL
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    post_lf_row = r; r += 1

    # Post bases
    ws.cell(r, 1, "Post bases").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "1 per post"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f"=C{post_count_row}")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Simpson ABU or equivalent, standoff").font = SMALL; ws.cell(r, 5).border = BOX
    base_row = r; r += 1

    r += 1
    section(ws, r, "FOOTINGS + CONCRETE", span=5); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Footing count
    ws.cell(r, 1, "Footing count").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "1 per post"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=PostCount")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Verify diameter per local code / soil").font = WARN; ws.cell(r, 5).border = BOX
    foot_count_row = r; r += 1

    # Concrete CY per footing = π × (dia/2/12)² × (depth/12) / 27
    ws.cell(r, 1, "Concrete per footing (CY)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "π × (dia/24)² × (depth/12) / 27")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=PI()*(FootDia/24)^2*(FootDepth/12)/27")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0.00"; q.border = BOX
    ws.cell(r, 4, "CY").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Cylindrical footing volume").font = SMALL; ws.cell(r, 5).border = BOX
    conc_each_row = r; r += 1

    # Concrete total
    ws.cell(r, 1, "Concrete TOTAL (with 10% waste)").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    f = ws.cell(r, 2, "footings × per-footing × 1.10")
    f.font = MONO; f.alignment = LEFT; f.border = BOX; f.fill = PANEL_FILL
    q = ws.cell(r, 3, f"=C{foot_count_row}*C{conc_each_row}*1.10")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0.00"; q.border = BOX; q.fill = YELLOW
    ws.cell(r, 4, "CY").font = LBL; ws.cell(r, 4).border = BOX; ws.cell(r, 4).fill = PANEL_FILL
    ws.cell(r, 5, "ConcPsi mix per INPUTS").font = SMALL; ws.cell(r, 5).border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    conc_total_row = r; r += 1

    # Named ranges
    wb.defined_names["JoistCount"] = DefinedName("JoistCount", attr_text=f"STRUCTURE!$C${joist_count_row}")
    wb.defined_names["JoistLF"] = DefinedName("JoistLF", attr_text=f"STRUCTURE!$C${joist_lf_row}")
    wb.defined_names["JoistHangers"] = DefinedName("JoistHangers", attr_text=f"STRUCTURE!$C${hanger_row}")
    wb.defined_names["BeamLF"] = DefinedName("BeamLF", attr_text=f"STRUCTURE!$C${beam_lf_row}")
    wb.defined_names["PostLF"] = DefinedName("PostLF", attr_text=f"STRUCTURE!$C${post_lf_row}")
    wb.defined_names["PostBases"] = DefinedName("PostBases", attr_text=f"STRUCTURE!$C${base_row}")
    wb.defined_names["FootingCount"] = DefinedName("FootingCount", attr_text=f"STRUCTURE!$C${foot_count_row}")
    wb.defined_names["ConcreteCY"] = DefinedName("ConcreteCY", attr_text=f"STRUCTURE!$C${conc_total_row}")
    return ws


def sheet_decking(wb):
    ws = wb.create_sheet("DECKING")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Decking, Railing, Fasteners", "Boards, balusters, screws by deck footprint", width=5)

    r = 4
    section(ws, r, "DECKING BOARDS", span=5); r += 1
    headers = ["Item", "Formula", "Quantity", "Unit", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Deck area
    ws.cell(r, 1, "Deck area (sq ft)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "DeckLen × DeckWidth"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=DeckLen*DeckWidth")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "sf").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "").border = BOX
    area_row = r; r += 1

    # Decking board count
    # Each board covers (BoardWidth in / 12) ft × BoardLen ft = (BoardWidth/12) × BoardLen sq ft
    # Boards needed = ROUNDUP( deck_area × (1+waste) / coverage )
    ws.cell(r, 1, "Decking board count").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(area × (1+waste) / (BoardWidth/12 × BoardLen))")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f"=ROUNDUP(C{area_row}*(1+DeckWaste)/(BoardWidth/12*BoardLen),0)")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX; q.fill = YELLOW
    ws.cell(r, 4, "boards").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Composite or PT per INPUTS").font = SMALL; ws.cell(r, 5).border = BOX
    boards_row = r; r += 1

    # Deck screws — 2 per joist per board
    ws.cell(r, 1, "Deck screws (count)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "2 × boards × joists"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f"=2*C{boards_row}*JoistCount")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "#,##0"; q.border = BOX
    ws.cell(r, 4, "screws").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Hidden clip systems may differ").font = SMALL; ws.cell(r, 5).border = BOX
    deck_screws_row = r; r += 1

    r += 1
    section(ws, r, "RAILING", span=5); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Top + bottom rail = railing LF × 2
    ws.cell(r, 1, "Rail LF (top + bottom)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "RailLF × 2 (top + bottom rail)")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=RailLF*2")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "LF").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "2x4 PT or composite rail").font = SMALL; ws.cell(r, 5).border = BOX
    rail_lf_row = r; r += 1

    # Balusters: IRC R312.1.3 — 4" max sphere passage. Spacing 4" OC.
    ws.cell(r, 1, "Balusters (4\" OC max)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(RailLF × 12 / 4)")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=ROUNDUP(RailLF*12/4,0)")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX; q.fill = YELLOW
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "IRC 4-inch sphere rule").font = SMALL; ws.cell(r, 5).border = BOX
    balusters_row = r; r += 1

    # Post caps (1 per railing post, default 1 per 6 ft + corners)
    ws.cell(r, 1, "Railing posts").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(RailLF / 6) + 2 (corners)")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=ROUNDUP(RailLF/6,0)+2")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Max 6 ft spacing typical").font = SMALL; ws.cell(r, 5).border = BOX
    rail_posts_row = r; r += 1

    r += 1
    section(ws, r, "STAIRS", span=5); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Stringer count: 3 typical (2 outer + 1 middle for ≤4ft, +1 for wider)
    ws.cell(r, 1, "Stair stringers").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(StairWidth/1.5) + 1")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=IF(StairCount=0,0,ROUNDUP(StairWidth/1.5,0)+1)")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Max 18\" spacing per stringer").font = SMALL; ws.cell(r, 5).border = BOX
    stringer_row = r; r += 1

    # Tread boards = StairCount × StairWidth × 2 (two boards per tread typical)
    ws.cell(r, 1, "Tread board LF").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "StairCount × StairWidth × 2")
    f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=StairCount*StairWidth*2")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "LF").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Two 5/4×6 per tread").font = SMALL; ws.cell(r, 5).border = BOX
    tread_row = r; r += 1

    # Named ranges
    wb.defined_names["DeckArea"] = DefinedName("DeckArea", attr_text=f"DECKING!$C${area_row}")
    wb.defined_names["BoardCount"] = DefinedName("BoardCount", attr_text=f"DECKING!$C${boards_row}")
    wb.defined_names["DeckScrews"] = DefinedName("DeckScrews", attr_text=f"DECKING!$C${deck_screws_row}")
    wb.defined_names["RailLF2"] = DefinedName("RailLF2", attr_text=f"DECKING!$C${rail_lf_row}")
    wb.defined_names["Balusters"] = DefinedName("Balusters", attr_text=f"DECKING!$C${balusters_row}")
    wb.defined_names["RailPosts"] = DefinedName("RailPosts", attr_text=f"DECKING!$C${rail_posts_row}")
    wb.defined_names["StairStringers"] = DefinedName("StairStringers", attr_text=f"DECKING!$C${stringer_row}")
    wb.defined_names["TreadLF"] = DefinedName("TreadLF", attr_text=f"DECKING!$C${tread_row}")
    return ws


def sheet_labor(wb):
    ws = wb.create_sheet("LABOR")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Labor Hours", "Concrete + framing + decking + railing + stairs", width=5)

    r = 4
    section(ws, r, "BASE HOURS PER TASK", span=5); r += 1
    headers = ["Task", "Qty unit", "Hr per unit", "Hours", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Concrete/footings — per footing
    ws.cell(r, 1, "Footing dig + pour (per footing)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=FootingCount")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 2.5)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.0"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Hand-dig + tube + pour"); n.font = SMALL; n.border = BOX
    foot_h_row = r; r += 1

    # Framing — per LF of joists + beams
    ws.cell(r, 1, "Framing (per LF joists + beams)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=JoistLF+BeamLF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.04)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Cut, hang, fasten"); n.font = SMALL; n.border = BOX
    frame_h_row = r; r += 1

    # Decking install — per board
    ws.cell(r, 1, "Decking install (per board)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=BoardCount")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.20)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.00"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Cut, space, screw"); n.font = SMALL; n.border = BOX
    deck_h_row = r; r += 1

    # Railing — per LF of railing
    ws.cell(r, 1, "Railing install (per LF)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=RailLF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.40)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.00"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Posts, rails, balusters"); n.font = SMALL; n.border = BOX
    rail_h_row = r; r += 1

    # Stairs — per step
    ws.cell(r, 1, "Stair build (per step)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=StairCount")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 1.0)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.0"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Stringers, treads, risers"); n.font = SMALL; n.border = BOX
    stair_h_row = r; r += 1

    # Layout + cleanup
    ws.cell(r, 1, "Layout + cleanup").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, 1); sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 8.0)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.0"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Site prep, batter boards, haul"); n.font = SMALL; n.border = BOX
    setup_row = r; r += 1

    ws.cell(r, 1, "TOTAL LABOR HOURS").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{foot_h_row}:D{setup_row})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0.0"; s.fill = YELLOW; s.border = BOX
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    total_hours_row = r
    wb.defined_names["TotalHours"] = DefinedName("TotalHours", attr_text=f"LABOR!$D${total_hours_row}")
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "DECKING MATERIALS — PER LF (2026)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("Decking — PT 5/4×6 per LF",          2.50, "LF"),
        ("Decking — Composite per LF",         4.50, "LF"),
        ("Decking — Hardwood (ipe) per LF",   12.00, "LF"),
        ("Joist 2x10 PT per LF",               3.20, "LF"),
        ("Beam 2x10 PT per LF",                3.20, "LF"),
        ("Post 6x6 PT per LF",                 6.50, "LF"),
        ("Rail (2x4 PT) per LF",               1.40, "LF"),
        ("Baluster (square aluminum)",         5.00, "ea"),
        ("Joist hanger (Simpson LUS)",         3.50, "ea"),
        ("Post base (Simpson ABU)",           18.00, "ea"),
        ("Deck screws (per 100)",              4.00, "100 ct"),
        ("Concrete (per CY, delivered)",     220.00, "CY"),
        ("Stair stringer (2x12 PT, 8 ft)",   45.00, "ea"),
        ("Misc fasteners + flashing (flat)",  85.00, "job"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceDeckPT"] = DefinedName("PriceDeckPT", attr_text=f"PRICING!$B${pricing_start}")
    wb.defined_names["PriceDeckComp"] = DefinedName("PriceDeckComp", attr_text=f"PRICING!$B${pricing_start+1}")
    wb.defined_names["PriceDeckHW"] = DefinedName("PriceDeckHW", attr_text=f"PRICING!$B${pricing_start+2}")
    wb.defined_names["PriceJoist"] = DefinedName("PriceJoist", attr_text=f"PRICING!$B${pricing_start+3}")
    wb.defined_names["PriceBeam"] = DefinedName("PriceBeam", attr_text=f"PRICING!$B${pricing_start+4}")
    wb.defined_names["PricePost"] = DefinedName("PricePost", attr_text=f"PRICING!$B${pricing_start+5}")
    wb.defined_names["PriceRail"] = DefinedName("PriceRail", attr_text=f"PRICING!$B${pricing_start+6}")
    wb.defined_names["PriceBalu"] = DefinedName("PriceBalu", attr_text=f"PRICING!$B${pricing_start+7}")
    wb.defined_names["PriceHanger"] = DefinedName("PriceHanger", attr_text=f"PRICING!$B${pricing_start+8}")
    wb.defined_names["PriceBase"] = DefinedName("PriceBase", attr_text=f"PRICING!$B${pricing_start+9}")
    wb.defined_names["PriceScrew100"] = DefinedName("PriceScrew100", attr_text=f"PRICING!$B${pricing_start+10}")
    wb.defined_names["PriceConcCY"] = DefinedName("PriceConcCY", attr_text=f"PRICING!$B${pricing_start+11}")
    wb.defined_names["PriceStringer"] = DefinedName("PriceStringer", attr_text=f"PRICING!$B${pricing_start+12}")
    wb.defined_names["PriceMisc"] = DefinedName("PriceMisc", attr_text=f"PRICING!$B${pricing_start+13}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Labor rate (lead carpenter)",   65.00, "per hour"),
        ("Profit margin (decimal)",        0.25, "0.25 = 25%"),
        ("Sales tax (decimal)",            0.07, "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit or "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LaborRate"] = DefinedName("LaborRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+2}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Deck size", '=DeckLen&" × "&DeckWidth&" ft (" &DeckArea&" sf)"'); r += 1
    label_value(ws, r, "Decking material", "=DeckingMat"); r += 1
    label_value(ws, r, "Joist size", "=JoistSize"); r += 1
    label_value(ws, r, "Total labor hours", "=TotalHours", fmt="0.0"); r += 1

    r += 1
    section(ws, r, "MATERIALS", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    mat_start = r
    mat_items = [
        ("Decking boards", "=BoardCount*BoardLen",
         '=IF(DeckingMat="PT",PriceDeckPT,IF(DeckingMat="Composite",PriceDeckComp,PriceDeckHW))'),
        ("Joists", "=JoistLF", "=PriceJoist"),
        ("Beams", "=BeamLF", "=PriceBeam"),
        ("Posts", "=PostLF", "=PricePost"),
        ("Rail (top + bottom)", "=RailLF2", "=PriceRail"),
        ("Balusters", "=Balusters", "=PriceBalu"),
        ("Joist hangers", "=JoistHangers", "=PriceHanger"),
        ("Post bases", "=PostBases", "=PriceBase"),
        ("Deck screws (per 100)", "=ROUNDUP(DeckScrews/100,0)", "=PriceScrew100"),
        ("Concrete (CY, delivered)", "=ROUNDUP(ConcreteCY,0)", "=PriceConcCY"),
        ("Stair stringers", "=StairStringers", "=PriceStringer"),
        ("Misc fasteners + flashing", 1, "=PriceMisc"),
    ]
    for desc, qty, price in mat_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        p = ws.cell(r, 3, price); p.font = MONO; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    mat_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{mat_start}:D{mat_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    mat_sub_row = r; r += 1

    r += 1
    section(ws, r, "LABOR", span=4); r += 1
    ws.cell(r, 1, "Labor (total hours × rate)").font = LBL; ws.cell(r, 1).border = BOX
    c = ws.cell(r, 2, "=TotalHours"); c.font = MONO; c.alignment = RIGHT; c.number_format = "0.0"; c.border = BOX
    p = ws.cell(r, 3, "=LaborRate"); p.font = MONO; p.alignment = RIGHT
    p.number_format = '"$"#,##0.00'; p.border = BOX
    t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
    t.number_format = '"$"#,##0.00'; t.border = BOX
    labor_total_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{mat_sub_row}+D{labor_total_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3):
        c = ws.cell(r, col, ""); c.fill = YELLOW; c.border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. ESTIMATE ONLY — joist, beam, post, and "
        "footing sizes must be verified per IRC 2018 / DCA-6 / local AHJ "
        "before fabrication. Custom railing, lighting, skirting, and "
        "stairs beyond what's listed are billed separately.")
    n.font = WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_structure(wb)
    sheet_decking(wb)
    sheet_labor(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "STRUCTURE", "DECKING", "LABOR", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Deck-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
