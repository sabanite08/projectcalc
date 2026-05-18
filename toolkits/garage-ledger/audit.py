"""Audit one or all Garage Ledger variants.

Usage:  python audit.py <variant>      # audit one
        python audit.py all            # audit all 6

For each variant: LibreOffice headless recalc -> openpyxl data_only=True
-> scan every cell for Excel errors -> spot-check key DASHBOARD + ROI
formulas against pure-Python expected values (computed from the sample
expense rows shipped with the variant).
"""

import math
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from variants import VARIANTS

SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"
DIST = Path(__file__).parent / "dist"
EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def recalc_via_libreoffice(src):
    tmp = Path(tempfile.mkdtemp(prefix="garage_audit_"))
    cmd = [SOFFICE, "--headless", "--calc", "--norestore",
           "--convert-to", "xlsx", "--outdir", str(tmp), str(src)]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if res.returncode != 0:
        raise RuntimeError(f"soffice convert failed: {res.stderr or res.stdout}")
    out = tmp / src.name
    if not out.exists():
        raise RuntimeError(f"soffice did not produce {out}")
    return out


def scan_for_errors(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    found = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(e in v for e in EXCEL_ERRORS):
                    found.append((sheet, cell.coordinate, v))
    wb.close()
    return found


def expected_results(cfg):
    """Pure-Python recomputation of what the workbook should show.

    Uses the variant's sample_expenses rows (which are also what the workbook
    ships with) to compute totals.
    """
    samples = cfg.get("sample_expenses", [])
    total = 0.0
    by_cat = {}
    diy_hours = 0.0
    diy_value = 0.0
    for (date, cat, vendor, partnum, desc, qty, unit) in samples:
        line_total = float(qty) * float(unit)
        total += line_total
        by_cat[cat] = by_cat.get(cat, 0.0) + line_total
        if cat.startswith("DIY Hours"):
            diy_hours += float(qty)
            diy_value += line_total

    build_budget = float(cfg["default_vehicle"]["build_budget"])
    target_sale = float(cfg["default_vehicle"]["target_sale_value"])
    roi_target = float(cfg["default_vehicle"]["roi_target_pct"])

    budget_remaining = build_budget - total
    pct_used = total / build_budget if build_budget else 0
    true_cost = total + diy_value
    sale_to_hit_roi = true_cost * (1 + roi_target)

    return {
        "DashTotal": total,
        "DashBudget": build_budget,
        "DashBudgetRemaining": budget_remaining,
        "DashPctUsed": pct_used,
        "DashDiyHours": diy_hours,
        "DashDiyValue": diy_value,
        "DashTrueCost": true_cost,
        "RoiCash": total,
        "RoiDiy": diy_value,
        "RoiTrueCost": true_cost,
        "RoiTarget": target_sale,
        "RoiSaleToHitTarget": sale_to_hit_roi,
        "ByCat": by_cat,
    }


def approx(a, b, tol_pct=0.005, tol_abs=0.01):
    if a is None or b is None:
        return False
    if a == b:
        return True
    if abs(a - b) <= tol_abs:
        return True
    denom = max(abs(a), abs(b))
    return denom > 0 and abs(a - b) / denom <= tol_pct


def audit_variant(variant_key):
    cfg = VARIANTS[variant_key]
    src = DIST / cfg["slug"] / cfg["xlsx_basename"]
    if not src.exists():
        print(f"FAIL: {src} does not exist — run `python build.py {variant_key}` first.")
        return False

    print(f"\n=== {cfg['display_name']} ({src.relative_to(Path(__file__).parent)}) ===")
    print("Recalculating via LibreOffice headless...")
    recalced = recalc_via_libreoffice(src)

    errors = scan_for_errors(recalced)
    if errors:
        print(f"FAIL: {len(errors)} Excel error cell(s):")
        for sheet, coord, val in errors[:20]:
            print(f"  {sheet}!{coord} = {val}")
        return False
    print("OK: no Excel errors in any cell")

    wb = load_workbook(recalced, data_only=True)
    exp = expected_results(cfg)

    # Pull DASHBOARD values by inspecting the label/value pairs we know are
    # in column A/B starting at row 5.
    dash = wb["DASHBOARD"]
    dash_vals = {}
    for row in dash.iter_rows():
        label = row[0].value
        val = row[1].value if len(row) >= 2 else None
        if isinstance(label, str) and val is not None:
            dash_vals[label.strip()] = val

    roi = wb["ROI"]
    roi_vals = {}
    for row in roi.iter_rows():
        label = row[0].value
        val = row[1].value if len(row) >= 2 else None
        if isinstance(label, str) and val is not None:
            roi_vals[label.strip()] = val

    # By-category checks
    by_cat_workbook = {}
    in_section = False
    for row in dash.iter_rows():
        first = row[0].value
        if isinstance(first, str) and first.strip().startswith("Category"):
            in_section = True
            continue
        if in_section:
            cat = row[0].value
            tot = row[1].value
            if isinstance(cat, str) and isinstance(tot, (int, float)):
                by_cat_workbook[cat] = tot

    checks = [
        ("DASH total spent",   dash_vals.get("Total spent (paid + unpaid)"),     exp["DashTotal"]),
        ("DASH build budget",  dash_vals.get("Build budget"),                    exp["DashBudget"]),
        ("DASH budget remain", dash_vals.get("Budget remaining"),                exp["DashBudgetRemaining"]),
        ("DASH DIY hours",     dash_vals.get("DIY hours logged (sum of qty)"),   exp["DashDiyHours"]),
        ("DASH DIY value",     dash_vals.get("DIY labor value (hrs × shop rate)"), exp["DashDiyValue"]),
        ("DASH true cost",     dash_vals.get("Build total + DIY labor value"),   exp["DashTrueCost"]),
        ("ROI cash spent",     roi_vals.get("Cash spent (from DASHBOARD)"),      exp["RoiCash"]),
        ("ROI DIY value",      roi_vals.get("DIY labor value (hrs × shop rate)"), exp["RoiDiy"]),
        ("ROI true cost",      roi_vals.get("True cost (cash + DIY labor)"),     exp["RoiTrueCost"]),
        ("ROI sale-to-hit",    roi_vals.get("Sale price to hit ROI target"),     exp["RoiSaleToHitTarget"]),
    ]

    # Add category checks for every non-zero expected category
    for cat, expected_total in exp["ByCat"].items():
        if expected_total == 0:
            continue
        got = by_cat_workbook.get(cat)
        checks.append((f"by-cat: {cat[:30]}", got, expected_total))

    print(f"\nNumeric checks (workbook vs. pure-Python expected):")
    print(f"  {'Label':<35} {'Workbook':>14} {'Expected':>14}   Result")
    fails = 0
    for label, got, expv in checks:
        try:
            got_n = float(got) if got is not None else None
            exp_n = float(expv)
        except (TypeError, ValueError):
            print(f"  {label:<35} {'(not numeric)':>14} {'':>14}   FAIL")
            fails += 1
            continue
        ok = approx(got_n, exp_n)
        flag = "OK  " if ok else "FAIL"
        print(f"  {label:<35} {got_n:>14,.2f} {exp_n:>14,.2f}   {flag}")
        if not ok:
            fails += 1

    wb.close()
    shutil.rmtree(recalced.parent, ignore_errors=True)

    if fails:
        print(f"\nFAIL: {fails} of {len(checks)} checks failed")
        return False
    print(f"\nPASS: {len(checks)} checks, 0 Excel errors")
    return True


def main():
    if len(sys.argv) < 2:
        print("Usage: python audit.py <variant>  OR  python audit.py all")
        sys.exit(1)

    arg = sys.argv[1]
    if arg == "all":
        results = {}
        for key in VARIANTS:
            results[key] = audit_variant(key)
        print("\n\n=== SUMMARY ===")
        for key, ok in results.items():
            print(f"  {VARIANTS[key]['display_name']:<40} {'PASS' if ok else 'FAIL'}")
        if not all(results.values()):
            sys.exit(1)
    else:
        if arg not in VARIANTS:
            print(f"Unknown variant: {arg}")
            sys.exit(1)
        if not audit_variant(arg):
            sys.exit(1)


if __name__ == "__main__":
    main()
