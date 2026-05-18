"""Generate 5 Etsy listing PNGs for any Garage Ledger variant.

Usage:  python make_listing_screenshots.py <variant>
        python make_listing_screenshots.py all

Output per variant: dist/<slug>/listing/01_cover.png ... 05_roi.png

The cover at 01_cover.png is a real workbook screenshot. The bold-marketing
splash that becomes the Etsy rank-1 thumbnail is generated separately by
toolkits/make_design_covers.py and saved as 01_cover_design.png.
"""

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import pymupdf
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image, ImageDraw, ImageFont

from variants import VARIANTS

TOOLKIT_FOLDER = Path(__file__).parent
PRICE_BY_SLUG = {
    "master": "$39",
    "ls-swap": "$59",
    "cummins-swap": "$59",
    "overland": "$59",
    "restomod": "$59",
    "track-build": "$59",
}

BG = (14, 14, 12)
BG_2 = (22, 22, 19)
LINE = (42, 42, 38)
INK = (244, 241, 234)
INK_2 = (184, 179, 167)
INK_3 = (118, 114, 106)
HI_VIS = (255, 212, 0)

FN_BLACK = "C:/Windows/Fonts/ariblk.ttf"
FN_BOLD = "C:/Windows/Fonts/arialbd.ttf"
FN_REG = "C:/Windows/Fonts/arial.ttf"
FN_MONO = "C:/Windows/Fonts/consola.ttf"
FN_MONO_BOLD = "C:/Windows/Fonts/consolab.ttf"

W = H = 2000
SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"

SHEETS_TO_FEATURE = [
    ("BUILD INFO",       "BUILD INFO — vehicle + budget + targets"),
    ("EXPENSE LOG",      "EXPENSE LOG — every receipt, categorized"),
    ("PARTS INVENTORY",  "PARTS INVENTORY — parts + warranty alerts"),
    ("TIMELINE",         "TIMELINE — milestones + photo links"),
    ("DASHBOARD",        "DASHBOARD — total spent, by-category, DIY value"),
    ("ROI",              "ROI / SALE COMP — your build vs. similar sales"),
]
COVER_SHEET = "DASHBOARD"

# Per-sheet print areas — crops the long input-row sheets to the first ~20
# data rows so screenshots don't show a wall of empty yellow cells.
# Sheets not listed here render with their entire used range.
PRINT_AREAS = {
    "EXPENSE LOG":     "A1:K25",   # header rows 1-5 + 20 data rows
    "PARTS INVENTORY": "A1:I25",
    "TIMELINE":        "A1:E25",
}


def f(path, size):
    return ImageFont.truetype(path, size)


def text_centered(d, text, font, fill, cx, y):
    bbox = d.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    d.text((cx - w // 2, y), text, fill=fill, font=font)


def text_right(d, text, font, fill, x_right, y):
    bbox = d.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    d.text((x_right - w, y), text, fill=fill, font=font)


def render_sheets_to_png(xlsx_path):
    work_dir = Path(tempfile.mkdtemp(prefix="gl_screens_"))

    src = work_dir / xlsx_path.name
    shutil.copy(xlsx_path, src)
    wb = load_workbook(src)
    sheet_order = list(wb.sheetnames)
    for name in sheet_order:
        ws = wb[name]
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.orientation = "landscape"
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
        ws.print_options.horizontalCentered = True
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        # Crop the long-input sheets so screenshots don't show empty yellow rows
        ws.print_area = PRINT_AREAS.get(name, None)
    wb.save(src)

    subprocess.run([
        SOFFICE, "--headless", "--calc",
        "--convert-to", "pdf",
        "--outdir", str(work_dir),
        str(src),
    ], check=True, capture_output=True, timeout=120)

    pdf_path = work_dir / src.with_suffix(".pdf").name
    doc = pymupdf.open(pdf_path)
    out_images = {}
    zoom = 2.5
    mat = pymupdf.Matrix(zoom, zoom)
    for idx, sheet in enumerate(sheet_order):
        if idx >= len(doc):
            break
        page = doc[idx]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        out_images[sheet] = crop_trailing_whitespace(img)
    doc.close()
    shutil.rmtree(work_dir, ignore_errors=True)
    return out_images


def crop_trailing_whitespace(img):
    """Trim trailing white rows off the bottom of a rendered PDF page so the
    sheet table fills more of the final 2000x2000 listing image instead of
    leaving half the frame white."""
    import numpy as np
    arr = np.asarray(img.convert("L"))
    # A row is "content" if any pixel is darker than 245 (off-white)
    row_has_content = (arr < 245).any(axis=1)
    if not row_has_content.any():
        return img
    last_content_row = int(np.where(row_has_content)[0][-1])
    # 60px padding below the last content row so the table isn't crushed
    new_h = min(arr.shape[0], last_content_row + 60)
    if new_h >= arr.shape[0] - 8:
        return img  # already tight; no-op
    return img.crop((0, 0, img.width, new_h))


def build_sheet_image(screenshot, caption, toolkit_name, price):
    img = Image.new("RGB", (W, H), BG)
    d = ImageDraw.Draw(img)

    bar_h = 140
    d.rectangle([0, 0, W, bar_h], fill=BG_2)
    d.rectangle([0, bar_h, W, bar_h + 6], fill=HI_VIS)

    s = 26
    bx, by = 60, 56
    d.polygon([(bx, by + s // 2), (bx + s // 2, by), (bx + s, by + s // 2),
               (bx + s // 2, by + s)], fill=HI_VIS)
    d.text((bx + s + 18, by - 4), "PROJECTCALC", fill=INK, font=f(FN_BLACK, 32))

    d.text((bx + s + 320, by - 4), toolkit_name, fill=HI_VIS, font=f(FN_MONO_BOLD, 24))
    text_right(d, caption, f(FN_MONO_BOLD, 24), INK, W - 60, by - 2)

    pad_x = 80
    pad_top = 200
    pad_bottom = 100
    avail_w = W - pad_x * 2
    avail_h = H - pad_top - pad_bottom

    sw, sh = screenshot.size
    scale = min(avail_w / sw, avail_h / sh)
    new_w = int(sw * scale)
    new_h = int(sh * scale)
    scaled = screenshot.resize((new_w, new_h), Image.LANCZOS)

    sx = pad_x + (avail_w - new_w) // 2
    sy = pad_top + (avail_h - new_h) // 2

    shadow_off = 16
    d.rectangle([sx + shadow_off, sy + shadow_off, sx + new_w + shadow_off,
                 sy + new_h + shadow_off], fill=(0, 0, 0))
    d.rectangle([sx - 2, sy - 2, sx + new_w + 2, sy + new_h + 2],
                fill=(250, 250, 247), outline=LINE, width=2)
    img.paste(scaled, (sx, sy))
    d = ImageDraw.Draw(img)

    d.line([(60, H - 80), (W - 60, H - 80)], fill=LINE, width=2)
    d.text((60, H - 60), "PROJECTCALC.APP", fill=HI_VIS, font=f(FN_BLACK, 24))
    text_right(d, f"{price}  ·  INSTANT DOWNLOAD  ·  EXCEL · GOOGLE SHEETS · LIBREOFFICE",
               f(FN_MONO_BOLD, 22), INK_2, W - 60, H - 56)

    return img


def build_cover(screenshot, toolkit_name, tagline, price):
    img = Image.new("RGB", (W, H), BG)
    d = ImageDraw.Draw(img)

    step = 60
    for x in range(0, W, step):
        d.line([(x, 0), (x, H)], fill=(20, 20, 18), width=1)
    for y in range(0, H, step):
        d.line([(0, y), (W, y)], fill=(20, 20, 18), width=1)

    s = 28
    bx, by = 80, 80
    d.polygon([(bx, by + s // 2), (bx + s // 2, by), (bx + s, by + s // 2),
               (bx + s // 2, by + s)], fill=HI_VIS)
    d.text((bx + s + 18, by - 4), "PROJECTCALC", fill=INK, font=f(FN_BLACK, 32))
    text_right(d, "GARAGE LEDGER  ·  v1", f(FN_MONO_BOLD, 24), HI_VIS, W - 80, by - 2)

    title_y = 180
    d.text((80, title_y), "GARAGE", fill=INK, font=f(FN_BLACK, 130))
    d.text((80, title_y + 136), "LEDGER.", fill=HI_VIS, font=f(FN_BLACK, 110))

    tag_y = title_y + 30
    d.text((W - 720, tag_y), tagline[:60], fill=INK_2, font=f(FN_REG, 22))

    chip_w, chip_h = 280, 150
    cx_chip = W - 80 - chip_w
    cy_chip = tag_y + 60
    d.rectangle([cx_chip, cy_chip, cx_chip + chip_w, cy_chip + chip_h], fill=HI_VIS)
    text_centered(d, price, f(FN_BLACK, 92), BG, cx_chip + chip_w // 2, cy_chip + 8)
    text_centered(d, "INSTANT DOWNLOAD", f(FN_MONO_BOLD, 20), BG, cx_chip + chip_w // 2, cy_chip + 116)

    pad_x = 80
    hero_top = 640
    hero_bottom = 1820
    avail_w = W - pad_x * 2
    avail_h = hero_bottom - hero_top

    sw, sh = screenshot.size
    scale = min(avail_w / sw, avail_h / sh)
    new_w = int(sw * scale)
    new_h = int(sh * scale)
    scaled = screenshot.resize((new_w, new_h), Image.LANCZOS)

    sx = pad_x + (avail_w - new_w) // 2
    sy = hero_top + (avail_h - new_h) // 2

    text_centered(d, f"REAL WORKBOOK PREVIEW  ·  {toolkit_name}",
                  f(FN_MONO_BOLD, 24), HI_VIS, W // 2, hero_top - 50)

    d.rectangle([sx + 14, sy + 14, sx + new_w + 14, sy + new_h + 14], fill=(0, 0, 0))
    d.rectangle([sx - 2, sy - 2, sx + new_w + 2, sy + new_h + 2],
                fill=(250, 250, 247), outline=LINE, width=2)
    img.paste(scaled, (sx, sy))
    d = ImageDraw.Draw(img)

    y = 1880
    d.line([(80, y), (W - 80, y)], fill=LINE, width=2)
    d.text((80, y + 24), "PROJECTCALC.APP", fill=HI_VIS, font=f(FN_BLACK, 26))
    text_right(d, "EXCEL  ·  GOOGLE SHEETS  ·  LIBREOFFICE",
               f(FN_MONO_BOLD, 22), INK_3, W - 80, y + 28)

    return img


def make_for_variant(variant_key):
    cfg = VARIANTS[variant_key]
    xlsx = TOOLKIT_FOLDER / "dist" / cfg["slug"] / cfg["xlsx_basename"]
    out = TOOLKIT_FOLDER / "dist" / cfg["slug"] / "listing"
    out.mkdir(parents=True, exist_ok=True)
    price = PRICE_BY_SLUG[cfg["slug"]]
    toolkit_name = cfg["display_name"].upper()

    print(f"\n[{cfg['slug']}] Rendering sheets from {xlsx.name}")
    images = render_sheets_to_png(xlsx)

    cover_img = build_cover(images[COVER_SHEET], toolkit_name, cfg["tagline"], price)
    cover_img.save(out / "01_cover.png", "PNG", optimize=True)
    print(f"  01_cover.png  (using {COVER_SHEET})")

    for i, (sheet, caption) in enumerate(SHEETS_TO_FEATURE, start=2):
        if sheet not in images:
            print(f"  WARN: sheet {sheet!r} not found, skipping")
            continue
        slug = sheet.lower().replace(" ", "_").replace("-", "_")
        out_name = f"0{i}_{slug}.png"
        img = build_sheet_image(images[sheet], caption, toolkit_name, price)
        img.save(out / out_name, "PNG", optimize=True)
        print(f"  {out_name}  ({sheet})")


def main():
    if len(sys.argv) < 2:
        print("Usage: python make_listing_screenshots.py <variant>|all")
        sys.exit(1)
    arg = sys.argv[1]
    if arg == "all":
        for key in VARIANTS:
            make_for_variant(key)
    else:
        if arg not in VARIANTS:
            print(f"Unknown variant: {arg}")
            sys.exit(1)
        make_for_variant(arg)


if __name__ == "__main__":
    main()
