"""Build a Garage Ledger workbook for one variant.

Usage:  python build.py <variant>
        variant in: master ls_swap cummins_swap overland restomod track_build

Output: dist/<variant>/<Garage-Ledger-...-v1>.xlsx
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation

from variants import VARIANTS

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"
WARN_BG = "FFFFF3CC"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)
WARN_FILL = PatternFill("solid", fgColor=WARN_BG)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Constants for cross-sheet referencing
EXPENSE_DATA_START = 6
EXPENSE_DATA_END = 205  # 200 expense rows
PARTS_DATA_START = 6
PARTS_DATA_END = 105    # 100 parts rows
TIMELINE_DATA_START = 6
TIMELINE_DATA_END = 55  # 50 timeline rows


def sheet_readme(wb, cfg):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, cfg["display_name"], f"v1 · 2026 · projectcalc.app — {cfg['tagline']}", width=2)

    rows = [
        ("WHAT'S INSIDE",
         f"Seven linked tabs that turn every receipt, part, and shop hour of "
         f"your {cfg['default_vehicle']['build_type'].lower()} into a single "
         "dashboard: total spent, spend-by-category, parts inventory with "
         "warranty alerts, milestone timeline, and a sale-comp / ROI sheet."),
        ("HOW TO USE",
         "1. Fill BUILD INFO once — vehicle, budget, target sale, your shop rate.\n"
         "2. Log every receipt on EXPENSE LOG. Pick a category from the "
         "dropdown so DASHBOARD aggregates correctly.\n"
         "3. Add parts to PARTS INVENTORY as they ship — warranty expiration "
         "auto-calculates from the install date.\n"
         "4. Drop milestone dates + photo links on TIMELINE.\n"
         "5. DASHBOARD updates live as you log. ROI / SALE COMP shows what "
         "the build is worth vs. similar builds you research."),
        ("DASHBOARD CATEGORIES",
         "DASHBOARD's pivot is built from the categories on EXPENSE LOG's "
         "dropdown. Add/remove categories by editing the data-validation "
         "list (Data → Validation in Excel, Tools → Validity in LibreOffice). "
         "DASHBOARD reads from the EXPENSE LOG with SUMIFS so it picks up "
         "anything you log there."),
        ("DIY HOURS",
         "Log shop hours as expenses: pick category \"DIY Hours (logged)\", "
         "qty = hours, unit cost = your shop rate (from BUILD INFO). "
         "DASHBOARD's \"DIY labor value\" totals those rows so you can see "
         "the dollar value of the time you've put in — important for the "
         "ROI sheet."),
        ("WARRANTIES",
         "On PARTS INVENTORY, enter the warranty period (months) and the "
         "install date. The warranty-expires column computes automatically. "
         "DASHBOARD flags parts expiring within the next 60 days."),
        ("SALE COMP / ROI",
         "On ROI / SALE COMP, enter up to 5 comparable builds you find "
         "(Bring a Trailer, Cars & Bids, Facebook groups, ClassicCars). "
         "The sheet computes implied profit/loss vs each comp and your "
         "break-even sale price."),
        ("LICENSE",
         "Single user. Use it on as many of your own builds as you want. "
         "Please don't resell or redistribute the file. Bug reports & "
         "feature requests: bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_build_info(wb, cfg):
    ws = wb.create_sheet("BUILD INFO")
    auto_width(ws, [38, 22, 30])
    banner(ws, "Build Info", "Yellow cells = enter your project values", width=3)

    v = cfg["default_vehicle"]
    r = 4
    section(ws, r, "VEHICLE", span=3); r += 1
    fields = [
        ("Year", v["year"], "0"),
        ("Make", v["make"], None),
        ("Model", v["model"], None),
        ("VIN", v["vin"], None),
        ("Build type", v["build_type"], None),
    ]
    veh_start = r
    for label, default, fmt in fields:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["VehYear"] = DefinedName("VehYear", attr_text=f"'BUILD INFO'!$B${veh_start}")
    wb.defined_names["VehMake"] = DefinedName("VehMake", attr_text=f"'BUILD INFO'!$B${veh_start+1}")
    wb.defined_names["VehModel"] = DefinedName("VehModel", attr_text=f"'BUILD INFO'!$B${veh_start+2}")

    r += 1
    section(ws, r, "TIMELINE", span=3); r += 1
    timeline_fields = [
        ("Start date", v["start_date"], None),
        ("Target finish date", v["target_date"], None),
    ]
    tl_start = r
    for label, default, fmt in timeline_fields:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["StartDate"] = DefinedName("StartDate", attr_text=f"'BUILD INFO'!$B${tl_start}")
    wb.defined_names["TargetDate"] = DefinedName("TargetDate", attr_text=f"'BUILD INFO'!$B${tl_start+1}")

    r += 1
    section(ws, r, "BUDGET + TARGETS", span=3); r += 1
    budget_fields = [
        ("Build budget", v["build_budget"], '"$"#,##0'),
        ("Target sale value (0 = keeper)", v["target_sale_value"], '"$"#,##0'),
        ("Your shop rate (per hour)", v["shop_rate"], '"$"#,##0'),
        ("ROI target (decimal)", v["roi_target_pct"], "0.0%"),
    ]
    bud_start = r
    for label, default, fmt in budget_fields:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["BuildBudget"] = DefinedName("BuildBudget", attr_text=f"'BUILD INFO'!$B${bud_start}")
    wb.defined_names["TargetSale"] = DefinedName("TargetSale", attr_text=f"'BUILD INFO'!$B${bud_start+1}")
    wb.defined_names["ShopRate"] = DefinedName("ShopRate", attr_text=f"'BUILD INFO'!$B${bud_start+2}")
    wb.defined_names["RoiTarget"] = DefinedName("RoiTarget", attr_text=f"'BUILD INFO'!$B${bud_start+3}")

    r += 2
    note = ws.cell(r, 1,
        "Shop rate is what you'd charge if you were billing this work to a "
        "customer ($65-95/hr is typical for skilled hobbyist / pro). DIY "
        "Hours logged on EXPENSE LOG use this rate to compute labor value "
        "on DASHBOARD and ROI / SALE COMP.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=3)
    return ws


def sheet_expense_log(wb, cfg):
    ws = wb.create_sheet("EXPENSE LOG")
    auto_width(ws, [12, 26, 22, 14, 36, 8, 12, 12, 8, 16, 30])
    banner(ws, "Expense Log", "Every receipt; pick category from dropdown so DASHBOARD updates", width=11)

    r = 4
    section(ws, r, "EXPENSES (yellow cells; total auto-computes from Qty × Unit Cost)", span=11); r += 1

    headers = ["Date", "Category", "Vendor", "Part #", "Description",
               "Qty", "Unit Cost", "Total", "Paid", "Payment Method", "Receipt Link"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    # r should now equal EXPENSE_DATA_START
    assert r == EXPENSE_DATA_START, f"EXPENSE_DATA_START mismatch: r={r}"

    # Pre-fill sample rows from variant config (always 3-4 rows in the variant)
    samples = cfg.get("sample_expenses", [])

    # Build 200 rows of formulas
    for row in range(EXPENSE_DATA_START, EXPENSE_DATA_END + 1):
        # If we have a sample for this row, fill it in; else blank yellow inputs
        sample_idx = row - EXPENSE_DATA_START
        if sample_idx < len(samples):
            date, cat, vendor, partnum, desc, qty, unit = samples[sample_idx]
            ws.cell(row, 1, date).fill = YELLOW
            ws.cell(row, 2, cat).fill = YELLOW
            ws.cell(row, 3, vendor).fill = YELLOW
            ws.cell(row, 4, partnum).fill = YELLOW
            ws.cell(row, 5, desc).fill = YELLOW
            ws.cell(row, 6, qty).fill = YELLOW
            ws.cell(row, 7, unit).fill = YELLOW
            ws.cell(row, 9, "Y").fill = YELLOW
        else:
            for col in (1, 2, 3, 4, 5, 6, 7, 9, 10, 11):
                ws.cell(row, col, None).fill = YELLOW

        # Total = qty * unit cost (or blank if either missing)
        t = ws.cell(row, 8, f'=IF(OR(F{row}="",G{row}=""),"",F{row}*G{row})')
        t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'

        # Borders + formatting on every cell in the row
        for col in range(1, 12):
            cell = ws.cell(row, col)
            cell.border = BOX
            if col == 6:
                cell.number_format = "0.00"
                cell.alignment = RIGHT
            elif col == 7:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = RIGHT
            elif col == 9:
                cell.alignment = CENTER
            else:
                if cell.font.name != "Consolas":
                    cell.font = LBL
                cell.alignment = LEFT

    # Data validation: Category dropdown
    cats_list = ",".join(f'"{c}"' for c in cfg["categories"])
    # Excel/LibreOffice prefer a named-range or comma-list. We use a comma list.
    # Quote the entire formula1 in case of long lists.
    if len(",".join(cfg["categories"])) < 250:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(cfg["categories"])}"',
            allow_blank=True,
        )
        dv.error = "Pick a category from the list (or edit the validation)"
        dv.errorTitle = "Invalid category"
        ws.add_data_validation(dv)
        dv.add(f"B{EXPENSE_DATA_START}:B{EXPENSE_DATA_END}")

    # Data validation: Paid Y/N
    dv_paid = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_paid)
    dv_paid.add(f"I{EXPENSE_DATA_START}:I{EXPENSE_DATA_END}")

    # Named range for total expenses column (for DASHBOARD + ROI)
    wb.defined_names["ExpCat"] = DefinedName(
        "ExpCat", attr_text=f"'EXPENSE LOG'!$B${EXPENSE_DATA_START}:$B${EXPENSE_DATA_END}")
    wb.defined_names["ExpTotal"] = DefinedName(
        "ExpTotal", attr_text=f"'EXPENSE LOG'!$H${EXPENSE_DATA_START}:$H${EXPENSE_DATA_END}")
    wb.defined_names["ExpDate"] = DefinedName(
        "ExpDate", attr_text=f"'EXPENSE LOG'!$A${EXPENSE_DATA_START}:$A${EXPENSE_DATA_END}")

    return ws


def sheet_parts(wb, cfg):
    ws = wb.create_sheet("PARTS INVENTORY")
    auto_width(ws, [14, 36, 22, 8, 18, 14, 12, 14, 14])
    banner(ws, "Parts Inventory", "Parts you've bought; warranty expiration auto-computes", width=9)

    r = 4
    section(ws, r, "PARTS (yellow cells; warranty-expires + status flag auto)", span=9); r += 1

    headers = ["Part #", "Description", "Vendor", "Qty", "Location",
               "Install Date", "Warranty (mo)", "Warranty Expires", "Status"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    assert r == PARTS_DATA_START

    samples = cfg.get("sample_parts", [])
    for row in range(PARTS_DATA_START, PARTS_DATA_END + 1):
        sample_idx = row - PARTS_DATA_START
        if sample_idx < len(samples):
            part_num, desc, vendor, qty, location, install_date, warranty_mo, status = samples[sample_idx]
            ws.cell(row, 1, part_num).fill = YELLOW
            ws.cell(row, 2, desc).fill = YELLOW
            ws.cell(row, 3, vendor).fill = YELLOW
            ws.cell(row, 4, qty).fill = YELLOW
            ws.cell(row, 5, location).fill = YELLOW
            ws.cell(row, 6, install_date).fill = YELLOW
            ws.cell(row, 7, warranty_mo).fill = YELLOW
            ws.cell(row, 9, status).fill = YELLOW
        else:
            # Yellow inputs
            for col in (1, 2, 3, 4, 5, 6, 7, 9):
                ws.cell(row, col, None).fill = YELLOW
        # Warranty expires = install date + warranty months (in days approx)
        we = ws.cell(row, 8,
            f'=IF(OR(F{row}="",G{row}=""),"",F{row}+G{row}*30)')
        we.font = MONO_BOLD; we.alignment = RIGHT; we.number_format = "mm/dd/yyyy"

        for col in range(1, 10):
            cell = ws.cell(row, col)
            cell.border = BOX
            if col in (4, 7):
                cell.number_format = "0"
                cell.alignment = RIGHT
            elif col == 6:
                cell.number_format = "mm/dd/yyyy"
                cell.alignment = RIGHT
            elif col == 9:
                cell.alignment = CENTER
            else:
                if cell.font.name != "Consolas":
                    cell.font = LBL
                cell.alignment = LEFT

    dv_status = DataValidation(
        type="list",
        formula1='"Installed,On Shelf,Ordered,Returned,Sold"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"I{PARTS_DATA_START}:I{PARTS_DATA_END}")

    wb.defined_names["PartsWarrantyExp"] = DefinedName(
        "PartsWarrantyExp",
        attr_text=f"'PARTS INVENTORY'!$H${PARTS_DATA_START}:$H${PARTS_DATA_END}")
    wb.defined_names["PartsStatus"] = DefinedName(
        "PartsStatus",
        attr_text=f"'PARTS INVENTORY'!$I${PARTS_DATA_START}:$I${PARTS_DATA_END}")
    return ws


def sheet_timeline(wb, cfg):
    ws = wb.create_sheet("TIMELINE")
    auto_width(ws, [14, 30, 50, 36, 30])
    banner(ws, "Timeline", "Milestone dates, phase, photo link, notes", width=5)

    r = 4
    section(ws, r, "MILESTONES (yellow cells; variant phases pre-filled, edit as needed)", span=5); r += 1

    headers = ["Date", "Phase", "Description", "Photo Link", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    assert r == TIMELINE_DATA_START

    # Pre-fill from variant's milestones
    milestones = cfg.get("milestones", [])
    for row in range(TIMELINE_DATA_START, TIMELINE_DATA_END + 1):
        mi = row - TIMELINE_DATA_START
        if mi < len(milestones):
            phase, desc = milestones[mi]
            ws.cell(row, 1, "").fill = YELLOW
            ws.cell(row, 2, phase).fill = YELLOW
            ws.cell(row, 3, desc).fill = YELLOW
            ws.cell(row, 4, "").fill = YELLOW
            ws.cell(row, 5, "").fill = YELLOW
        else:
            for col in (1, 2, 3, 4, 5):
                ws.cell(row, col, None).fill = YELLOW
        for col in range(1, 6):
            cell = ws.cell(row, col)
            cell.border = BOX
            if col == 1:
                cell.number_format = "mm/dd/yyyy"
                cell.alignment = RIGHT
            else:
                if cell.font.name != "Consolas":
                    cell.font = LBL
                cell.alignment = LEFT
    return ws


def sheet_dashboard(wb, cfg):
    ws = wb.create_sheet("DASHBOARD")
    auto_width(ws, [32, 16, 14, 14, 28])
    banner(ws, "Dashboard", "Auto-aggregated from EXPENSE LOG + PARTS INVENTORY", width=5)

    r = 4
    section(ws, r, "PROJECT TOTALS", span=5); r += 1
    label_value(ws, r, "Total spent (paid + unpaid)",
        "=SUMIF(ExpTotal,\">0\")", fmt='"$"#,##0.00').font = MONO_BOLD
    total_row = r; r += 1
    label_value(ws, r, "Build budget",
        "=BuildBudget", fmt='"$"#,##0.00').font = MONO_BOLD; r += 1
    label_value(ws, r, "Budget remaining",
        f"=BuildBudget-B{total_row}", fmt='"$"#,##0.00').font = MONO_BOLD
    rem_row = r; r += 1
    label_value(ws, r, "% of budget used",
        f"=IF(BuildBudget=0,0,B{total_row}/BuildBudget)", fmt="0.0%").font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "DIY LABOR VALUE", span=5); r += 1
    label_value(ws, r, "DIY hours logged (sum of qty)",
        f'=SUMIFS(\'EXPENSE LOG\'!F${EXPENSE_DATA_START}:F${EXPENSE_DATA_END},'
        f'\'EXPENSE LOG\'!B${EXPENSE_DATA_START}:B${EXPENSE_DATA_END},"DIY Hours*")',
        fmt="0.0").font = MONO_BOLD
    diy_hours_row = r; r += 1
    label_value(ws, r, "DIY labor value (hrs × shop rate)",
        f'=SUMIFS(ExpTotal,ExpCat,"DIY Hours*")',
        fmt='"$"#,##0.00').font = MONO_BOLD
    diy_value_row = r; r += 1
    label_value(ws, r, "Build total + DIY labor value",
        f"=B{total_row}+B{diy_value_row}", fmt='"$"#,##0.00').font = MONO_BOLD
    true_cost_row = r; r += 1

    # Named ranges for ROI sheet
    wb.defined_names["DashTotal"] = DefinedName("DashTotal", attr_text=f"DASHBOARD!$B${total_row}")
    wb.defined_names["DashDiyValue"] = DefinedName("DashDiyValue", attr_text=f"DASHBOARD!$B${diy_value_row}")
    wb.defined_names["DashTrueCost"] = DefinedName("DashTrueCost", attr_text=f"DASHBOARD!$B${true_cost_row}")

    r += 1
    section(ws, r, "SPEND BY CATEGORY (auto)", span=5); r += 1
    headers = ["Category", "Total", "% of build", "", ""]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    for cat in cfg["categories"]:
        ws.cell(r, 1, cat).font = LBL; ws.cell(r, 1).border = BOX
        # Escape category name for SUMIF — Excel handles quoted strings
        # but to be safe we replace " with "" inside SUMIF
        cat_for_formula = cat.replace('"', '""')
        f_total = ws.cell(r, 2, f'=SUMIF(ExpCat,"{cat_for_formula}",ExpTotal)')
        f_total.font = MONO_BOLD; f_total.alignment = RIGHT
        f_total.number_format = '"$"#,##0.00'; f_total.border = BOX

        f_pct = ws.cell(r, 3, f'=IF(B{total_row}=0,0,B{r}/B{total_row})')
        f_pct.font = MONO; f_pct.alignment = RIGHT
        f_pct.number_format = "0.0%"; f_pct.border = BOX
        for col in (4, 5):
            ws.cell(r, col, "").border = BOX
        r += 1

    r += 1
    section(ws, r, "PARTS UNDER WARRANTY (expiring in next 60 days)", span=5); r += 1
    label_value(ws, r, "Parts with warranty expiring soon",
        f'=SUMPRODUCT((PartsWarrantyExp<>"")*(PartsWarrantyExp<=TODAY()+60)*(PartsWarrantyExp>TODAY()))',
        fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Parts currently installed",
        f'=COUNTIF(PartsStatus,"Installed")', fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Parts on shelf (not yet installed)",
        f'=COUNTIF(PartsStatus,"On Shelf")', fmt="0").font = MONO_BOLD; r += 1

    return ws


def sheet_roi(wb, cfg):
    ws = wb.create_sheet("ROI")
    auto_width(ws, [38, 16, 16, 16, 26])
    banner(ws, "ROI / Sale Comp", "What the build is worth vs. similar builds", width=5)

    r = 4
    section(ws, r, "YOUR INVESTMENT", span=5); r += 1
    label_value(ws, r, "Cash spent (from DASHBOARD)",
        "=DashTotal", fmt='"$"#,##0.00').font = MONO_BOLD
    cash_row = r; r += 1
    label_value(ws, r, "DIY labor value (hrs × shop rate)",
        "=DashDiyValue", fmt='"$"#,##0.00').font = MONO_BOLD
    diy_row = r; r += 1
    label_value(ws, r, "True cost (cash + DIY labor)",
        f"=B{cash_row}+B{diy_row}", fmt='"$"#,##0.00').font = MONO_BOLD
    true_row = r; r += 1
    label_value(ws, r, "Your target sale value",
        "=TargetSale", fmt='"$"#,##0.00').font = MONO_BOLD
    target_row = r; r += 1
    label_value(ws, r, "Target ROI % (from BUILD INFO)",
        "=RoiTarget", fmt="0.0%").font = MONO_BOLD
    roi_tgt_row = r; r += 1
    label_value(ws, r, "Break-even sale (cash only)",
        f"=B{cash_row}", fmt='"$"#,##0.00').font = MONO_BOLD; r += 1
    label_value(ws, r, "Break-even sale (true cost incl. DIY)",
        f"=B{true_row}", fmt='"$"#,##0.00').font = MONO_BOLD; r += 1
    label_value(ws, r, "Sale price to hit ROI target",
        f"=B{true_row}*(1+RoiTarget)", fmt='"$"#,##0.00').font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "COMPARABLE BUILDS (research 5 similar sales)", span=5); r += 1
    headers = ["Source / URL", "Sale Price", "Comp vs. cash", "Comp vs. true cost", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    comp_start = r
    for i in range(5):
        ws.cell(r, 1, f"Comp {i+1}: paste URL here").fill = YELLOW
        ws.cell(r, 2, "").fill = YELLOW
        ws.cell(r, 5, "").fill = YELLOW
        # Comp - cash
        v_cash = ws.cell(r, 3, f'=IF(B{r}="","",B{r}-B{cash_row})')
        v_cash.number_format = '"$"#,##0.00'; v_cash.font = MONO_BOLD
        v_cash.alignment = RIGHT
        # Comp - true cost
        v_true = ws.cell(r, 4, f'=IF(B{r}="","",B{r}-B{true_row})')
        v_true.number_format = '"$"#,##0.00'; v_true.font = MONO_BOLD
        v_true.alignment = RIGHT
        for col in range(1, 6):
            cell = ws.cell(r, col)
            cell.border = BOX
            if col == 1 and cell.font.name != "Consolas":
                cell.font = LBL
        r += 1

    ws.cell(r, 1, "Comp average").font = LBL_BOLD; ws.cell(r, 1).fill = PANEL_FILL
    ws.cell(r, 1).border = BOX
    avg = ws.cell(r, 2, f'=IFERROR(AVERAGEIF(B{comp_start}:B{comp_start+4},">0"),0)')
    avg.font = MONO_BOLD; avg.alignment = RIGHT
    avg.number_format = '"$"#,##0.00'; avg.fill = YELLOW; avg.border = BOX
    avg_vs_cash = ws.cell(r, 3, f'=IF(B{r}=0,0,B{r}-B{cash_row})')
    avg_vs_cash.font = MONO_BOLD; avg_vs_cash.alignment = RIGHT
    avg_vs_cash.number_format = '"$"#,##0.00'; avg_vs_cash.border = BOX
    avg_vs_true = ws.cell(r, 4, f'=IF(B{r}=0,0,B{r}-B{true_row})')
    avg_vs_true.font = MONO_BOLD; avg_vs_true.alignment = RIGHT
    avg_vs_true.number_format = '"$"#,##0.00'; avg_vs_true.border = BOX
    ws.cell(r, 5, "").border = BOX
    r += 2

    note = ws.cell(r, 1,
        "Sources to research comps: Bring a Trailer (sold listings), Cars & "
        "Bids, ClassicCars.com (sold archive), Facebook build-specific groups, "
        "Hemmings sold filter. Aim for 5 sales of similar year/build-type/"
        "condition in the last 12 months. Adjust for your build's quality "
        "and modifications relative to each comp.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=5)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    return ws


def build(variant_key):
    cfg = VARIANTS[variant_key]
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb, cfg)
    sheet_build_info(wb, cfg)
    sheet_expense_log(wb, cfg)
    sheet_parts(wb, cfg)
    sheet_timeline(wb, cfg)
    sheet_dashboard(wb, cfg)
    sheet_roi(wb, cfg)

    order = ["README", "BUILD INFO", "EXPENSE LOG", "PARTS INVENTORY",
             "TIMELINE", "DASHBOARD", "ROI"]
    wb._sheets = [wb[name] for name in order]

    out_dir = Path(__file__).parent / "dist" / cfg["slug"]
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / cfg["xlsx_basename"]
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out.relative_to(Path(__file__).parent)}  ({size_kb:.1f} KB)")
    return out


def main():
    if len(sys.argv) < 2:
        print("Usage: python build.py <variant>")
        print(f"Available: {' '.join(VARIANTS.keys())}")
        sys.exit(1)
    variant_key = sys.argv[1]
    if variant_key not in VARIANTS:
        print(f"Unknown variant: {variant_key}")
        print(f"Available: {' '.join(VARIANTS.keys())}")
        sys.exit(1)
    build(variant_key)


if __name__ == "__main__":
    main()
