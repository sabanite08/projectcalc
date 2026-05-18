"""Audit the Painting Pro Toolkit workbook.

1. LibreOffice headless recalc (convert xlsx -> xlsx, which recomputes formulas).
2. openpyxl data_only=True scan for #REF!/#DIV/0!/#VALUE!/#NAME?/#NULL!/#NUM!/#N/A.
3. Numeric spot checks: recompute expected results in pure Python from the same
   default inputs the workbook ships with, and assert the workbook agrees.

Run: python audit.py
Exits 0 on PASS, 1 on FAIL.
"""

import math
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"
XLSX = Path(__file__).parent / "dist" / "Painting-Pro-Toolkit-v1.xlsx"

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def recalc_via_libreoffice(src: Path) -> Path:
    """Convert xlsx -> xlsx in a temp dir; LO recomputes formulas as a side effect."""
    tmp = Path(tempfile.mkdtemp(prefix="painting_audit_"))
    cmd = [
        SOFFICE, "--headless", "--calc", "--norestore",
        "--convert-to", "xlsx", "--outdir", str(tmp), str(src),
    ]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if res.returncode != 0:
        raise RuntimeError(f"soffice convert failed: {res.stderr or res.stdout}")
    out = tmp / src.name
    if not out.exists():
        raise RuntimeError(f"soffice did not produce {out}")
    return out


def scan_for_errors(xlsx: Path):
    wb = load_workbook(xlsx, data_only=True)
    found = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(e in v for e in EXCEL_ERRORS):
                    found.append((sheet, cell.coordinate, v))
    wb.close()
    return found


def get_value(wb, sheet, coord):
    return wb[sheet][coord].value


def expected_results():
    """Hand-compute what the workbook should produce from its default inputs."""
    # INPUTS defaults
    ceil_height = 9.0
    cov_top = 350
    cov_prime = 250
    cov_trim = 400
    wall_coats = 2
    ceil_coats = 1
    trim_coats = 2
    prime_walls = "Y"
    prep_level = 2  # multiplier 1.4
    texture_factor = 1.00

    # SURFACES rooms (length, width, doors, windows, trim LF)
    rooms = [
        (18, 14, 2, 3, 60),
        (14, 12, 2, 2, 50),
        (12, 12, 1, 2, 44),
        (14, 13, 2, 2, 54),
        (12, 11, 1, 1, 46),
        (11, 10, 1, 1, 42),
        (9, 7, 1, 1, 30),
        (12, 4, 2, 0, 32),
        (0, 0, 0, 0, 0),
        (0, 0, 0, 0, 0),
    ]
    total_wall = 0
    total_ceil = 0
    total_trim_sf = 0
    total_trim_lf = 0
    total_doors = 0
    for length, width, doors, windows, trim in rooms:
        wall = max(0, 2 * (length + width) * ceil_height - doors * 21 - windows * 15)
        ceil = length * width
        trim_sf = trim * 0.5
        total_wall += wall
        total_ceil += ceil
        total_trim_sf += trim_sf
        total_trim_lf += trim
        total_doors += doors

    # PAINT (gallons, ROUNDUP per surface)
    wall_adj = total_wall * texture_factor
    ceil_adj = total_ceil * texture_factor

    walls_primer_sf = wall_adj if prime_walls == "Y" else 0
    walls_primer_gal = math.ceil(walls_primer_sf / cov_prime)
    ceil_primer_gal = math.ceil(ceil_adj / cov_prime)
    primer_gal = walls_primer_gal + ceil_primer_gal

    wall_top_gal = math.ceil((wall_adj * wall_coats) / cov_top)
    ceil_top_gal = math.ceil((ceil_adj * ceil_coats) / cov_top)
    trim_top_gal = math.ceil((total_trim_sf * trim_coats) / cov_trim)

    # LABOR
    prep_mult_lookup = {1: 1.0, 2: 1.4, 3: 2.0, 4: 3.0}
    prep_mult = prep_mult_lookup[prep_level]
    prep_hours = (total_wall + total_ceil) * 0.005 * prep_mult
    wall_hours = (total_wall * wall_coats) * 0.006
    ceil_hours = (total_ceil * ceil_coats) * 0.005
    trim_hours = (total_trim_lf * trim_coats) * 0.05
    door_hours = total_doors * 0.75
    setup_hours = 4.0
    total_hours = prep_hours + wall_hours + ceil_hours + trim_hours + door_hours + setup_hours

    # PRICING defaults
    price_primer = 28.0
    price_wall_top = 45.0
    price_ceil_top = 35.0
    price_trim_top = 55.0
    price_brushes = 120.0
    price_plastic = 60.0
    price_caulk = 35.0
    price_drops = 40.0
    labor_rate = 55.0
    margin = 0.25
    tax_rate = 0.07

    materials = (
        primer_gal * price_primer +
        wall_top_gal * price_wall_top +
        ceil_top_gal * price_ceil_top +
        trim_top_gal * price_trim_top +
        1 * price_brushes +
        1 * price_plastic +
        1 * price_caulk +
        1 * price_drops
    )
    labor_cost = total_hours * labor_rate
    cost_sub = materials + labor_cost
    margin_amt = cost_sub * margin
    pretax = cost_sub + margin_amt
    tax = pretax * tax_rate
    total_due = pretax + tax

    return {
        "TotalWallSF": total_wall,
        "TotalCeilSF": total_ceil,
        "TotalTrimSF": total_trim_sf,
        "TotalTrimLF": total_trim_lf,
        "PrimerGal": primer_gal,
        "WallTopGal": wall_top_gal,
        "CeilTopGal": ceil_top_gal,
        "TrimTopGal": trim_top_gal,
        "TotalHours": total_hours,
        "Materials": materials,
        "Labor": labor_cost,
        "CostSub": cost_sub,
        "Margin": margin_amt,
        "Pretax": pretax,
        "Tax": tax,
        "TotalDue": total_due,
    }


def approx(a, b, tol_pct=0.005, tol_abs=0.01):
    """Match within tol_pct (default 0.5%) OR tol_abs (default 1 cent)."""
    if a is None or b is None:
        return False
    if a == b:
        return True
    if abs(a - b) <= tol_abs:
        return True
    denom = max(abs(a), abs(b))
    return denom > 0 and abs(a - b) / denom <= tol_pct


def main():
    if not XLSX.exists():
        print(f"FAIL: {XLSX} does not exist — run build.py first.")
        sys.exit(1)

    print(f"Source: {XLSX}")
    print("Recalculating via LibreOffice headless...")
    recalced = recalc_via_libreoffice(XLSX)
    print(f"  -> {recalced}")

    # Step 1: scan for errors
    errors = scan_for_errors(recalced)
    if errors:
        print(f"\nFAIL: {len(errors)} Excel error cell(s):")
        for sheet, coord, val in errors[:20]:
            print(f"  {sheet}!{coord} = {val}")
        sys.exit(1)
    print("OK: no #REF!/#DIV/0!/#VALUE!/#NAME?/#NULL!/#NUM!/#N/A in any cell")

    # Step 2: numeric spot checks
    wb = load_workbook(recalced, data_only=True)
    exp = expected_results()

    # Map workbook cells (need to know addresses — read from formulas workbook)
    wb_f = load_workbook(XLSX, data_only=False)

    def resolve(name):
        dn = wb_f.defined_names[name]
        # attr_text like "SURFACES!$G$15"
        ref = dn.value
        sheet, coord = ref.split("!")
        coord = coord.replace("$", "")
        return wb[sheet][coord].value

    # Note: QUOTE subtotals don't have named ranges; read by sheet+coord
    # Find them dynamically by scanning QUOTE for labels
    quote = wb["QUOTE"]
    quote_vals = {}
    for row in quote.iter_rows():
        label_cell = row[0]
        total_cell = row[3] if len(row) >= 4 else None
        if isinstance(label_cell.value, str) and total_cell is not None:
            quote_vals[label_cell.value.strip()] = total_cell.value

    checks = [
        ("TotalWallSF",  resolve("TotalWallSF"),  exp["TotalWallSF"]),
        ("TotalCeilSF",  resolve("TotalCeilSF"),  exp["TotalCeilSF"]),
        ("TotalTrimSF",  resolve("TotalTrimSF"),  exp["TotalTrimSF"]),
        ("TotalTrimLF",  resolve("TotalTrimLF"),  exp["TotalTrimLF"]),
        ("PrimerGal",    resolve("PrimerGal"),    exp["PrimerGal"]),
        ("WallTopGal",   resolve("WallTopGal"),   exp["WallTopGal"]),
        ("CeilTopGal",   resolve("CeilTopGal"),   exp["CeilTopGal"]),
        ("TrimTopGal",   resolve("TrimTopGal"),   exp["TrimTopGal"]),
        ("TotalHours",   resolve("TotalHours"),   exp["TotalHours"]),
        ("Materials sub", quote_vals.get("Materials subtotal"),               exp["Materials"]),
        ("Cost sub",      quote_vals.get("Cost subtotal (materials + labor)"), exp["CostSub"]),
        ("Profit margin", quote_vals.get("Profit margin"),                     exp["Margin"]),
        ("Pre-tax total", quote_vals.get("Pre-tax total"),                     exp["Pretax"]),
        ("Sales tax",     quote_vals.get("Sales tax"),                         exp["Tax"]),
        ("TOTAL DUE",     quote_vals.get("TOTAL DUE"),                         exp["TotalDue"]),
    ]

    print("\nNumeric checks (workbook vs. pure-Python expected):")
    print(f"  {'Label':<22} {'Workbook':>14} {'Expected':>14}   Result")
    fails = 0
    for label, got, expv in checks:
        try:
            got_n = float(got) if got is not None else None
            exp_n = float(expv)
        except (TypeError, ValueError):
            print(f"  {label:<22} {'(not numeric)':>14} {'':>14}   FAIL")
            fails += 1
            continue
        ok = approx(got_n, exp_n)
        flag = "OK  " if ok else "FAIL"
        print(f"  {label:<22} {got_n:>14,.2f} {exp_n:>14,.2f}   {flag}")
        if not ok:
            fails += 1

    wb.close()
    wb_f.close()
    shutil.rmtree(recalced.parent, ignore_errors=True)

    if fails:
        print(f"\nFAIL: {fails} of {len(checks)} numeric checks failed")
        sys.exit(1)
    print(f"\nPASS: {len(checks)} checks, 0 Excel errors, {XLSX.name} is clean")


if __name__ == "__main__":
    main()
