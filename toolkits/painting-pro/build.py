"""Build the Painting Pro Toolkit workbook.

Run: python build.py
Output: dist/Painting-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Painting Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a list of rooms into paintable surface "
         "areas, gallons of primer + topcoat, prep and paint labor hours, and "
         "a print-ready customer quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs — fill them in.\n"
         "2. The SURFACES tab returns net paintable wall area (gross minus "
         "openings), ceiling area, and trim linear feet for every room.\n"
         "3. The PAINT tab converts those into gallons of primer and topcoat "
         "by coverage rate and number of coats.\n"
         "4. The LABOR tab returns prep + paint + cleanup hours by surface.\n"
         "5. The QUOTE tab is print-ready — update prices in PRICING first."),
        ("COVERAGE RATES",
         "Latex topcoat: 350 sq ft / gallon (smooth drywall, rolled).\n"
         "Latex topcoat: 200 sq ft / gallon (textured / rough / sprayed).\n"
         "Primer: 250 sq ft / gallon (one coat).\n"
         "Oil enamel trim: 400 sq ft / gallon (brushed).\n"
         "Trim coverage in this toolkit is converted from linear feet using "
         "a 0.5 ft equivalent surface width (base + cap)."),
        ("PREP LEVELS (labor multiplier)",
         "Level 1 (light dust, taping only): 1.00x base prep hours.\n"
         "Level 2 (drywall patches, light sanding): 1.40x.\n"
         "Level 3 (heavy patch, skim, stain block): 2.00x.\n"
         "Level 4 (lead-safe / hazardous prep): 3.00x. Use full RRP "
         "protocol if pre-1978."),
        ("PRICING",
         "Paint and labor unit costs on PRICING are 2026 ballpark — edit "
         "to your supplier and pay rate. Margin compounds on (material + "
         "labor). Defaults assume mid-grade interior latex."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please don't "
         "redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Residence Interior Repaint"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "PROJECT-LEVEL DEFAULTS", span=3); r += 1
    proj = [
        ("Ceiling height (ft)", 9, "0.0"),
        ("Topcoat coverage (sq ft / gal)", 350, "0"),
        ("Primer coverage (sq ft / gal)", 250, "0"),
        ("Trim coverage (sq ft / gal)", 400, "0"),
        ("Wall topcoat # of coats", 2, "0"),
        ("Ceiling topcoat # of coats", 1, "0"),
        ("Trim topcoat # of coats", 2, "0"),
        ("Prime walls (Y/N)", "Y", None),
        ("Prep level (1-4)", 2, "0"),
        ("Surface texture factor", 1.00, "0.00"),
    ]
    proj_start = r
    for label, default, fmt in proj:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["CeilHeight"] = DefinedName("CeilHeight", attr_text=f"INPUTS!$B${proj_start}")
    wb.defined_names["CovTop"] = DefinedName("CovTop", attr_text=f"INPUTS!$B${proj_start+1}")
    wb.defined_names["CovPrime"] = DefinedName("CovPrime", attr_text=f"INPUTS!$B${proj_start+2}")
    wb.defined_names["CovTrim"] = DefinedName("CovTrim", attr_text=f"INPUTS!$B${proj_start+3}")
    wb.defined_names["WallCoats"] = DefinedName("WallCoats", attr_text=f"INPUTS!$B${proj_start+4}")
    wb.defined_names["CeilCoats"] = DefinedName("CeilCoats", attr_text=f"INPUTS!$B${proj_start+5}")
    wb.defined_names["TrimCoats"] = DefinedName("TrimCoats", attr_text=f"INPUTS!$B${proj_start+6}")
    wb.defined_names["PrimeWalls"] = DefinedName("PrimeWalls", attr_text=f"INPUTS!$B${proj_start+7}")
    wb.defined_names["PrepLevel"] = DefinedName("PrepLevel", attr_text=f"INPUTS!$B${proj_start+8}")
    wb.defined_names["TextureFactor"] = DefinedName("TextureFactor", attr_text=f"INPUTS!$B${proj_start+9}")

    r += 1
    n = ws.cell(r, 1, "Prep options: 1=light dust  2=patch+sand  3=heavy skim  4=lead-safe (RRP)"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Texture factor: smooth=1.00  light texture=1.15  knockdown=1.30  popcorn=1.50"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Per-room geometry goes on the SURFACES tab (enter L × W per room)"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_surfaces(wb):
    ws = wb.create_sheet("SURFACES")
    auto_width(ws, [22, 10, 10, 10, 10, 10, 12, 14, 14, 14])
    banner(ws, "Surfaces by Room", "Enter L × W × openings; net paintable area calculated", width=10)

    r = 4
    section(ws, r, "ROOM TAKEOFF (yellow cells = enter values)", span=10); r += 1
    headers = ["Room", "Length", "Width", "Doors", "Windows", "Trim LF",
               "Wall sq ft", "Ceiling sq ft", "Trim equiv sq ft", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    rooms = [
        ("Living Room", 18, 14, 2, 3, 60, "main"),
        ("Kitchen", 14, 12, 2, 2, 50, "above cabs"),
        ("Dining", 12, 12, 1, 2, 44, ""),
        ("Master Bedroom", 14, 13, 2, 2, 54, ""),
        ("Bedroom 2", 12, 11, 1, 1, 46, ""),
        ("Bedroom 3", 11, 10, 1, 1, 42, ""),
        ("Bathroom", 9, 7, 1, 1, 30, "trim + door only"),
        ("Hallway", 12, 4, 2, 0, 32, ""),
        ("", 0, 0, 0, 0, 0, ""),
        ("", 0, 0, 0, 0, 0, ""),
    ]
    room_start = r
    for name, length, width, doors, windows, trim, note in rooms:
        ws.cell(r, 1, name).font = LBL; ws.cell(r, 1).border = BOX
        for col, val, fmt in [(2, length, "0.0"), (3, width, "0.0"),
                              (4, doors, "0"), (5, windows, "0"), (6, trim, "0")]:
            c = ws.cell(r, col, val); c.font = MONO; c.alignment = RIGHT
            c.number_format = fmt; c.fill = YELLOW; c.border = BOX
        # Wall sq ft = 2*(L+W)*ceiling height - doors*21 - windows*15
        # Doors typical 3'x7' = 21 sf opening; windows typical 3'x5' = 15 sf
        w = ws.cell(r, 7, f"=MAX(0,2*(B{r}+C{r})*CeilHeight-D{r}*21-E{r}*15)")
        w.font = MONO_BOLD; w.alignment = RIGHT; w.number_format = "0"; w.border = BOX
        # Ceiling sq ft = L*W
        c = ws.cell(r, 8, f"=B{r}*C{r}")
        c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        # Trim equiv sq ft = trim LF * 0.5 (base + cap avg 6" combined)
        t = ws.cell(r, 9, f"=F{r}*0.5")
        t.font = MONO_BOLD; t.alignment = RIGHT; t.number_format = "0"; t.border = BOX
        n = ws.cell(r, 10, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1
    room_end = r - 1

    ws.cell(r, 1, "TOTAL").font = LBL_BOLD; ws.cell(r, 1).fill = PANEL_FILL; ws.cell(r, 1).border = BOX
    for col in range(2, 7):
        ws.cell(r, col, "").fill = PANEL_FILL; ws.cell(r, col).border = BOX
    tw = ws.cell(r, 7, f"=SUM(G{room_start}:G{room_end})")
    tw.font = MONO_BOLD; tw.alignment = RIGHT; tw.number_format = "#,##0"; tw.fill = YELLOW; tw.border = BOX
    tc = ws.cell(r, 8, f"=SUM(H{room_start}:H{room_end})")
    tc.font = MONO_BOLD; tc.alignment = RIGHT; tc.number_format = "#,##0"; tc.fill = YELLOW; tc.border = BOX
    tt = ws.cell(r, 9, f"=SUM(I{room_start}:I{room_end})")
    tt.font = MONO_BOLD; tt.alignment = RIGHT; tt.number_format = "#,##0"; tt.fill = YELLOW; tt.border = BOX
    ws.cell(r, 10, "").border = BOX
    total_row = r
    wb.defined_names["TotalWallSF"] = DefinedName("TotalWallSF", attr_text=f"SURFACES!$G${total_row}")
    wb.defined_names["TotalCeilSF"] = DefinedName("TotalCeilSF", attr_text=f"SURFACES!$H${total_row}")
    wb.defined_names["TotalTrimSF"] = DefinedName("TotalTrimSF", attr_text=f"SURFACES!$I${total_row}")
    wb.defined_names["TotalTrimLF"] = DefinedName("TotalTrimLF", attr_text=f"SURFACES!$F${total_row}")
    # TotalTrimLF references column F which is empty in the TOTAL row — fix:
    # We need a row sum for trim LF too
    r += 1

    # Compute total trim LF (needed for labor)
    ws.cell(r, 1, "TOTAL trim linear ft").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in range(2, 6):
        ws.cell(r, col, "").fill = PANEL_FILL; ws.cell(r, col).border = BOX
    tlf = ws.cell(r, 6, f"=SUM(F{room_start}:F{room_end})")
    tlf.font = MONO_BOLD; tlf.alignment = RIGHT; tlf.number_format = "#,##0"; tlf.fill = YELLOW; tlf.border = BOX
    for col in range(7, 11):
        ws.cell(r, col, "").fill = PANEL_FILL; ws.cell(r, col).border = BOX
    # Override TotalTrimLF named range to point at the correct cell
    wb.defined_names["TotalTrimLF"] = DefinedName("TotalTrimLF", attr_text=f"SURFACES!$F${r}")
    r += 2

    note = ws.cell(r, 1,
        "Wall area assumes 3'×7' doors (21 sf) and 3'×5' windows "
        "(15 sf) subtracted from gross wall area. Trim is converted to "
        "equivalent paintable sq ft by 0.5 ft equivalent width (typical "
        "3 1/4\" base + 2 1/4\" door & window casing combined). Adjust the "
        "openings count for unusual sizes.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=10)
    return ws


def sheet_paint(wb):
    ws = wb.create_sheet("PAINT")
    auto_width(ws, [34, 14, 12, 14, 32])
    banner(ws, "Paint Quantities", "Gallons of primer + topcoat by surface", width=5)

    r = 4
    section(ws, r, "SURFACE AREAS (from SURFACES, adjusted for texture)", span=5); r += 1
    label_value(ws, r, "Wall sq ft (texture-adjusted)",
        "=TotalWallSF*TextureFactor", fmt="#,##0").font = MONO_BOLD
    wall_adj_row = r; r += 1
    label_value(ws, r, "Ceiling sq ft (texture-adjusted)",
        "=TotalCeilSF*TextureFactor", fmt="#,##0").font = MONO_BOLD
    ceil_adj_row = r; r += 1
    label_value(ws, r, "Trim equiv sq ft",
        "=TotalTrimSF", fmt="#,##0").font = MONO_BOLD
    trim_adj_row = r; r += 1

    r += 1
    section(ws, r, "PRIMER (one coat each)", span=5); r += 1
    headers = ["Surface", "Sq ft", "Coverage", "Gallons", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    # Walls primer (only if PrimeWalls = "Y")
    ws.cell(r, 1, "Walls (primer)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, f'=IF(UPPER(PrimeWalls)="Y",B{wall_adj_row},0)')
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    cov = ws.cell(r, 3, "=CovPrime"); cov.font = MONO; cov.alignment = RIGHT
    cov.number_format = "0"; cov.border = BOX
    gal = ws.cell(r, 4, f"=ROUNDUP(B{r}/C{r},0)")
    gal.font = MONO_BOLD; gal.alignment = RIGHT; gal.number_format = "0"; gal.border = BOX
    n = ws.cell(r, 5, "Skip if walls are same color and clean"); n.font = SMALL; n.border = BOX
    wall_primer_row = r; r += 1
    # Ceiling primer (only if no existing paint stain blocker scenario; default on)
    ws.cell(r, 1, "Ceiling (primer)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, f"=B{ceil_adj_row}")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    cov = ws.cell(r, 3, "=CovPrime"); cov.font = MONO; cov.alignment = RIGHT
    cov.number_format = "0"; cov.border = BOX
    gal = ws.cell(r, 4, f"=ROUNDUP(B{r}/C{r},0)")
    gal.font = MONO_BOLD; gal.alignment = RIGHT; gal.number_format = "0"; gal.border = BOX
    n = ws.cell(r, 5, "Stain-block primer for ceilings"); n.font = SMALL; n.border = BOX
    ceil_primer_row = r; r += 1

    ws.cell(r, 1, "Primer subtotal (gal)").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{wall_primer_row}:D{ceil_primer_row})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0"; s.fill = YELLOW; s.border = BOX
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    primer_total_row = r
    wb.defined_names["PrimerGal"] = DefinedName("PrimerGal", attr_text=f"PAINT!$D${primer_total_row}")
    r += 2

    section(ws, r, "TOPCOAT (per coat counts on INPUTS)", span=5); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Walls topcoat
    ws.cell(r, 1, "Walls (topcoat)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, f"=B{wall_adj_row}*WallCoats")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    cov = ws.cell(r, 3, "=CovTop"); cov.font = MONO; cov.alignment = RIGHT
    cov.number_format = "0"; cov.border = BOX
    gal = ws.cell(r, 4, f"=ROUNDUP(B{r}/C{r},0)")
    gal.font = MONO_BOLD; gal.alignment = RIGHT; gal.number_format = "0"; gal.border = BOX
    n = ws.cell(r, 5, "Mid-grade latex"); n.font = SMALL; n.border = BOX
    wall_top_row = r; r += 1
    # Ceiling topcoat
    ws.cell(r, 1, "Ceiling (topcoat)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, f"=B{ceil_adj_row}*CeilCoats")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    cov = ws.cell(r, 3, "=CovTop"); cov.font = MONO; cov.alignment = RIGHT
    cov.number_format = "0"; cov.border = BOX
    gal = ws.cell(r, 4, f"=ROUNDUP(B{r}/C{r},0)")
    gal.font = MONO_BOLD; gal.alignment = RIGHT; gal.number_format = "0"; gal.border = BOX
    n = ws.cell(r, 5, "Flat ceiling paint"); n.font = SMALL; n.border = BOX
    ceil_top_row = r; r += 1
    # Trim
    ws.cell(r, 1, "Trim (topcoat)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, f"=B{trim_adj_row}*TrimCoats")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    cov = ws.cell(r, 3, "=CovTrim"); cov.font = MONO; cov.alignment = RIGHT
    cov.number_format = "0"; cov.border = BOX
    gal = ws.cell(r, 4, f"=ROUNDUP(B{r}/C{r},0)")
    gal.font = MONO_BOLD; gal.alignment = RIGHT; gal.number_format = "0"; gal.border = BOX
    n = ws.cell(r, 5, "Semi-gloss enamel"); n.font = SMALL; n.border = BOX
    trim_top_row = r; r += 1

    ws.cell(r, 1, "Topcoat subtotal (gal)").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{wall_top_row}:D{trim_top_row})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0"; s.fill = YELLOW; s.border = BOX
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    top_total_row = r
    wb.defined_names["WallTopGal"] = DefinedName("WallTopGal", attr_text=f"PAINT!$D${wall_top_row}")
    wb.defined_names["CeilTopGal"] = DefinedName("CeilTopGal", attr_text=f"PAINT!$D${ceil_top_row}")
    wb.defined_names["TrimTopGal"] = DefinedName("TrimTopGal", attr_text=f"PAINT!$D${trim_top_row}")
    r += 2

    note = ws.cell(r, 1,
        "Gallons round UP to whole units (paint sells by the gallon/quart). "
        "Coverage rates default to 350 sf/gal for smooth drywall topcoat — "
        "drop to 200 sf/gal on textured walls by setting the texture factor "
        "on INPUTS (1.15 light texture, 1.30 knockdown, 1.50 popcorn).")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=5)
    return ws


def sheet_labor(wb):
    ws = wb.create_sheet("LABOR")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Labor Hours", "Prep + paint + cleanup, by surface", width=5)

    r = 4
    section(ws, r, "PREP MULTIPLIER", span=5); r += 1
    label_value(ws, r, "Prep level (from INPUTS)", "=PrepLevel", fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Prep multiplier",
        '=CHOOSE(PrepLevel,1,1.4,2,3)', fmt="0.00").font = MONO_BOLD
    prep_mult_row = r; r += 1

    r += 1
    section(ws, r, "BASE HOURS PER SURFACE", span=5); r += 1
    headers = ["Surface", "Sq ft / LF", "Hr per unit", "Hours", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Prep hours (per sq ft of wall + ceiling)
    ws.cell(r, 1, "Prep (walls + ceiling)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalWallSF+TotalCeilSF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.005)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}*B{prep_mult_row}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Patching, sanding, taping, plastic"); n.font = SMALL; n.border = BOX
    prep_hours_row = r; r += 1

    # Walls paint hours
    ws.cell(r, 1, "Walls (cut + roll, per coat)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalWallSF*WallCoats")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.006)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Cut-in + roll, mid-grade"); n.font = SMALL; n.border = BOX
    walls_hours_row = r; r += 1

    # Ceiling paint
    ws.cell(r, 1, "Ceiling (per coat)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalCeilSF*CeilCoats")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.005)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Roller + extension pole"); n.font = SMALL; n.border = BOX
    ceil_hours_row = r; r += 1

    # Trim paint (per LF, per coat)
    ws.cell(r, 1, "Trim (per LF, per coat)").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalTrimLF*TrimCoats")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.05)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Brush, slow"); n.font = SMALL; n.border = BOX
    trim_hours_row = r; r += 1

    # Doors hours (count)
    ws.cell(r, 1, "Doors (per door, both sides)").font = LBL; ws.cell(r, 1).border = BOX
    # Sum doors from SURFACES (column D, rows ROOM_DATA_START..ROOM_DATA_END)
    # The room data lives in SURFACES rows 6-15 (10 rows starting row 6)
    sf = ws.cell(r, 2, "=SUM(SURFACES!D6:D15)")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.75)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.00"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Includes mask/unmask hardware"); n.font = SMALL; n.border = BOX
    doors_hours_row = r; r += 1

    # Setup / cleanup
    ws.cell(r, 1, "Setup + daily cleanup").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, 1); sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 4.0)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.0"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Flat job allowance"); n.font = SMALL; n.border = BOX
    setup_hours_row = r; r += 1

    # Total
    ws.cell(r, 1, "TOTAL LABOR HOURS").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{prep_hours_row}:D{setup_hours_row})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0.0"; s.fill = YELLOW; s.border = BOX
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    total_hours_row = r
    wb.defined_names["TotalHours"] = DefinedName("TotalHours", attr_text=f"LABOR!$D${total_hours_row}")
    r += 2

    note = ws.cell(r, 1,
        "Base hours/unit are mid-grade production rates for an experienced "
        "painter. Slow trim hand-brushing on detailed colonial profiles can "
        "double the per-LF time. Prep multiplier compounds on the prep line only.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=5)
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "PAINT & MATERIALS (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("Primer, mid-grade latex (PVA)",        28.00, "gal"),
        ("Wall topcoat, mid-grade latex eggshell", 45.00, "gal"),
        ("Ceiling paint, flat",                  35.00, "gal"),
        ("Trim enamel, semi-gloss",              55.00, "gal"),
        ("Brushes, rollers, trays (per job)",   120.00, "set"),
        ("Plastic, paper, tape (per job)",       60.00, "set"),
        ("Caulk + spackle (per job)",            35.00, "set"),
        ("Drop cloths, ladders (per job)",       40.00, "set"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceTable"] = DefinedName("PriceTable", attr_text=f"PRICING!$A${pricing_start}:$C${r-1}")
    # Specific named refs for each consumable so QUOTE can pull cleanly
    wb.defined_names["PricePrimer"] = DefinedName("PricePrimer", attr_text=f"PRICING!$B${pricing_start}")
    wb.defined_names["PriceWallTop"] = DefinedName("PriceWallTop", attr_text=f"PRICING!$B${pricing_start+1}")
    wb.defined_names["PriceCeilTop"] = DefinedName("PriceCeilTop", attr_text=f"PRICING!$B${pricing_start+2}")
    wb.defined_names["PriceTrimTop"] = DefinedName("PriceTrimTop", attr_text=f"PRICING!$B${pricing_start+3}")
    wb.defined_names["PriceBrushes"] = DefinedName("PriceBrushes", attr_text=f"PRICING!$B${pricing_start+4}")
    wb.defined_names["PricePlastic"] = DefinedName("PricePlastic", attr_text=f"PRICING!$B${pricing_start+5}")
    wb.defined_names["PriceCaulk"] = DefinedName("PriceCaulk", attr_text=f"PRICING!$B${pricing_start+6}")
    wb.defined_names["PriceDrops"] = DefinedName("PriceDrops", attr_text=f"PRICING!$B${pricing_start+7}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Labor rate (lead painter)",         55.00, "per hour"),
        ("Profit margin (decimal)",            0.25, "0.25 = 25%"),
        ("Sales tax (decimal)",                0.07, "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit or "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LaborRate"] = DefinedName("LaborRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+2}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Wall sq ft", "=TotalWallSF", fmt="#,##0"); r += 1
    label_value(ws, r, "Ceiling sq ft", "=TotalCeilSF", fmt="#,##0"); r += 1
    label_value(ws, r, "Trim linear ft", "=TotalTrimLF", fmt="#,##0"); r += 1
    label_value(ws, r, "Total labor hours", "=TotalHours", fmt="0.0"); r += 1

    r += 1
    section(ws, r, "MATERIALS", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    mat_start = r
    mat_items = [
        ("Primer", "=PrimerGal", "=PricePrimer"),
        ("Wall topcoat", "=WallTopGal", "=PriceWallTop"),
        ("Ceiling topcoat", "=CeilTopGal", "=PriceCeilTop"),
        ("Trim enamel", "=TrimTopGal", "=PriceTrimTop"),
        ("Brushes / rollers / trays", 1, "=PriceBrushes"),
        ("Plastic / paper / tape", 1, "=PricePlastic"),
        ("Caulk / spackle", 1, "=PriceCaulk"),
        ("Drops / ladders", 1, "=PriceDrops"),
    ]
    for desc, qty, price in mat_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        p = ws.cell(r, 3, price); p.font = MONO; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    mat_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{mat_start}:D{mat_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    mat_sub_row = r; r += 1

    r += 1
    section(ws, r, "LABOR", span=4); r += 1
    ws.cell(r, 1, "Labor (total hours × rate)").font = LBL; ws.cell(r, 1).border = BOX
    c = ws.cell(r, 2, "=TotalHours"); c.font = MONO; c.alignment = RIGHT; c.number_format = "0.0"; c.border = BOX
    p = ws.cell(r, 3, "=LaborRate"); p.font = MONO; p.alignment = RIGHT
    p.number_format = '"$"#,##0.00'; p.border = BOX
    t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
    t.number_format = '"$"#,##0.00'; t.border = BOX
    labor_total_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{mat_sub_row}+D{labor_total_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3):
        c = ws.cell(r, col, ""); c.fill = YELLOW; c.border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Paint color matching, custom enamel, "
        "and additional coats beyond what's listed are billed separately. "
        "Lead-safe (RRP) protocols add ~30% to labor on pre-1978 surfaces.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_surfaces(wb)
    sheet_paint(wb)
    sheet_labor(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "SURFACES", "PAINT", "LABOR", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Painting-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
