"""Build the Electrical Pro Toolkit workbook.

Run: python build.py
Output: dist/Electrical-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Electrical Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a residential electrical scope into "
         "a NEC 220 service-load calc, branch-circuit takeoff, voltage-drop "
         "check, and a print-ready customer quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs.\n"
         "2. PANEL-LOAD runs the NEC 220 Part III standard-method calc — "
         "service amperage and recommended panel size.\n"
         "3. BRANCH-CIRCUITS lists every circuit, ampacity (310.16), "
         "breaker (240.6), and OCPD type.\n"
         "4. VOLTAGE-DROP confirms ≤ 3% drop on each branch (NEC informational).\n"
         "5. PRICING — set wire/device/labor unit costs once.\n"
         "6. Print QUOTE."),
        ("NEC 220 STANDARD METHOD",
         "General lighting + receptacles: 3 VA × sq ft. Add 1500 VA per "
         "small-appliance branch (min 2 kitchen + 1 laundry). Apply demand "
         "factor: first 3000 VA × 100%, next 117,000 × 35%. Add fixed "
         "appliances + dryer + range + EV charger at NEC values."),
        ("AMPACITY (NEC 310.16)",
         "75°C column for typical residential terminations. Common: 14 AWG "
         "= 15A, 12 AWG = 20A, 10 AWG = 30A, 8 AWG = 40A (50A copper), "
         "6 AWG = 55A (65A copper), 4 AWG = 70A (85A), 2 AWG = 95A (115A), "
         "1/0 = 125A (150A), 2/0 = 145A (175A), 4/0 = 195A (230A)."),
        ("VOLTAGE-DROP",
         "VD = 2 × K × I × L / cmil for single-phase. K = 12.9 for copper. "
         "Toolkit flags any branch > 3% as ‘UPSIZE’."),
        ("CONDUIT FILL",
         "Reference table on BRANCH-CIRCUITS lists max copper THHN/THWN-2 "
         "wires per ½\", ¾\", 1\" EMT (NEC Ch 9 Table 4 + Annex C)."),
        ("PRICING",
         "Unit prices on PRICING are 2026 ballpark — edit to match your "
         "supply house. Labor includes service swap, branch circuit per "
         "device, and trim/finish."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label); a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top"); a.fill = PANEL_FILL
        b = ws.cell(r, 2, body); b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Service Upgrade"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "DWELLING + SERVICE", span=3); r += 1
    bldg = [
        ("Conditioned floor area (sq ft)", 1800, "0"),
        ("Service voltage (V)", 240, "0"),
        ("Service phase", "Single-phase 240/120", None),
        ("Existing service (A)", 100, "0"),
        ("Target service (A)", 200, "0"),
        ("Number of small-appliance branches (min 2)", 2, "0"),
        ("Laundry circuits (min 1)", 1, "0"),
        ("Number of bathrooms (×20A circuit)", 2, "0"),
    ]
    bldg_start = r
    for label, default, fmt in bldg:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["FloorArea"] = DefinedName("FloorArea", attr_text=f"INPUTS!$B${bldg_start}")
    wb.defined_names["Volt"] = DefinedName("Volt", attr_text=f"INPUTS!$B${bldg_start+1}")
    wb.defined_names["Phase"] = DefinedName("Phase", attr_text=f"INPUTS!$B${bldg_start+2}")
    wb.defined_names["ExistingAmps"] = DefinedName("ExistingAmps", attr_text=f"INPUTS!$B${bldg_start+3}")
    wb.defined_names["TargetAmps"] = DefinedName("TargetAmps", attr_text=f"INPUTS!$B${bldg_start+4}")
    wb.defined_names["SAB"] = DefinedName("SAB", attr_text=f"INPUTS!$B${bldg_start+5}")
    wb.defined_names["LaundryC"] = DefinedName("LaundryC", attr_text=f"INPUTS!$B${bldg_start+6}")
    wb.defined_names["NumBaths"] = DefinedName("NumBaths", attr_text=f"INPUTS!$B${bldg_start+7}")

    r += 1
    section(ws, r, "FIXED APPLIANCES (VA, blank if none)", span=3); r += 1
    appl = [
        ("Electric range (VA, NEC 8000)",        8000, "0"),
        ("Electric dryer (VA, NEC 5000)",        5000, "0"),
        ("Water heater electric (VA)",           4500, "0"),
        ("Heat pump (VA, indoor + outdoor)",     7200, "0"),
        ("AC condenser only (VA)",               4800, "0"),
        ("Furnace blower (VA)",                  600,  "0"),
        ("Dishwasher (VA)",                      1200, "0"),
        ("Garbage disposal (VA)",                900,  "0"),
        ("Microwave dedicated (VA)",             1500, "0"),
        ("EV charger 40A 240V (VA = 9600)",      9600, "0"),
        ("Pool / spa pump (VA)",                 0,    "0"),
        ("Other / largest motor (VA)",           0,    "0"),
    ]
    appl_start = r
    for label, default, fmt in appl:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["Range"] = DefinedName("Range", attr_text=f"INPUTS!$B${appl_start}")
    wb.defined_names["Dryer"] = DefinedName("Dryer", attr_text=f"INPUTS!$B${appl_start+1}")
    wb.defined_names["WaterHeat"] = DefinedName("WaterHeat", attr_text=f"INPUTS!$B${appl_start+2}")
    wb.defined_names["HeatPump"] = DefinedName("HeatPump", attr_text=f"INPUTS!$B${appl_start+3}")
    wb.defined_names["AC"] = DefinedName("AC", attr_text=f"INPUTS!$B${appl_start+4}")
    wb.defined_names["Furnace"] = DefinedName("Furnace", attr_text=f"INPUTS!$B${appl_start+5}")
    wb.defined_names["Dishwasher"] = DefinedName("Dishwasher", attr_text=f"INPUTS!$B${appl_start+6}")
    wb.defined_names["Disposal"] = DefinedName("Disposal", attr_text=f"INPUTS!$B${appl_start+7}")
    wb.defined_names["Microwave"] = DefinedName("Microwave", attr_text=f"INPUTS!$B${appl_start+8}")
    wb.defined_names["EVCharger"] = DefinedName("EVCharger", attr_text=f"INPUTS!$B${appl_start+9}")
    wb.defined_names["PoolPump"] = DefinedName("PoolPump", attr_text=f"INPUTS!$B${appl_start+10}")
    wb.defined_names["Other"] = DefinedName("Other", attr_text=f"INPUTS!$B${appl_start+11}")

    r += 1
    n = ws.cell(r, 1, "Service phase: Single-phase 240/120 | Three-phase 208Y/120 | Three-phase 480Y/277")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_panel_load(wb):
    ws = wb.create_sheet("PANEL-LOAD")
    auto_width(ws, [42, 14, 14, 30])
    banner(ws, "NEC 220 Part III Standard Method", "Single-phase 240/120 dwelling service load", width=4)

    r = 4
    section(ws, r, "GENERAL LIGHTING + RECEPTACLES", span=4); r += 1
    label_value(ws, r, "General lighting (3 VA × sq ft)", "=FloorArea*3", fmt="#,##0").font = MONO_BOLD
    gl_row = r; r += 1
    label_value(ws, r, "Small-appliance branches (1500 VA each)", "=SAB*1500", fmt="#,##0").font = MONO_BOLD
    sab_row = r; r += 1
    label_value(ws, r, "Laundry circuit (1500 VA each)", "=LaundryC*1500", fmt="#,##0").font = MONO_BOLD
    lc_row = r; r += 1

    label_value(ws, r, "Lighting + receptacle subtotal (VA)",
        f"=B{gl_row}+B{sab_row}+B{lc_row}", fmt="#,##0").font = MONO_BOLD
    sub1_row = r; r += 1

    r += 1
    section(ws, r, "DEMAND FACTOR APPLIED (NEC 220.42)", span=4); r += 1
    label_value(ws, r, "First 3000 VA × 100%",
        f"=MIN(B{sub1_row},3000)", fmt="#,##0").font = MONO_BOLD
    df1_row = r; r += 1
    label_value(ws, r, "Remainder × 35%",
        f"=MAX(0,B{sub1_row}-3000)*0.35", fmt="#,##0").font = MONO_BOLD
    df2_row = r; r += 1
    label_value(ws, r, "Lighting after demand (VA)",
        f"=B{df1_row}+B{df2_row}", fmt="#,##0").font = MONO_BOLD
    light_after_row = r; r += 1

    r += 1
    section(ws, r, "FIXED APPLIANCES + LARGE LOADS", span=4); r += 1
    headers = ["Appliance", "Rated VA", "Demand %", "Demand VA"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    appl_rows = [
        ("Electric range",   "=Range",       "75%",  "=Range*0.75"),
        ("Electric dryer",   "=Dryer",       "100%", "=Dryer*1.0"),
        ("Water heater",     "=WaterHeat",   "100%", "=WaterHeat*1.0"),
        ("Heat pump (heating side)", "=HeatPump", "100%", "=HeatPump*1.0"),
        ("AC condenser",     "=AC",          "100%", "=AC*1.0"),
        ("Furnace blower",   "=Furnace",     "100%", "=Furnace*1.0"),
        ("Dishwasher",       "=Dishwasher",  "75%",  "=Dishwasher*0.75"),
        ("Disposal",         "=Disposal",    "75%",  "=Disposal*0.75"),
        ("Microwave",        "=Microwave",   "75%",  "=Microwave*0.75"),
        ("EV charger",       "=EVCharger",   "100%", "=EVCharger*1.0"),
        ("Pool pump",        "=PoolPump",    "100%", "=PoolPump*1.0"),
        ("Other / largest motor + 25%", "=Other", "125%", "=Other*1.25"),
    ]
    appl_start = r
    for label, va, dem, demva in appl_rows:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, va); c.font = MONO; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        d_ = ws.cell(r, 3, dem); d_.font = MONO; d_.alignment = CENTER; d_.border = BOX
        e = ws.cell(r, 4, demva); e.font = MONO_BOLD; e.alignment = RIGHT; e.number_format = "#,##0"; e.border = BOX
        r += 1
    appl_end = r - 1

    ws.cell(r, 1, "Appliance demand subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{appl_start}:D{appl_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "#,##0"; s.border = BOX; s.fill = PANEL_FILL
    appl_sub_row = r; r += 1

    r += 1
    section(ws, r, "TOTAL SERVICE DEMAND", span=4); r += 1
    label_value(ws, r, "Total demand VA",
        f"=B{light_after_row}+D{appl_sub_row}", fmt="#,##0").font = MONO_BOLD
    total_va_row = r; r += 1
    label_value(ws, r, "Total demand A (÷ 240V)",
        f"=B{total_va_row}/Volt", fmt="0.0").font = MONO_BOLD
    total_a_row = r; r += 1
    label_value(ws, r, "Recommended service size",
        f'=IF(B{total_a_row}<=80,"100A",IF(B{total_a_row}<=160,"200A",IF(B{total_a_row}<=320,"400A","ENGINEER")))',
        fmt=None).font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    rec_row = r; r += 1
    label_value(ws, r, "Headroom vs. 200A panel (%)",
        f"=(200-B{total_a_row})/200", fmt="0.0%").font = MONO_BOLD; r += 1

    wb.defined_names["TotalServiceA"] = DefinedName("TotalServiceA", attr_text=f"'PANEL-LOAD'!$B${total_a_row}")
    wb.defined_names["TotalServiceVA"] = DefinedName("TotalServiceVA", attr_text=f"'PANEL-LOAD'!$B${total_va_row}")

    r += 1
    note = ws.cell(r, 1,
        "ESTIMATE ONLY — NEC 220 Part III standard method, single-phase. "
        "Optional method (220.82) often returns a smaller service for "
        "all-electric homes. Always confirm with your AHJ for permit submittal.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)
    return ws


def sheet_branch_circuits(wb):
    ws = wb.create_sheet("BRANCH-CIRCUITS")
    auto_width(ws, [30, 12, 12, 14, 14, 26])
    banner(ws, "Branch Circuit Schedule", "Wire size (NEC 310.16, 75°C), breaker (NEC 240.6), notes", width=6)

    r = 4
    headers = ["Circuit", "AWG", "Amps", "Breaker", "OCPD Type", "Note"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1

    rows = [
        ("Lighting (general)",         "14",    "15A",  "15A",  "STD",   "Most rooms; arc-fault on bedrooms"),
        ("Receptacles (general)",      "12",    "20A",  "20A",  "AFCI",  "All habitable spaces 210.12"),
        ("Small appliance branch #1",  "12",    "20A",  "20A",  "GFCI",  "Kitchen counter (210.52)"),
        ("Small appliance branch #2",  "12",    "20A",  "20A",  "GFCI",  "Kitchen counter (210.52)"),
        ("Bathroom receptacles",       "12",    "20A",  "20A",  "GFCI",  "Per bath (one circuit/bath OK)"),
        ("Laundry receptacle",         "12",    "20A",  "20A",  "GFCI",  "Dedicated, no other loads"),
        ("Garage / outdoor receptacles","12",   "20A",  "20A",  "GFCI",  "210.8(A)(2,3)"),
        ("Refrigerator (dedicated)",   "12",    "20A",  "20A",  "STD",   "May share SAB if listed"),
        ("Dishwasher",                 "12",    "20A",  "20A",  "GFCI",  "210.8(D) effective 2020"),
        ("Disposal",                   "12",    "20A",  "20A",  "STD",   "Switch on counter, motor-rated"),
        ("Microwave (dedicated)",      "12",    "20A",  "20A",  "STD",   "Range hood combo unit"),
        ("Range (8 kW or less)",       "10",    "30A",  "30A",  "STD",   "≤ 8 kW = 40A; ≥ 8 kW = 50A"),
        ("Range (8.75-12 kW)",         "8",     "40A",  "40A",  "STD",   "Most modern ranges"),
        ("Dryer (electric, 30A)",      "10",    "30A",  "30A",  "STD",   "NEMA 14-30, 4-wire"),
        ("Water heater (electric)",    "10",    "30A",  "30A",  "STD",   "4500W typical"),
        ("Heat pump (3 ton)",          "8",     "40A",  "40A",  "HACR",  "Per nameplate MCA / MOCP"),
        ("AC condenser (3 ton)",       "10",    "30A",  "30A",  "HACR",  "Per nameplate"),
        ("Furnace blower",             "14",    "15A",  "15A",  "STD",   "Often shared with AC at AHJ"),
        ("EV charger 40A continuous",  "8",     "50A",  "50A",  "GFCI",  "625.42; 125% continuous"),
        ("Pool pump 240V",             "12",    "20A",  "20A",  "GFCI",  "680.21 GFCI required"),
    ]
    for c1, c2, c3, c4, c5, c6 in rows:
        ws.cell(r, 1, c1).font = LBL; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, c2).font = MONO_BOLD; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, c3).font = MONO; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        ws.cell(r, 4, c4).font = MONO_BOLD; ws.cell(r, 4).alignment = CENTER; ws.cell(r, 4).border = BOX
        ws.cell(r, 5, c5).font = MONO; ws.cell(r, 5).alignment = CENTER; ws.cell(r, 5).border = BOX
        n = ws.cell(r, 6, c6); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "CONDUIT FILL — MAX THHN/THWN-2 COPPER PER NEC CH 9", span=6); r += 1
    headers2 = ["Wire AWG", "1/2\" EMT", "3/4\" EMT", "1\" EMT", "1-1/4\" EMT", ""]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    fill = [
        ("14",  9,  16, 26, 45),
        ("12",  7,  12, 19, 33),
        ("10",  4,  7,  12, 21),
        ("8",   2,  4,  7,  12),
        ("6",   1,  3,  5,  9),
        ("4",   1,  2,  3,  6),
        ("3",   1,  1,  3,  5),
        ("2",   1,  1,  3,  4),
        ("1",   0,  1,  1,  3),
        ("1/0", 0,  1,  1,  3),
        ("2/0", 0,  1,  1,  2),
        ("4/0", 0,  0,  1,  1),
    ]
    for awg, c12, c34, c1, c114 in fill:
        ws.cell(r, 1, awg).font = MONO_BOLD; ws.cell(r, 1).alignment = CENTER; ws.cell(r, 1).border = BOX
        for col, v in zip((2, 3, 4, 5), (c12, c34, c1, c114)):
            c = ws.cell(r, col, v); c.font = MONO; c.alignment = CENTER; c.number_format = "0"; c.border = BOX
        ws.cell(r, 6, "").border = BOX
        r += 1

    note = ws.cell(r, 1,
        "Conduit fill from NEC Annex C Table C.1 (EMT, THHN/THWN-2). "
        "Derate ampacity per 310.15(C) when more than 3 current-carrying "
        "conductors share a raceway.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    return ws


def sheet_voltage_drop(wb):
    ws = wb.create_sheet("VOLTAGE-DROP")
    auto_width(ws, [28, 12, 12, 12, 14, 14, 14])
    banner(ws, "Voltage Drop (NEC 210.19 Informational)", "VD = 2 × K × I × L / cmil  ·  K=12.9 copper", width=7)

    r = 4
    headers = ["Circuit", "Voltage", "Amps", "Length (ft)", "AWG", "VD (V)", "% Drop"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1

    # Common branch examples — yellow editable, formulas compute drop
    rows = [
        ("Kitchen SAB",             120, 16,  35,  "12"),
        ("Bathroom",                120, 16,  40,  "12"),
        ("Garage receptacles",      120, 16,  60,  "12"),
        ("EV charger",              240, 32,  75,  "8"),
        ("Range",                   240, 33,  20,  "8"),
        ("Dryer",                   240, 22,  35,  "10"),
        ("AC condenser",            240, 18,  45,  "10"),
        ("Heat pump",               240, 30,  40,  "8"),
        ("Sub-panel feeder (60A)",  240, 60,  120, "6"),
        ("Pool pump",               240, 12,  150, "12"),
    ]
    # Circular mils per AWG (copper)
    cmil = {"14": 4110, "12": 6530, "10": 10380, "8": 16510, "6": 26240,
            "4": 41740, "3": 52620, "2": 66360, "1": 83690, "1/0": 105600,
            "2/0": 133100, "3/0": 167800, "4/0": 211600}

    for circ, v, i_amps, length, awg in rows:
        ws.cell(r, 1, circ).font = LBL; ws.cell(r, 1).border = BOX
        c1 = ws.cell(r, 2, v); c1.font = MONO; c1.alignment = CENTER; c1.fill = YELLOW; c1.border = BOX
        c2 = ws.cell(r, 3, i_amps); c2.font = MONO; c2.alignment = RIGHT; c2.number_format = "0"; c2.fill = YELLOW; c2.border = BOX
        c3 = ws.cell(r, 4, length); c3.font = MONO; c3.alignment = RIGHT; c3.number_format = "0"; c3.fill = YELLOW; c3.border = BOX
        c4 = ws.cell(r, 5, awg); c4.font = MONO_BOLD; c4.alignment = CENTER; c4.fill = YELLOW; c4.border = BOX
        # VD formula: 2 * 12.9 * I * L / cmil_lookup
        cmil_val = cmil.get(awg, 6530)
        vd = ws.cell(r, 6, f"=2*12.9*C{r}*D{r}/{cmil_val}")
        vd.font = MONO_BOLD; vd.alignment = RIGHT; vd.number_format = "0.00"; vd.border = BOX
        pct = ws.cell(r, 7, f"=F{r}/B{r}")
        pct.font = MONO_BOLD; pct.alignment = RIGHT; pct.number_format = "0.0%"; pct.border = BOX
        r += 1

    r += 1
    section(ws, r, "FLAGS", span=7); r += 1
    note = ws.cell(r, 1,
        "Branches with > 3% drop will not deliver rated current efficiently. "
        "NEC 210.19 informational note — not strictly required, but most AHJs "
        "and good design call for ≤ 3% on branch circuits + ≤ 2% on feeders, "
        "totaling ≤ 5% from service to outlet. Upsize one AWG step per "
        "doubling of length above the limit.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "WIRE + CONDUIT (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    items = [
        ("Romex 14/2 NM-B (250 ft coil)",       95.00,  "coil"),
        ("Romex 12/2 NM-B (250 ft coil)",       145.00, "coil"),
        ("Romex 12/3 NM-B (250 ft coil)",       195.00, "coil"),
        ("Romex 10/2 NM-B (125 ft coil)",       155.00, "coil"),
        ("Romex 10/3 NM-B (125 ft coil)",       210.00, "coil"),
        ("Romex 8/3 NM-B (125 ft coil)",        320.00, "coil"),
        ("Romex 6/3 NM-B (125 ft coil)",        425.00, "coil"),
        ("THHN 12 AWG (per ft)",                0.42,   "lin ft"),
        ("THHN 8 AWG (per ft)",                 1.10,   "lin ft"),
        ("THHN 6 AWG (per ft)",                 1.65,   "lin ft"),
        ("EMT 1/2\" (10 ft)",                   8.50,   "stick"),
        ("EMT 3/4\" (10 ft)",                   12.50,  "stick"),
        ("EMT 1\" (10 ft)",                     18.50,  "stick"),
        ("EMT connectors 1/2\"",                1.40,   "ea"),
        ("Single-gang plastic box (NM)",        0.55,   "ea"),
        ("Double-gang plastic box (NM)",        0.95,   "ea"),
        ("Round ceiling box (fan-rated)",       4.85,   "ea"),
        ("4\" sq metal box + mud ring",         3.20,   "ea"),
        ("Switch (single-pole, 15A)",           1.85,   "ea"),
        ("Switch (3-way, 15A)",                 4.50,   "ea"),
        ("Receptacle (15A duplex)",             1.65,   "ea"),
        ("Receptacle (20A duplex, GFCI)",       18.00,  "ea"),
        ("Tamper-resistant 15A",                3.50,   "ea"),
        ("AFCI 15A breaker",                    48.00,  "ea"),
        ("AFCI 20A breaker",                    52.00,  "ea"),
        ("Standard 15/20A breaker",             8.00,   "ea"),
        ("2-pole 30A breaker",                  22.00,  "ea"),
        ("2-pole 50A breaker",                  35.00,  "ea"),
        ("200A main breaker panel (40 spc)",    485.00, "ea"),
        ("100A subpanel (24 spc)",              225.00, "ea"),
        ("EV charger 40A NEMA 14-50 outlet",    32.00,  "ea"),
        ("Whole-house surge protector (Type 2)",185.00, "ea"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceTable"] = DefinedName("PriceTable", attr_text=f"PRICING!$A${pricing_start}:$C${r-1}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Lead electrician labor",       115.00, "per hour"),
        ("Helper / apprentice labor",    65.00,  "per hour"),
        ("Service swap labor (200A)",    18.00,  "hours"),
        ("Branch circuit (per device)",  85.00,  "per device"),
        ("Sub-panel install",            450.00, "flat"),
        ("Permit + inspection",          275.00, "flat"),
        ("Profit margin (decimal)",      0.25,   "0.25 = 25%"),
        ("Sales tax (decimal)",          0.07,   "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit:
            v.number_format = "0.0%"
        elif unit == "hours":
            v.number_format = "0.0"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LeadRate"] = DefinedName("LeadRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["HelperRate"] = DefinedName("HelperRate", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["SwapHrs"] = DefinedName("SwapHrs", attr_text=f"PRICING!$B${labor_start+2}")
    wb.defined_names["DeviceRate"] = DefinedName("DeviceRate", attr_text=f"PRICING!$B${labor_start+3}")
    wb.defined_names["SubPanel"] = DefinedName("SubPanel", attr_text=f"PRICING!$B${labor_start+4}")
    wb.defined_names["Permit"] = DefinedName("Permit", attr_text=f"PRICING!$B${labor_start+5}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+6}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+7}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Service size",
        '=IF(TotalServiceA<=80,"100A",IF(TotalServiceA<=160,"200A",IF(TotalServiceA<=320,"400A","ENGINEER")))'); r += 1
    label_value(ws, r, "Calculated demand", "=ROUND(TotalServiceA,0)&\" A @ \"&Volt&\" V\""); r += 1
    label_value(ws, r, "Floor area", "=FloorArea&\" sq ft\""); r += 1

    r += 1
    section(ws, r, "MATERIALS (typical 200A service swap + EV)", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    eq_start = r
    eq_items = [
        ("200A main breaker panel (40 spc)",  1,  "200A main breaker panel (40 spc)"),
        ("Whole-house surge protector",       1,  "Whole-house surge protector (Type 2)"),
        ("Standard 15/20A breakers",          18, "Standard 15/20A breaker"),
        ("AFCI 15A breakers",                 4,  "AFCI 15A breaker"),
        ("AFCI 20A breakers",                 6,  "AFCI 20A breaker"),
        ("2-pole 30A (dryer + WH)",           2,  "2-pole 30A breaker"),
        ("2-pole 50A (range + EV)",           2,  "2-pole 50A breaker"),
        ("Romex 14/2 NM-B (coil)",            2,  "Romex 14/2 NM-B (250 ft coil)"),
        ("Romex 12/2 NM-B (coil)",            3,  "Romex 12/2 NM-B (250 ft coil)"),
        ("Romex 10/2 NM-B (coil)",            1,  "Romex 10/2 NM-B (125 ft coil)"),
        ("Romex 8/3 NM-B (coil)",             1,  "Romex 8/3 NM-B (125 ft coil)"),
        ("Switches (single-pole)",            22, "Switch (single-pole, 15A)"),
        ("Receptacles (15A TR)",              42, "Tamper-resistant 15A"),
        ("GFCI receptacles (kitchen+bath)",   8,  "Receptacle (20A duplex, GFCI)"),
        ("Single-gang plastic boxes",         48, "Single-gang plastic box (NM)"),
        ("EV charger 14-50 outlet",           1,  "EV charger 40A NEMA 14-50 outlet"),
    ]
    for desc, qty, price_label in eq_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        p = ws.cell(r, 3, f'=VLOOKUP("{price_label}",PriceTable,2,FALSE)')
        p.font = MONO; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    eq_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{eq_start}:D{eq_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    eq_subtotal_row = r; r += 1

    r += 1
    section(ws, r, "LABOR & FEES", span=4); r += 1
    labor_lines = [
        ("Service swap labor (lead, hr)",       "=SwapHrs*LeadRate"),
        ("Service swap labor (helper, hr)",     "=SwapHrs*HelperRate"),
        ("Branch circuit labor (50 devices)",   "=50*DeviceRate"),
        ("EV charger circuit labor",            "=DeviceRate*2"),
        ("Permit + inspection",                 "=Permit"),
    ]
    labor_start = r
    for desc, formula in labor_lines:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        for col in (2, 3): ws.cell(r, col, "").border = BOX
        t = ws.cell(r, 4, formula); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    labor_end = r - 1

    ws.cell(r, 1, "Labor subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{labor_start}:D{labor_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    labor_subtotal_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{eq_subtotal_row}+D{labor_subtotal_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Material prices subject to change. Work "
        "scheduled within 2-3 weeks of signed acceptance and utility coordination.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_panel_load(wb)
    sheet_branch_circuits(wb)
    sheet_voltage_drop(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "PANEL-LOAD", "BRANCH-CIRCUITS", "VOLTAGE-DROP", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Electrical-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
