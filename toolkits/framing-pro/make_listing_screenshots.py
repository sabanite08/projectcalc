"""Generate Etsy listing images using REAL workbook screenshots.

Pipeline:
  1. Load HVAC-Pro-Toolkit-v1.xlsx
  2. Force every sheet to print as a single page (fitToPage)
  3. Convert to PDF via LibreOffice
  4. Render each PDF page to PNG via pymupdf
  5. Composite each into a 2000×2000 brand-framed listing image
  6. Build the cover

Run: python make_listing_screenshots.py
Output: dist/listing/01_cover.png ... 05_quote.png  (overwrites the design mockups)
"""

from pathlib import Path
import subprocess
import shutil
import tempfile
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
import pymupdf
from PIL import Image, ImageDraw, ImageFont

# ──────────────────────────────────────────────────────────────────────
# Per-toolkit configuration — change these to retarget the script
# ──────────────────────────────────────────────────────────────────────
TOOLKIT_FOLDER = Path(__file__).parent
XLSX_NAME = "Framing-Pro-Toolkit-v1.xlsx"
TOOLKIT_NAME = "FRAMING PRO TOOLKIT"
TAGLINE = "Lumber takeoff · Cut list · Plywood · Quote"
PRICE = "$39"

# Which sheets to feature, in order. Each becomes one listing image (after the cover).
# (sheet_name_in_xlsx, caption_for_brand_bar)
SHEETS_TO_FEATURE = [
    ("README",       "README — what's inside"),
    ("INPUTS",       "INPUTS — yellow cells, type your numbers"),
    ("TAKEOFF",      "TAKEOFF — studs, joists, rafters, fasteners"),
    ("QUOTE",        "QUOTE — print-ready customer quote"),
]
# The sheet shown on the cover (hero image)
COVER_SHEET = "QUOTE"

# ──────────────────────────────────────────────────────────────────────
# Brand
# ──────────────────────────────────────────────────────────────────────
BG = (14, 14, 12)
BG_2 = (22, 22, 19)
LINE = (42, 42, 38)
INK = (244, 241, 234)
INK_2 = (184, 179, 167)
INK_3 = (118, 114, 106)
HI_VIS = (255, 212, 0)

FN_BLACK = "C:/Windows/Fonts/ariblk.ttf"
FN_BOLD = "C:/Windows/Fonts/arialbd.ttf"
FN_REG = "C:/Windows/Fonts/arial.ttf"
FN_MONO = "C:/Windows/Fonts/consola.ttf"
FN_MONO_BOLD = "C:/Windows/Fonts/consolab.ttf"

W = H = 2000
SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"

OUT = TOOLKIT_FOLDER / "dist" / "listing"
OUT.mkdir(parents=True, exist_ok=True)


def f(path, size):
    return ImageFont.truetype(path, size)


def text_centered(d, text, font, fill, cx, y):
    bbox = d.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    d.text((cx - w // 2, y), text, fill=fill, font=font)


def text_right(d, text, font, fill, x_right, y):
    bbox = d.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    d.text((x_right - w, y), text, fill=fill, font=font)


# ──────────────────────────────────────────────────────────────────────
# Step 1-3: render every sheet to its own PDF page
# ──────────────────────────────────────────────────────────────────────
def render_sheets_to_png(xlsx_path: Path) -> dict:
    """Returns {sheet_name: PIL.Image} at high resolution."""
    work_dir = Path(tempfile.mkdtemp(prefix="pc_screens_"))

    # Copy + force fit-to-1-page on every sheet
    src = work_dir / xlsx_path.name
    shutil.copy(xlsx_path, src)
    wb = load_workbook(src)
    sheet_order = list(wb.sheetnames)
    for name in sheet_order:
        ws = wb[name]
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.orientation = "landscape"
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
        ws.print_options.horizontalCentered = True
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.print_area = None
    wb.save(src)

    # Convert to PDF
    subprocess.run([
        SOFFICE, "--headless", "--calc",
        "--convert-to", "pdf",
        "--outdir", str(work_dir),
        str(src),
    ], check=True, capture_output=True, timeout=120)

    pdf_path = work_dir / src.with_suffix(".pdf").name
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not produced at {pdf_path}")

    # Render each page as PNG via pymupdf — one page per sheet
    doc = pymupdf.open(pdf_path)
    if len(doc) != len(sheet_order):
        # Sometimes a sheet still spills to 2 pages despite fit-to-page.
        # Just take the first page corresponding to each sheet by ordinal mapping
        # we can't tell which page belongs to which sheet without more work.
        # Best effort: if equal, 1:1; if not, take pages 0..n in order.
        print(f"  warning: {len(doc)} PDF pages for {len(sheet_order)} sheets — using 1:N mapping anyway")

    out_images = {}
    zoom = 2.5  # ~ 180 DPI for 72-pt PDF
    mat = pymupdf.Matrix(zoom, zoom)
    for idx, sheet in enumerate(sheet_order):
        if idx >= len(doc):
            break
        page = doc[idx]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        out_images[sheet] = img

    doc.close()
    shutil.rmtree(work_dir, ignore_errors=True)
    return out_images


# ──────────────────────────────────────────────────────────────────────
# Step 4: composite a screenshot into a branded listing image
# ──────────────────────────────────────────────────────────────────────
def build_sheet_image(screenshot: Image.Image, caption: str) -> Image.Image:
    """Top brand bar + dominant screenshot below."""
    img = Image.new("RGB", (W, H), BG)
    d = ImageDraw.Draw(img)

    # Top brand bar (black with hi-vis accent)
    bar_h = 140
    d.rectangle([0, 0, W, bar_h], fill=BG_2)
    d.rectangle([0, bar_h, W, bar_h + 6], fill=HI_VIS)

    # Brand mark + caption
    s = 26
    bx, by = 60, 56
    d.polygon([(bx, by + s // 2), (bx + s // 2, by), (bx + s, by + s // 2), (bx + s // 2, by + s)], fill=HI_VIS)
    d.text((bx + s + 18, by - 4), "PROJECTCALC", fill=INK, font=f(FN_BLACK, 32))

    # Toolkit name + sheet caption
    d.text((bx + s + 320, by - 4), TOOLKIT_NAME, fill=HI_VIS, font=f(FN_MONO_BOLD, 26))
    text_right(d, caption, f(FN_MONO_BOLD, 26), INK, W - 60, by - 2)

    # Screenshot area: from y=200 to y=1900, width fills with margin
    pad_x = 80
    pad_top = 200
    pad_bottom = 100
    avail_w = W - pad_x * 2
    avail_h = H - pad_top - pad_bottom

    # Add white "paper" background behind the screenshot so it reads as Excel
    # Scale screenshot to fit avail with letterboxing
    sw, sh = screenshot.size
    scale = min(avail_w / sw, avail_h / sh)
    new_w = int(sw * scale)
    new_h = int(sh * scale)
    scaled = screenshot.resize((new_w, new_h), Image.LANCZOS)

    # Paper frame
    sx = pad_x + (avail_w - new_w) // 2
    sy = pad_top + (avail_h - new_h) // 2

    # Drop shadow
    shadow_off = 16
    d.rectangle([sx + shadow_off, sy + shadow_off, sx + new_w + shadow_off, sy + new_h + shadow_off],
                fill=(0, 0, 0))

    # Paper background (white-ish)
    d.rectangle([sx - 2, sy - 2, sx + new_w + 2, sy + new_h + 2], fill=(250, 250, 247), outline=LINE, width=2)

    img.paste(scaled, (sx, sy))
    d = ImageDraw.Draw(img)

    # Bottom: tagline strip
    d.line([(60, H - 80), (W - 60, H - 80)], fill=LINE, width=2)
    d.text((60, H - 60), "PROJECTCALC.APP", fill=HI_VIS, font=f(FN_BLACK, 24))
    text_right(d, f"{PRICE}  ·  INSTANT DOWNLOAD  ·  EXCEL · GOOGLE SHEETS · LIBREOFFICE",
               f(FN_MONO_BOLD, 22), INK_2, W - 60, H - 56)

    return img


def build_cover(screenshot: Image.Image) -> Image.Image:
    """Hero cover: compact title block on top + dominant screenshot below.

    Layout (top to bottom):
      0–80     brand bar
      80–620   title + tagline + bullets + price chip
      620–1820 BIG screenshot (fills width)
      1820–2000 footer
    """
    img = Image.new("RGB", (W, H), BG)
    d = ImageDraw.Draw(img)

    # Subtle grid
    step = 60
    for x in range(0, W, step):
        d.line([(x, 0), (x, H)], fill=(20, 20, 18), width=1)
    for y in range(0, H, step):
        d.line([(0, y), (W, y)], fill=(20, 20, 18), width=1)

    # Brand mark top-left
    s = 28
    bx, by = 80, 80
    d.polygon([(bx, by + s // 2), (bx + s // 2, by), (bx + s, by + s // 2), (bx + s // 2, by + s)], fill=HI_VIS)
    d.text((bx + s + 18, by - 4), "PROJECTCALC", fill=INK, font=f(FN_BLACK, 32))

    # PRO TOOLKIT v1 tag, top right
    text_right(d, "PRO TOOLKIT  ·  v1", f(FN_MONO_BOLD, 24), HI_VIS, W - 80, by - 2)

    # ── Title block (left side, compact) ──
    # Title size scales by trade-name length
    parts = TOOLKIT_NAME.split(" PRO TOOLKIT")
    trade = parts[0].strip() if parts else TOOLKIT_NAME

    if len(trade) >= 9:    # PLUMBING, ELECTRICAL
        trade_size = 110
    elif len(trade) >= 7:  # FRAMING, CONCRETE
        trade_size = 130
    else:                  # HVAC
        trade_size = 160

    title_y = 180
    d.text((80, title_y), trade, fill=INK, font=f(FN_BLACK, trade_size))
    second_y = title_y + int(trade_size * 1.05)
    d.text((80, second_y), "PRO TOOLKIT.", fill=HI_VIS, font=f(FN_BLACK, 86))

    # Tagline + price chip on the right
    tag_y = title_y + 30
    d.text((W - 720, tag_y), TAGLINE, fill=INK_2, font=f(FN_REG, 26))

    # Price chip — big, top-right
    chip_w, chip_h = 280, 150
    cx_chip = W - 80 - chip_w
    cy_chip = tag_y + 60
    d.rectangle([cx_chip, cy_chip, cx_chip + chip_w, cy_chip + chip_h], fill=HI_VIS)
    text_centered(d, PRICE, f(FN_BLACK, 92), BG, cx_chip + chip_w // 2, cy_chip + 8)
    text_centered(d, "INSTANT DOWNLOAD", f(FN_MONO_BOLD, 20), BG, cx_chip + chip_w // 2, cy_chip + 116)

    # ── Hero screenshot — fills middle band ──
    pad_x = 80
    hero_top = 640
    hero_bottom = 1820
    avail_w = W - pad_x * 2
    avail_h = hero_bottom - hero_top

    sw, sh = screenshot.size
    scale = min(avail_w / sw, avail_h / sh)
    new_w = int(sw * scale)
    new_h = int(sh * scale)
    scaled = screenshot.resize((new_w, new_h), Image.LANCZOS)

    sx = pad_x + (avail_w - new_w) // 2
    sy = hero_top + (avail_h - new_h) // 2

    # Caption above screenshot
    text_centered(d, "REAL WORKBOOK PREVIEW  ·  CUSTOMER QUOTE",
                  f(FN_MONO_BOLD, 24), HI_VIS, W // 2, hero_top - 50)

    # Drop shadow + paper frame
    d.rectangle([sx + 14, sy + 14, sx + new_w + 14, sy + new_h + 14], fill=(0, 0, 0))
    d.rectangle([sx - 2, sy - 2, sx + new_w + 2, sy + new_h + 2],
                fill=(250, 250, 247), outline=LINE, width=2)
    img.paste(scaled, (sx, sy))
    d = ImageDraw.Draw(img)

    # Footer
    y = 1880
    d.line([(80, y), (W - 80, y)], fill=LINE, width=2)
    d.text((80, y + 24), "PROJECTCALC.APP", fill=HI_VIS, font=f(FN_BLACK, 26))
    text_right(d, "EXCEL  ·  GOOGLE SHEETS  ·  LIBREOFFICE",
               f(FN_MONO_BOLD, 22), INK_3, W - 80, y + 28)

    return img


def main():
    xlsx = TOOLKIT_FOLDER / "dist" / XLSX_NAME
    print(f"Rendering sheets from {xlsx.name}")
    images = render_sheets_to_png(xlsx)

    # Build cover (uses COVER_SHEET screenshot)
    cover_img = build_cover(images[COVER_SHEET])
    cover_img.save(OUT / "01_cover.png", "PNG", optimize=True)
    print(f"  01_cover.png  (using {COVER_SHEET})")

    # Build sheet images
    for i, (sheet, caption) in enumerate(SHEETS_TO_FEATURE, start=2):
        if sheet not in images:
            print(f"  WARN: sheet {sheet!r} not found, skipping")
            continue
        out_name = f"0{i}_{sheet.lower().replace('-', '_')}.png"
        img = build_sheet_image(images[sheet], caption)
        img.save(OUT / out_name, "PNG", optimize=True)
        print(f"  {out_name}  ({sheet})")

    print(f"\nDone. Output: {OUT}")


if __name__ == "__main__":
    main()
