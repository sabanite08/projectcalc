"""Build the Framing Pro Toolkit workbook.

Run: python build.py
Output: dist/Framing-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Framing Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)
    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a wall, floor, and roof layout into "
         "a complete framing material takeoff — studs, plates, joists, "
         "rafters, headers, plywood, fasteners — plus a cut list and "
         "print-ready quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs.\n"
         "2. The TAKEOFF tab returns lumber count by member.\n"
         "3. CUT-LIST converts everything to standard 2× lengths and "
         "shows board feet by species.\n"
         "4. PLYWOOD computes sheathing + subfloor sheets with waste.\n"
         "5. PRICING — set unit costs once.\n"
         "6. Print QUOTE."),
        ("STUD SPACING + LAYOUT",
         "Default: 16\" OC studs, 24\" OC joists/rafters. Stud count = "
         "(linear ft × 12 / 16) + 1 + corner/T allowances + 1 cripple per "
         "rough opening."),
        ("HEADER QUICK-PICK (IRC R602.7)",
         "≤ 4 ft span: 2-2x4. 4–6 ft: 2-2x6. 6–8 ft: 2-2x8. 8–10 ft: "
         "2-2x10. 10–12 ft: 2-2x12. Built-up SPF #2, single-story, ground "
         "snow ≤ 50 psf. Engineer for outliers."),
        ("BEAM SPAN (SPF #2 BUILT-UP)",
         "L/360 deflection-controlled. The TAKEOFF tab includes a "
         "simplified built-up beam check; verify with the IRC tables for "
         "permitted construction."),
        ("PLYWOOD COVERAGE",
         "4×8 sheet = 32 sq ft. Default waste 12% on subfloor / 10% on "
         "wall sheathing. Roof sheathing waste depends on pitch — toolkit "
         "uses 12% as a starting point."),
        ("PRICING",
         "Unit prices on PRICING are 2026 ballpark — edit to your "
         "supplier. Labor is per board foot for assembly + per sheet for "
         "sheathing."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label); a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top"); a.fill = PANEL_FILL
        b = ws.cell(r, 2, body); b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Garage Addition"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "WALLS (linear ft + height)", span=3); r += 1
    walls = [
        ("Total exterior wall length (lin ft)", 120, "0"),
        ("Total interior wall length (lin ft)", 80, "0"),
        ("Wall height (ft)", 9, "0"),
        ("Stud spacing (in)", 16, "0"),
        ("Number of corners (3 studs each)", 4, "0"),
        ("Number of T-intersections (2 studs each)", 6, "0"),
        ("Rough openings (windows + doors)", 8, "0"),
        ("Avg header span (ft)", 4, "0"),
        ("Plates per wall (1 bottom + 2 top)", 3, "0"),
    ]
    walls_start = r
    for label, default, fmt in walls:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["ExtWallFt"] = DefinedName("ExtWallFt", attr_text=f"INPUTS!$B${walls_start}")
    wb.defined_names["IntWallFt"] = DefinedName("IntWallFt", attr_text=f"INPUTS!$B${walls_start+1}")
    wb.defined_names["WallHt"] = DefinedName("WallHt", attr_text=f"INPUTS!$B${walls_start+2}")
    wb.defined_names["StudOC"] = DefinedName("StudOC", attr_text=f"INPUTS!$B${walls_start+3}")
    wb.defined_names["Corners"] = DefinedName("Corners", attr_text=f"INPUTS!$B${walls_start+4}")
    wb.defined_names["Tees"] = DefinedName("Tees", attr_text=f"INPUTS!$B${walls_start+5}")
    wb.defined_names["Openings"] = DefinedName("Openings", attr_text=f"INPUTS!$B${walls_start+6}")
    wb.defined_names["AvgHeaderFt"] = DefinedName("AvgHeaderFt", attr_text=f"INPUTS!$B${walls_start+7}")
    wb.defined_names["PlateCount"] = DefinedName("PlateCount", attr_text=f"INPUTS!$B${walls_start+8}")

    r += 1
    section(ws, r, "FLOOR (joists)", span=3); r += 1
    flr = [
        ("Floor area (sq ft)", 720, "0"),
        ("Joist span (ft)", 14, "0"),
        ("Joist spacing (in)", 16, "0"),
        ("Joist size", "2x10", None),  # 2x8 | 2x10 | 2x12
        ("Rim joists (lin ft)", 96, "0"),
    ]
    flr_start = r
    for label, default, fmt in flr:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["FloorArea"] = DefinedName("FloorArea", attr_text=f"INPUTS!$B${flr_start}")
    wb.defined_names["JoistSpan"] = DefinedName("JoistSpan", attr_text=f"INPUTS!$B${flr_start+1}")
    wb.defined_names["JoistOC"] = DefinedName("JoistOC", attr_text=f"INPUTS!$B${flr_start+2}")
    wb.defined_names["JoistSize"] = DefinedName("JoistSize", attr_text=f"INPUTS!$B${flr_start+3}")
    wb.defined_names["RimFt"] = DefinedName("RimFt", attr_text=f"INPUTS!$B${flr_start+4}")

    r += 1
    section(ws, r, "ROOF (rafters)", span=3); r += 1
    roof = [
        ("Roof footprint (sq ft, plan view)", 800, "0"),
        ("Roof pitch (rise per 12)", 6, "0"),
        ("Rafter spacing (in)", 24, "0"),
        ("Rafter span (ft, ridge to wall)", 14, "0"),
        ("Rafter size", "2x10", None),
        ("Ridge length (lin ft)", 32, "0"),
        ("Overhang per side (ft)", 1.5, "0.0"),
        ("Roof type", "Gable", None),
    ]
    roof_start = r
    for label, default, fmt in roof:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["RoofPlan"] = DefinedName("RoofPlan", attr_text=f"INPUTS!$B${roof_start}")
    wb.defined_names["Pitch"] = DefinedName("Pitch", attr_text=f"INPUTS!$B${roof_start+1}")
    wb.defined_names["RafterOC"] = DefinedName("RafterOC", attr_text=f"INPUTS!$B${roof_start+2}")
    wb.defined_names["RafterSpan"] = DefinedName("RafterSpan", attr_text=f"INPUTS!$B${roof_start+3}")
    wb.defined_names["RafterSize"] = DefinedName("RafterSize", attr_text=f"INPUTS!$B${roof_start+4}")
    wb.defined_names["RidgeFt"] = DefinedName("RidgeFt", attr_text=f"INPUTS!$B${roof_start+5}")
    wb.defined_names["Overhang"] = DefinedName("Overhang", attr_text=f"INPUTS!$B${roof_start+6}")
    wb.defined_names["RoofType"] = DefinedName("RoofType", attr_text=f"INPUTS!$B${roof_start+7}")

    r += 1
    section(ws, r, "SHEATHING + WASTE", span=3); r += 1
    sh = [
        ("Wall sheathing (Y/N)", "Y", None),
        ("Wall sheathing thickness", "7/16 OSB", None),
        ("Roof sheathing thickness", "5/8 OSB", None),
        ("Subfloor thickness", "3/4 T&G", None),
        ("Wall sheathing waste %", 10, "0"),
        ("Roof sheathing waste %", 12, "0"),
        ("Subfloor waste %", 12, "0"),
    ]
    sh_start = r
    for label, default, fmt in sh:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["WallShOK"] = DefinedName("WallShOK", attr_text=f"INPUTS!$B${sh_start}")
    wb.defined_names["WallShTh"] = DefinedName("WallShTh", attr_text=f"INPUTS!$B${sh_start+1}")
    wb.defined_names["RoofShTh"] = DefinedName("RoofShTh", attr_text=f"INPUTS!$B${sh_start+2}")
    wb.defined_names["SubflTh"] = DefinedName("SubflTh", attr_text=f"INPUTS!$B${sh_start+3}")
    wb.defined_names["WallShWaste"] = DefinedName("WallShWaste", attr_text=f"INPUTS!$B${sh_start+4}")
    wb.defined_names["RoofShWaste"] = DefinedName("RoofShWaste", attr_text=f"INPUTS!$B${sh_start+5}")
    wb.defined_names["SubflWaste"] = DefinedName("SubflWaste", attr_text=f"INPUTS!$B${sh_start+6}")

    r += 1
    n = ws.cell(r, 1, "Joist/rafter sizes: 2x6 | 2x8 | 2x10 | 2x12"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Wall sheathing options: 7/16 OSB | 1/2 OSB | 1/2 CDX"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_takeoff(wb):
    ws = wb.create_sheet("TAKEOFF")
    auto_width(ws, [38, 14, 14, 36])
    banner(ws, "Lumber Takeoff", "Stud, plate, joist, rafter, header counts — pulled from INPUTS", width=4)

    r = 4
    section(ws, r, "WALL FRAMING", span=4); r += 1
    headers = ["Member", "Quantity", "Unit", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Stud count: (total wall ft × 12 / OC) + 1 + corner/T extras + 1 cripple per opening + 2 jacks per opening
    total_wall = "(ExtWallFt+IntWallFt)"
    items = [
        ("Studs (precut, wall height)",
         f'=ROUNDUP({total_wall}*12/StudOC+1+Corners*2+Tees+Openings*3,0)',
         "ea",
         "OC + 1 + 2/corner + 1/T + 3 (jacks+king+cripple) per RO"),
        ("Bottom plates (PT for ext, SPF for int)",
         f"=ROUNDUP((ExtWallFt+IntWallFt)/8,0)",
         "8' pieces",
         "1 plate per wall, sold in 8-ft sticks"),
        ("Top plates (double, both walls)",
         f"=ROUNDUP((ExtWallFt+IntWallFt)*2/8,0)",
         "8' pieces",
         "Double top, both ext + int — 2× linear ft / 8"),
        ("Headers (built-up 2x material)",
         f'=ROUNDUP(Openings*AvgHeaderFt*2/8,0)',
         "8' pieces",
         "2 plies per header × span / 8 ft length"),
        ("Window/door jack studs (2 per opening)",
         f"=Openings*2",
         "ea",
         "Already counted above; line included for invoice clarity"),
        ("Cripples (1 above, 1 below per opening)",
         f"=Openings*2",
         "ea",
         "Short studs above + below RO"),
    ]
    for label, formula, unit, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "FLOOR FRAMING", span=4); r += 1
    items = [
        ("Floor joists",
         "=ROUNDUP(FloorArea/JoistSpan*12/JoistOC+1,0)",
         "ea",
         "(area / span × 12 / OC) + 1; sized per JoistSize"),
        ("Rim joists",
         "=ROUNDUP(RimFt/12,0)",
         "12' pieces",
         "Sold in 12-ft sticks; matches joist depth"),
        ("Joist hangers (one per joist end + rim)",
         "=ROUNDUP(FloorArea/JoistSpan*12/JoistOC*2+1,0)",
         "ea",
         "Both ends if hung off rim or beam"),
        ("Sill seal (foam, lin ft)",
         "=RimFt",
         "lin ft",
         "Between sill plate and foundation"),
    ]
    for label, formula, unit, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "ROOF FRAMING", span=4); r += 1
    # Slope multiplier
    items = [
        ("Slope multiplier",
         "=SQRT(Pitch^2+144)/12",
         "×",
         "Multiply plan ft by this for actual rafter length"),
        ("True roof area (sq ft)",
         "=RoofPlan*(SQRT(Pitch^2+144)/12)",
         "sq ft",
         "Plan area × slope multiplier"),
        ("Rafters (each side)",
         "=ROUNDUP(RidgeFt*12/RafterOC+1,0)",
         "ea per side",
         "Per slope; ×2 for gable roof total"),
        ("Rafter total (gable, both sides)",
         "=ROUNDUP(RidgeFt*12/RafterOC+1,0)*2",
         "ea",
         "Gable roof; hip adjusts per geometry"),
        ("Ridge board",
         "=ROUNDUP(RidgeFt/12,0)",
         "12' pieces",
         "1× depth deeper than rafter (e.g. 2x12 for 2x10 rafters)"),
        ("Rafter ties / collar ties",
         "=ROUNDUP(RidgeFt*12/RafterOC/3,0)",
         "ea",
         "Every 3rd rafter pair, upper ⅓ of attic"),
        ("Hurricane ties (one per rafter end)",
         "=ROUNDUP(RidgeFt*12/RafterOC+1,0)*2",
         "ea",
         "Both sides; required in coastal/high-wind zones"),
    ]
    for label, formula, unit, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "0.000"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "FASTENERS", span=4); r += 1
    items = [
        ("16d framing nails (sinkers)",
         f"=ROUNDUP({total_wall}*0.5,0)",
         "lb",
         "≈0.5 lb per linear ft of wall framing"),
        ("8d sheathing nails",
         "=ROUNDUP((RoofPlan*(SQRT(Pitch^2+144)/12)+FloorArea+(ExtWallFt*WallHt))/100,0)",
         "lb",
         "≈1 lb per 100 sq ft of sheathing"),
        ("Construction adhesive (subfloor)",
         "=ROUNDUP(FloorArea/100,0)",
         "tubes",
         "1 tube per 100 sq ft of subfloor"),
        ("Joist hanger nails (10d × 1.5\")",
         "=ROUNDUP(FloorArea/JoistSpan*12/JoistOC*5/30,0)",
         "boxes",
         "5 nails per hanger; 30 nails per box"),
    ]
    for label, formula, unit, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    return ws


def sheet_cut_list(wb):
    ws = wb.create_sheet("CUT-LIST")
    auto_width(ws, [30, 14, 18, 18, 30])
    banner(ws, "Cut List + Board Feet", "Standard 2× lengths and species board-foot count", width=5)

    r = 4
    section(ws, r, "STANDARD LENGTHS BY MEMBER", span=5); r += 1
    headers = ["Member", "Size", "Std Length", "Each (BF)", "Note"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    rows = [
        ("Stud (precut 92-5/8\")", "2x4", "92-5/8\"",        "5.16",  "8' wall pre-cut"),
        ("Stud (precut 104-5/8\")", "2x4", "104-5/8\"",      "5.81",  "9' wall pre-cut"),
        ("Stud (precut 116-5/8\")", "2x4", "116-5/8\"",      "6.49",  "10' wall pre-cut"),
        ("Top/bottom plate", "2x4", "8 ft",                  "5.33",  "Sold in 8/10/12/16"),
        ("Header", "2x10", "8 ft",                           "13.33", "Built-up 2-ply or 3-ply"),
        ("Floor joist", "2x10", "16 ft",                     "26.67", "Common joist length"),
        ("Rim joist", "2x10", "12 ft",                       "20.00", ""),
        ("Rafter", "2x10", "16 ft",                          "26.67", "Cut to length per pitch"),
        ("Ridge board", "2x12", "12 ft",                     "24.00", "1× depth > rafter"),
    ]
    for member, size, length, bf, note in rows:
        ws.cell(r, 1, member).font = LBL; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, size).font = MONO_BOLD; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, length).font = MONO; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        ws.cell(r, 4, bf).font = MONO; ws.cell(r, 4).alignment = RIGHT; ws.cell(r, 4).border = BOX
        n = ws.cell(r, 5, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "BOARD-FOOT BY SPECIES (estimate)", span=5); r += 1
    # Approx total BF = walls (studs × 5.5) + floor joists × 26 + rafters × 26 + headers + plates
    total_wall = "(ExtWallFt+IntWallFt)"
    label_value(ws, r, "Total wall framing BF",
        f"=ROUND(({total_wall}*12/StudOC+1+Corners*2+Tees+Openings*3)*5.5+({total_wall}*3*5.33),0)",
        fmt="#,##0").font = MONO_BOLD
    r += 1
    label_value(ws, r, "Total floor framing BF",
        "=ROUND((FloorArea/JoistSpan*12/JoistOC+1)*26.67+RimFt/12*20,0)",
        fmt="#,##0").font = MONO_BOLD
    r += 1
    label_value(ws, r, "Total roof framing BF",
        "=ROUND((RidgeFt*12/RafterOC+1)*2*26.67+RidgeFt/12*24,0)",
        fmt="#,##0").font = MONO_BOLD
    r += 1
    label_value(ws, r, "GRAND TOTAL BF",
        "=ROUND((((ExtWallFt+IntWallFt)*12/StudOC+1+Corners*2+Tees+Openings*3)*5.5+((ExtWallFt+IntWallFt)*3*5.33))+((FloorArea/JoistSpan*12/JoistOC+1)*26.67+RimFt/12*20)+((RidgeFt*12/RafterOC+1)*2*26.67+RidgeFt/12*24),0)",
        fmt="#,##0").font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1
    return ws


def sheet_plywood(wb):
    ws = wb.create_sheet("PLYWOOD")
    auto_width(ws, [38, 18, 14, 30])
    banner(ws, "Plywood / OSB Sheathing", "Sub-floor + wall + roof sheets (4×8 = 32 sq ft)", width=4)

    r = 4
    headers = ["Application", "Area (sq ft)", "Sheets", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("Subfloor (3/4 T&G)",
         "=FloorArea",
         "=ROUNDUP(FloorArea/32*(1+SubflWaste/100),0)",
         "12% waste; T&G saves blocking"),
        ("Wall sheathing (7/16 OSB)",
         '=IF(WallShOK="Y",ExtWallFt*WallHt,0)',
         '=IF(WallShOK="Y",ROUNDUP(ExtWallFt*WallHt/32*(1+WallShWaste/100),0),0)',
         "10% waste; full sheets at corners"),
        ("Roof sheathing (5/8 OSB)",
         "=RoofPlan*(SQRT(Pitch^2+144)/12)",
         "=ROUNDUP(RoofPlan*(SQRT(Pitch^2+144)/12)/32*(1+RoofShWaste/100),0)",
         "12% waste; pitch-adjusted area"),
    ]
    for label, area, sheets, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        a = ws.cell(r, 2, area); a.font = MONO; a.alignment = RIGHT; a.number_format = "#,##0"; a.border = BOX
        s = ws.cell(r, 3, sheets); s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "#,##0"; s.border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    # Total sheets
    label_value(ws, r, "TOTAL SHEETS",
        '=ROUNDUP(FloorArea/32*(1+SubflWaste/100),0)+IF(WallShOK="Y",ROUNDUP(ExtWallFt*WallHt/32*(1+WallShWaste/100),0),0)+ROUNDUP(RoofPlan*(SQRT(Pitch^2+144)/12)/32*(1+RoofShWaste/100),0)',
        fmt="#,##0").font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "LUMBER (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("Stud 2x4 SPF (precut 92-5/8\")",  4.85,  "ea"),
        ("Stud 2x4 SPF (precut 104-5/8\")", 5.95,  "ea"),
        ("Stud 2x6 SPF (precut)",            8.50,  "ea"),
        ("2x4 SPF 8-ft",                     5.20,  "stick"),
        ("2x4 SPF 12-ft",                    7.80,  "stick"),
        ("2x4 PT (ground contact) 8-ft",     8.40,  "stick"),
        ("2x6 SPF 8-ft",                     8.95,  "stick"),
        ("2x8 SPF 12-ft",                    16.50, "stick"),
        ("2x10 SPF 12-ft",                   22.00, "stick"),
        ("2x10 SPF 16-ft",                   29.50, "stick"),
        ("2x12 SPF 12-ft",                   28.00, "stick"),
        ("LVL 1-3/4\" × 11-7/8\"",           8.50,  "lin ft"),
        ("Joist hanger 2x10 (Simpson LUS)",  4.20,  "ea"),
        ("Hurricane tie (Simpson H2.5A)",    1.80,  "ea"),
        ("Sill seal foam (50 ft roll)",      14.00, "roll"),
        ("16d sinker nails (50 lb box)",     58.00, "box"),
        ("8d sheathing nails (50 lb box)",   62.00, "box"),
        ("Construction adhesive (29 oz)",    8.50,  "tube"),
        ("Joist hanger nails (1 lb box)",    9.00,  "box"),
        ("OSB 7/16\" 4×8",                   18.50, "sheet"),
        ("OSB 5/8\" 4×8 (roof)",             24.00, "sheet"),
        ("Subfloor 3/4\" T&G 4×8",           42.00, "sheet"),
        ("ZIP System 7/16 wall (4×8)",       38.00, "sheet"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceTable"] = DefinedName("PriceTable", attr_text=f"PRICING!$A${pricing_start}:$C${r-1}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Lead carpenter labor",       85.00, "per hour"),
        ("Helper labor",                50.00, "per hour"),
        ("Wall framing labor (per BF)", 0.85,  "per BF"),
        ("Floor framing labor (per BF)",0.65,  "per BF"),
        ("Roof framing labor (per BF)", 1.05,  "per BF"),
        ("Sheathing labor (per sheet)", 8.50,  "per sheet"),
        ("Permit + inspection",         185.00,"flat"),
        ("Dump fee + cleanup",          250.00,"flat"),
        ("Profit margin (decimal)",     0.25,  "0.25 = 25%"),
        ("Sales tax (decimal)",         0.07,  "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit:
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LeadRate"] = DefinedName("LeadRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["HelperRate"] = DefinedName("HelperRate", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["WallBF"] = DefinedName("WallBF", attr_text=f"PRICING!$B${labor_start+2}")
    wb.defined_names["FloorBF"] = DefinedName("FloorBF", attr_text=f"PRICING!$B${labor_start+3}")
    wb.defined_names["RoofBF"] = DefinedName("RoofBF", attr_text=f"PRICING!$B${labor_start+4}")
    wb.defined_names["SheathLab"] = DefinedName("SheathLab", attr_text=f"PRICING!$B${labor_start+5}")
    wb.defined_names["Permit"] = DefinedName("Permit", attr_text=f"PRICING!$B${labor_start+6}")
    wb.defined_names["DumpFee"] = DefinedName("DumpFee", attr_text=f"PRICING!$B${labor_start+7}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+8}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+9}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Floor area",   "=FloorArea&\" sq ft\""); r += 1
    label_value(ws, r, "Wall lin ft",  "=ExtWallFt+IntWallFt&\" lf\""); r += 1
    label_value(ws, r, "Roof plan",    "=RoofPlan&\" sq ft @ pitch \"&Pitch&\"/12\""); r += 1

    r += 1
    section(ws, r, "MATERIALS (typical garage / addition)", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    eq_start = r
    eq_items = [
        ("Studs 2x4 (precut, wall ht)", "=ROUNDUP((ExtWallFt+IntWallFt)*12/StudOC+1+Corners*2+Tees+Openings*3,0)",
         "Stud 2x4 SPF (precut 104-5/8\")"),
        ("Plates 2x4 8-ft (PT bottom + SPF top)", "=ROUNDUP((ExtWallFt+IntWallFt)*3/8,0)",
         "2x4 SPF 8-ft"),
        ("Headers (2x10 8-ft)", "=ROUNDUP(Openings*AvgHeaderFt*2/8,0)",
         "2x10 SPF 12-ft"),
        ("Floor joists (2x10 16-ft)", "=ROUNDUP(FloorArea/JoistSpan*12/JoistOC+1,0)",
         "2x10 SPF 16-ft"),
        ("Rim joists (2x10 12-ft)", "=ROUNDUP(RimFt/12,0)",
         "2x10 SPF 12-ft"),
        ("Rafters (2x10 16-ft)", "=ROUNDUP(RidgeFt*12/RafterOC+1,0)*2",
         "2x10 SPF 16-ft"),
        ("Ridge board (2x12 12-ft)", "=ROUNDUP(RidgeFt/12,0)",
         "2x12 SPF 12-ft"),
        ("Joist hangers", "=ROUNDUP(FloorArea/JoistSpan*12/JoistOC*2+1,0)",
         "Joist hanger 2x10 (Simpson LUS)"),
        ("Hurricane ties", "=ROUNDUP(RidgeFt*12/RafterOC+1,0)*2",
         "Hurricane tie (Simpson H2.5A)"),
        ("Sill seal", "=ROUNDUP(RimFt/50,0)",
         "Sill seal foam (50 ft roll)"),
        ("Subfloor 3/4 T&G", "=ROUNDUP(FloorArea/32*1.12,0)",
         "Subfloor 3/4\" T&G 4×8"),
        ("Wall sheathing 7/16 OSB", '=IF(WallShOK="Y",ROUNDUP(ExtWallFt*WallHt/32*1.10,0),0)',
         "OSB 7/16\" 4×8"),
        ("Roof sheathing 5/8 OSB", "=ROUNDUP(RoofPlan*(SQRT(Pitch^2+144)/12)/32*1.12,0)",
         "OSB 5/8\" 4×8 (roof)"),
        ("16d sinkers (50 lb box)", "1",
         "16d sinker nails (50 lb box)"),
        ("8d sheathing (50 lb box)", "1",
         "8d sheathing nails (50 lb box)"),
        ("Construction adhesive", "=ROUNDUP(FloorArea/100,0)",
         "Construction adhesive (29 oz)"),
    ]
    for desc, qty, price_label in eq_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        escaped = price_label.replace('"', '""')
        p = ws.cell(r, 3, f'=VLOOKUP("{escaped}",PriceTable,2,FALSE)')
        p.font = MONO; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    eq_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{eq_start}:D{eq_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    eq_subtotal_row = r; r += 1

    r += 1
    section(ws, r, "LABOR & FEES", span=4); r += 1
    labor_lines = [
        ("Wall framing labor (per BF)",
         "=(((ExtWallFt+IntWallFt)*12/StudOC+1+Corners*2+Tees+Openings*3)*5.5+((ExtWallFt+IntWallFt)*3*5.33))*WallBF"),
        ("Floor framing labor (per BF)",
         "=((FloorArea/JoistSpan*12/JoistOC+1)*26.67+RimFt/12*20)*FloorBF"),
        ("Roof framing labor (per BF)",
         "=((RidgeFt*12/RafterOC+1)*2*26.67+RidgeFt/12*24)*RoofBF"),
        ("Sheathing labor (per sheet)",
         '=(ROUNDUP(FloorArea/32*1.12,0)+IF(WallShOK="Y",ROUNDUP(ExtWallFt*WallHt/32*1.10,0),0)+ROUNDUP(RoofPlan*(SQRT(Pitch^2+144)/12)/32*1.12,0))*SheathLab'),
        ("Permit + inspection", "=Permit"),
        ("Dump fee + cleanup",  "=DumpFee"),
    ]
    labor_start = r
    for desc, formula in labor_lines:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        for col in (2, 3): ws.cell(r, col, "").border = BOX
        t = ws.cell(r, 4, formula); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    labor_end = r - 1

    ws.cell(r, 1, "Labor subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{labor_start}:D{labor_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    labor_subtotal_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{eq_subtotal_row}+D{labor_subtotal_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Lumber prices subject to market change. "
        "Work scheduled within 3-4 weeks of signed acceptance, weather permitting.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_takeoff(wb)
    sheet_cut_list(wb)
    sheet_plywood(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "TAKEOFF", "CUT-LIST", "PLYWOOD", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Framing-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
