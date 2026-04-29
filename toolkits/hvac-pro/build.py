"""Build the HVAC Pro Toolkit workbook.

Run: python build.py
Output: dist/HVAC-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

# Brand
HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "HVAC Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a house's geometry, envelope, and "
         "climate into a Manual J load, duct CFM allocation, ASHRAE 62.2 "
         "ventilation check, and a print-ready customer quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs — fill them in. "
         "Gray cells are calculated.\n"
         "2. The LOAD tab returns whole-house heating + cooling BTU/hr and "
         "the equipment ton/BTU you need to install.\n"
         "3. The DUCT tab takes that load and tells you total CFM, per-room "
         "CFM, and the round/rectangular duct sizes that hit it.\n"
         "4. The VENTILATION tab confirms ASHRAE 62.2 whole-house + spot "
         "exhaust requirements.\n"
         "5. The QUOTE tab is print-ready — update prices in PRICING first."),
        ("MANUAL J HTM (SIMPLIFIED)",
         "Heating load = sum of (U × area × ΔT) for every envelope "
         "surface (walls, ceiling, floor, windows, doors) plus infiltration "
         "(0.018 × ACH × volume × ΔT).\n"
         "Cooling load = conduction (ΔT) + window solar gain (area × SHGC "
         "× 100) + internal gains (230 BTU/h per occupant + 1,200 BTU/h "
         "kitchen) + infiltration.\n"
         "This is a simplified ACCA Manual J estimate — use a full Manual "
         "J report for permitted installs."),
        ("DUCT SIZING",
         "Total supply CFM defaults to 400 CFM/ton (cooling-driven). The "
         "DUCT tab allocates CFM per room by area share, then looks up the "
         "round duct size from the friction-rate table (0.10 in WC/100 ft, "
         "the standard residential default)."),
        ("VENTILATION (ASHRAE 62.2-2022)",
         "Whole-house Q_total = 0.03 × floor area + 7.5 × (Nbr + 1). "
         "Add local exhaust: 100 CFM kitchen, 50 CFM bath (intermittent) "
         "or 25 CFM bath (continuous)."),
        ("PRICING",
         "Equipment + labor unit costs on PRICING are 2026 ballpark — edit "
         "to your supplier and pay rate. Margin compounds on (material + "
         "labor)."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please don't "
         "redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Residence System Upgrade"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "HOUSE GEOMETRY", span=3); r += 1
    geom = [
        ("Conditioned floor area (sq ft)", 1800, "0"),
        ("Average ceiling height (ft)", 9, "0.0"),
        ("Number of bedrooms", 3, "0"),
        ("Number of occupants", 4, "0"),
    ]
    geom_start = r
    for label, default, fmt in geom:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["FloorArea"] = DefinedName("FloorArea", attr_text=f"INPUTS!$B${geom_start}")
    wb.defined_names["CeilHeight"] = DefinedName("CeilHeight", attr_text=f"INPUTS!$B${geom_start+1}")
    wb.defined_names["Nbr"] = DefinedName("Nbr", attr_text=f"INPUTS!$B${geom_start+2}")
    wb.defined_names["Occupants"] = DefinedName("Occupants", attr_text=f"INPUTS!$B${geom_start+3}")

    r += 1
    section(ws, r, "ENVELOPE (gross areas)", span=3); r += 1
    env = [
        ("Above-grade wall area (sq ft)", 1600, "0"),
        ("Wall R-value (h·ft²·°F/Btu)", 19, "0.0"),
        ("Window area (sq ft)", 280, "0"),
        ("Window U-factor (Btu/h·ft²·°F)", 0.30, "0.00"),
        ("Window SHGC (solar heat gain)", 0.30, "0.00"),
        ("Door area (sq ft)", 40, "0"),
        ("Door U-factor", 0.40, "0.00"),
        ("Ceiling area (sq ft)", 1800, "0"),
        ("Ceiling R-value", 38, "0.0"),
        ("Floor area exposed (sq ft)", 0, "0"),
        ("Floor R-value (slab/crawl)", 13, "0.0"),
        ("Infiltration ACH (natural)", 0.35, "0.00"),
    ]
    env_start = r
    for label, default, fmt in env:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["WallArea"] = DefinedName("WallArea", attr_text=f"INPUTS!$B${env_start}")
    wb.defined_names["WallR"] = DefinedName("WallR", attr_text=f"INPUTS!$B${env_start+1}")
    wb.defined_names["WindowArea"] = DefinedName("WindowArea", attr_text=f"INPUTS!$B${env_start+2}")
    wb.defined_names["WindowU"] = DefinedName("WindowU", attr_text=f"INPUTS!$B${env_start+3}")
    wb.defined_names["WindowSHGC"] = DefinedName("WindowSHGC", attr_text=f"INPUTS!$B${env_start+4}")
    wb.defined_names["DoorArea"] = DefinedName("DoorArea", attr_text=f"INPUTS!$B${env_start+5}")
    wb.defined_names["DoorU"] = DefinedName("DoorU", attr_text=f"INPUTS!$B${env_start+6}")
    wb.defined_names["CeilArea"] = DefinedName("CeilArea", attr_text=f"INPUTS!$B${env_start+7}")
    wb.defined_names["CeilR"] = DefinedName("CeilR", attr_text=f"INPUTS!$B${env_start+8}")
    wb.defined_names["FloorAreaExp"] = DefinedName("FloorAreaExp", attr_text=f"INPUTS!$B${env_start+9}")
    wb.defined_names["FloorR"] = DefinedName("FloorR", attr_text=f"INPUTS!$B${env_start+10}")
    wb.defined_names["ACH"] = DefinedName("ACH", attr_text=f"INPUTS!$B${env_start+11}")

    r += 1
    section(ws, r, "CLIMATE & SETPOINTS", span=3); r += 1
    clim = [
        ("Heating outdoor design temp (°F, 99%)", 10, "0"),
        ("Cooling outdoor design temp (°F, 1%)", 92, "0"),
        ("Indoor heating setpoint (°F)", 70, "0"),
        ("Indoor cooling setpoint (°F)", 75, "0"),
    ]
    clim_start = r
    for label, default, fmt in clim:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["HeatODT"] = DefinedName("HeatODT", attr_text=f"INPUTS!$B${clim_start}")
    wb.defined_names["CoolODT"] = DefinedName("CoolODT", attr_text=f"INPUTS!$B${clim_start+1}")
    wb.defined_names["HeatSet"] = DefinedName("HeatSet", attr_text=f"INPUTS!$B${clim_start+2}")
    wb.defined_names["CoolSet"] = DefinedName("CoolSet", attr_text=f"INPUTS!$B${clim_start+3}")

    r += 1
    section(ws, r, "EQUIPMENT & SCOPE", span=3); r += 1
    eq = [
        ("System type", "Split Heat Pump", None),  # Split Heat Pump | Furnace+AC | Mini-split
        ("Refrigerant", "R-410A", None),  # R-410A | R-32 | R-454B
        ("Existing ductwork (Y/N)", "N", None),
        ("Linear ft of new line set", 30, "0"),
        ("Filter MERV rating", 11, "0"),
        ("Static pressure budget (in WC)", 0.50, "0.00"),
    ]
    eq_start = r
    for label, default, fmt in eq:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["SystemType"] = DefinedName("SystemType", attr_text=f"INPUTS!$B${eq_start}")
    wb.defined_names["Refrigerant"] = DefinedName("Refrigerant", attr_text=f"INPUTS!$B${eq_start+1}")
    wb.defined_names["ExistingDuct"] = DefinedName("ExistingDuct", attr_text=f"INPUTS!$B${eq_start+2}")
    wb.defined_names["LineSet"] = DefinedName("LineSet", attr_text=f"INPUTS!$B${eq_start+3}")
    wb.defined_names["MERV"] = DefinedName("MERV", attr_text=f"INPUTS!$B${eq_start+4}")
    wb.defined_names["StaticBudget"] = DefinedName("StaticBudget", attr_text=f"INPUTS!$B${eq_start+5}")

    r += 1
    n = ws.cell(r, 1, "System options: Split Heat Pump | Furnace+AC | Mini-split"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Refrigerant options: R-410A | R-32 | R-454B"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1,
        "ACH guidance: tight new build 0.20 · average existing 0.35 · leaky/older 0.60+")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_load(wb):
    ws = wb.create_sheet("LOAD")
    auto_width(ws, [42, 18, 14, 32])
    banner(ws, "Manual J Load (Simplified HTM)", "Heating + cooling BTU/hr, equipment sizing", width=4)

    r = 4
    section(ws, r, "DESIGN TEMPERATURE DELTAS", span=4); r += 1
    label_value(ws, r, "Heating ΔT (°F)", "=HeatSet-HeatODT", fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Cooling ΔT (°F)", "=CoolODT-CoolSet", fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "House volume (cu ft)", "=FloorArea*CeilHeight", fmt="#,##0").font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "HEATING LOAD (BTU/hr)", span=4); r += 1
    headers = ["Component", "Q = U · A · ΔT", "Btu/h", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Each row: label / formula text / value formula / note
    heat_rows = [
        ("Walls", "(WallArea/WallR) × ΔT",
         "=WallArea/WallR*(HeatSet-HeatODT)", "U = 1/R, net of windows/doors not subtracted"),
        ("Windows", "WindowArea × U × ΔT",
         "=WindowArea*WindowU*(HeatSet-HeatODT)", "U-factor includes frame + glazing"),
        ("Doors", "DoorArea × U × ΔT",
         "=DoorArea*DoorU*(HeatSet-HeatODT)", "Insulated steel/fiberglass U ≈ 0.20–0.40"),
        ("Ceiling / Roof", "(CeilArea/CeilR) × ΔT",
         "=CeilArea/CeilR*(HeatSet-HeatODT)", "Attic R38 typical for IECC zone 5"),
        ("Floor (exposed)", "(FloorAreaExp/FloorR) × ΔT",
         "=FloorAreaExp/FloorR*(HeatSet-HeatODT)", "Slab perimeter / cantilever / crawl"),
        ("Infiltration", "0.018 × ACH × Vol × ΔT",
         "=0.018*ACH*FloorArea*CeilHeight*(HeatSet-HeatODT)", "0.018 = air heat capacity factor"),
    ]
    heat_start = r
    for label, formula_txt, formula, note in heat_rows:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, formula_txt).font = MONO; ws.cell(r, 2).alignment = LEFT; ws.cell(r, 2).border = BOX
        c = ws.cell(r, 3, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1
    heat_end = r - 1

    # Heating subtotal + safety factor
    ws.cell(r, 1, "Heating subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    ws.cell(r, 2, "").border = BOX; ws.cell(r, 2).fill = PANEL_FILL
    s = ws.cell(r, 3, f"=SUM(C{heat_start}:C{heat_end})"); s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "#,##0"; s.border = BOX; s.fill = PANEL_FILL
    ws.cell(r, 4, "").border = BOX; ws.cell(r, 4).fill = PANEL_FILL
    heat_sub_row = r; r += 1

    ws.cell(r, 1, "+ 10% safety / duct loss").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    ws.cell(r, 2, "").border = BOX
    c = ws.cell(r, 3, f"=C{heat_sub_row}*1.10"); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX; c.fill = YELLOW
    ws.cell(r, 4, "Design heating BTU/hr").font = SMALL; ws.cell(r, 4).border = BOX
    heat_design_row = r; r += 1

    r += 1
    section(ws, r, "COOLING LOAD (BTU/hr)", span=4); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    cool_rows = [
        ("Walls (conduction)", "(WallArea/WallR) × ΔT",
         "=WallArea/WallR*(CoolODT-CoolSet)", "Sensible only"),
        ("Windows (conduction)", "WindowArea × U × ΔT",
         "=WindowArea*WindowU*(CoolODT-CoolSet)", ""),
        ("Windows (solar gain)", "WindowArea × SHGC × 100",
         "=WindowArea*WindowSHGC*100", "Avg N/E/S/W mix; assumes typical shading"),
        ("Doors (conduction)", "DoorArea × U × ΔT",
         "=DoorArea*DoorU*(CoolODT-CoolSet)", ""),
        ("Ceiling / Roof", "(CeilArea/CeilR) × ΔT",
         "=CeilArea/CeilR*(CoolODT-CoolSet)", "Add 20% if uninsulated attic over"),
        ("Floor (exposed)", "(FloorAreaExp/FloorR) × ΔT",
         "=FloorAreaExp/FloorR*(CoolODT-CoolSet)", ""),
        ("Infiltration (sensible)", "0.018 × ACH × Vol × ΔT",
         "=0.018*ACH*FloorArea*CeilHeight*(CoolODT-CoolSet)", ""),
        ("Internal: occupants", "230 × occupants",
         "=230*Occupants", "ASHRAE sensible, seated office work"),
        ("Internal: kitchen/appliances", "1,200 (flat)",
         "=1200", "Range, oven, refrigerator, lighting"),
    ]
    cool_start = r
    for label, formula_txt, formula, note in cool_rows:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, formula_txt).font = MONO; ws.cell(r, 2).alignment = LEFT; ws.cell(r, 2).border = BOX
        c = ws.cell(r, 3, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1
    cool_end = r - 1

    ws.cell(r, 1, "Cooling subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    ws.cell(r, 2, "").border = BOX; ws.cell(r, 2).fill = PANEL_FILL
    s = ws.cell(r, 3, f"=SUM(C{cool_start}:C{cool_end})"); s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "#,##0"; s.border = BOX; s.fill = PANEL_FILL
    ws.cell(r, 4, "").border = BOX; ws.cell(r, 4).fill = PANEL_FILL
    cool_sub_row = r; r += 1

    ws.cell(r, 1, "+ 15% latent (1.15 × sensible)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    ws.cell(r, 2, "").border = BOX
    c = ws.cell(r, 3, f"=C{cool_sub_row}*1.15"); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX; c.fill = YELLOW
    ws.cell(r, 4, "Design cooling BTU/hr").font = SMALL; ws.cell(r, 4).border = BOX
    cool_design_row = r; r += 1

    r += 1
    section(ws, r, "EQUIPMENT SIZING", span=4); r += 1
    label_value(ws, r, "Cooling tons (round to 0.5 ton)",
        f"=CEILING(C{cool_design_row}/12000,0.5)", fmt="0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Heating BTU/hr (size to design)",
        f"=ROUND(C{heat_design_row},0)", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "BTU/sq ft (cooling check)",
        f"=C{cool_design_row}/FloorArea", fmt="0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Sq ft per ton (cooling check)",
        f"=FloorArea/(C{cool_design_row}/12000)", fmt="0").font = MONO_BOLD; r += 1

    # Named ranges so DUCT + QUOTE can reference design loads
    wb.defined_names["DesignHeatBTU"] = DefinedName("DesignHeatBTU", attr_text=f"LOAD!$C${heat_design_row}")
    wb.defined_names["DesignCoolBTU"] = DefinedName("DesignCoolBTU", attr_text=f"LOAD!$C${cool_design_row}")

    note = ws.cell(r + 1, 1,
        "ESTIMATE ONLY — simplified ACCA Manual J HTM. For permitted "
        "installs, run a full Manual J/S/D using ACCA-approved software.")
    note.font = SMALL
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=4)
    return ws


def sheet_duct(wb):
    ws = wb.create_sheet("DUCT")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Duct CFM + Static Pressure", "400 CFM/ton split across rooms; size by 0.10 in WC/100 ft", width=5)

    r = 4
    section(ws, r, "TOTAL AIRFLOW", span=5); r += 1
    label_value(ws, r, "Cooling tons (from LOAD)",
        "=CEILING(DesignCoolBTU/12000,0.5)", fmt="0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Total supply CFM (400 CFM/ton)",
        "=CEILING(DesignCoolBTU/12000,0.5)*400", fmt="#,##0").font = MONO_BOLD
    cfm_total_row = r; r += 1
    label_value(ws, r, "Return CFM (≈ total supply)",
        f"=$B${cfm_total_row}", fmt="#,##0").font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "ROUND DUCT SIZE BY CFM (0.10\" WC / 100 ft)", span=5); r += 1
    headers = ["CFM", "Round (in)", "Rect (in)", "Velocity (FPM)", ""]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    duct_table = [
        (50, 4, "6×4", 575),
        (75, 5, "8×4", 550),
        (100, 6, "8×4", 510),
        (150, 7, "10×4", 560),
        (200, 8, "10×6", 575),
        (300, 9, "12×6", 680),
        (400, 10, "12×6", 730),
        (500, 12, "14×8", 640),
        (700, 14, "16×8", 655),
        (900, 16, "18×10", 645),
        (1200, 18, "20×10", 680),
        (1600, 20, "24×12", 670),
        (2000, 22, "26×14", 660),
    ]
    table_start = r
    for cfm, rd, rect, fpm in duct_table:
        ws.cell(r, 1, cfm).font = MONO_BOLD; ws.cell(r, 1).alignment = RIGHT; ws.cell(r, 1).number_format = "#,##0"; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, rd).font = MONO; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, rect).font = MONO; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        ws.cell(r, 4, fpm).font = MONO; ws.cell(r, 4).alignment = RIGHT; ws.cell(r, 4).number_format = "#,##0"; ws.cell(r, 4).border = BOX
        ws.cell(r, 5, "").border = BOX
        r += 1
    table_end = r - 1
    wb.defined_names["DuctTable"] = DefinedName("DuctTable", attr_text=f"DUCT!$A${table_start}:$D${table_end}")

    r += 1
    section(ws, r, "PER-ROOM CFM ALLOCATION (yellow = enter sq ft)", span=5); r += 1
    headers2 = ["Room", "Sq Ft", "% of Total", "CFM", "Round Duct (in)"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    rooms = [
        ("Living Room", 320),
        ("Kitchen", 200),
        ("Dining", 150),
        ("Master Bedroom", 250),
        ("Bedroom 2", 150),
        ("Bedroom 3", 130),
        ("Master Bath", 90),
        ("Bath 2", 60),
        ("Hallway / Other", 100),
    ]
    room_start = r
    for name, sqft in rooms:
        ws.cell(r, 1, name).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, sqft); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.fill = YELLOW; c.border = BOX
        # % of total room area entered
        p = ws.cell(r, 3, f"=B{r}/SUM($B${room_start}:$B${room_start+len(rooms)-1})")
        p.font = MONO; p.alignment = RIGHT; p.number_format = "0.0%"; p.border = BOX
        # CFM = % * total CFM
        cf = ws.cell(r, 4, f"=ROUND(C{r}*$B${cfm_total_row},0)")
        cf.font = MONO_BOLD; cf.alignment = RIGHT; cf.number_format = "#,##0"; cf.border = BOX
        # Round duct: smallest table size that meets or exceeds the room CFM
        rd = ws.cell(r, 5,
            f'=IF(D{r}<=50,4,IF(D{r}<=75,5,IF(D{r}<=100,6,IF(D{r}<=150,7,'
            f'IF(D{r}<=200,8,IF(D{r}<=300,9,IF(D{r}<=400,10,IF(D{r}<=500,12,'
            f'IF(D{r}<=700,14,IF(D{r}<=900,16,IF(D{r}<=1200,18,IF(D{r}<=1600,20,22))))))))))))')
        rd.font = MONO_BOLD; rd.alignment = CENTER; rd.border = BOX
        r += 1

    r += 1
    section(ws, r, "STATIC PRESSURE BUDGET (TESP)", span=5); r += 1
    sp_items = [
        ("Filter (MERV " "11-13)", 0.20),
        ("Cooling coil (wet)", 0.20),
        ("Supply duct + boots", 0.10),
        ("Return duct + grille", 0.10),
        ("Accessories (UV, dehumid)", 0.00),
    ]
    sp_start = r
    for label, val in sp_items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, val); c.font = MONO; c.alignment = RIGHT; c.number_format = "0.00"; c.fill = YELLOW; c.border = BOX
        for col in (3, 4, 5):
            ws.cell(r, col, "").border = BOX
        r += 1
    sp_end = r - 1
    ws.cell(r, 1, "TESP total (in WC)").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    s = ws.cell(r, 2, f"=SUM(B{sp_start}:B{sp_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0.00"; s.border = BOX; s.fill = YELLOW
    ws.cell(r, 3, "vs. blower rating").font = SMALL; ws.cell(r, 3).border = BOX
    chk = ws.cell(r, 4, f'=IF(B{r}<=StaticBudget,"OK","OVER")'); chk.font = MONO_BOLD; chk.alignment = CENTER; chk.border = BOX
    ws.cell(r, 5, "").border = BOX
    r += 1

    note = ws.cell(r + 1, 1,
        "Static pressure rule of thumb: keep TESP below the blower's rated max "
        "(typically 0.50\" WC for residential PSC, 0.80\" for ECM). Filter and "
        "coil values are wet-clean; size up duct or add a return if OVER.")
    note.font = SMALL
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=5)
    return ws


def sheet_ventilation(wb):
    ws = wb.create_sheet("VENTILATION")
    auto_width(ws, [36, 18, 32])
    banner(ws, "Ventilation (ASHRAE 62.2-2022)", "Whole-house + local exhaust requirements", width=3)

    r = 4
    section(ws, r, "WHOLE-HOUSE VENTILATION (Q_total)", span=3); r += 1
    label_value(ws, r, "Floor area (sq ft)", "=FloorArea", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Bedrooms", "=Nbr", fmt="0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Q_total = 0.03 × A + 7.5 × (Nbr+1)",
        "=ROUND(0.03*FloorArea+7.5*(Nbr+1),0)", fmt="0").font = MONO_BOLD
    qtot_row = r; r += 1
    label_value(ws, r, "Infiltration credit (Q_inf)",
        "=ROUND((ACH*FloorArea*CeilHeight/60)*0.5,0)", fmt="0").font = MONO_BOLD
    qinf_row = r; r += 1
    label_value(ws, r, "Mechanical fan needed (CFM)",
        f"=MAX(0,B{qtot_row}-B{qinf_row})", fmt="0").font = MONO_BOLD; r += 1

    r += 1
    section(ws, r, "LOCAL EXHAUST (per room)", span=3); r += 1
    rows = [
        ("Kitchen range hood (intermittent)", 100, "CFM minimum, vented outdoors"),
        ("Kitchen continuous", 25, "if no range hood / continuous fan"),
        ("Bath fan, intermittent", 50, "per bath — 5 min/use minimum"),
        ("Bath fan, continuous", 25, "if no intermittent fan"),
        ("Dryer makeup air", 0, "code: required if vented dryer + tight envelope"),
    ]
    headers = ["Source", "CFM", "Note"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    for label, cfm, note in rows:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, cfm); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        n = ws.cell(r, 3, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    note = ws.cell(r, 1,
        "Whole-house ventilation must run continuously (or operate enough hours "
        "to deliver Q_total over 24 h). HRV/ERV is preferred in heating-dominated "
        "climates; balanced exhaust-only is acceptable per 62.2 with proper "
        "makeup air strategy.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=3)
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "EQUIPMENT (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("Heat pump condenser, 2 ton (15 SEER2)", 2400.00, "ea"),
        ("Heat pump condenser, 3 ton (15 SEER2)", 2900.00, "ea"),
        ("Heat pump condenser, 4 ton (15 SEER2)", 3400.00, "ea"),
        ("Air handler with electric heat strips", 1800.00, "ea"),
        ("Gas furnace, 80 kBTU 96% AFUE", 2200.00, "ea"),
        ("AC condenser, 3 ton (14.3 SEER2)", 2100.00, "ea"),
        ("Mini-split single-zone, 18k BTU", 1900.00, "ea"),
        ("Refrigerant line set (3/8 + 7/8 insulated)", 12.00, "lin ft"),
        ("Refrigerant R-410A", 28.00, "lb"),
        ("Refrigerant R-32 / R-454B", 45.00, "lb"),
        ("Round flex duct, R-8 insulated", 4.50, "lin ft"),
        ("Rectangular sheet metal duct", 14.00, "lin ft"),
        ("Supply register / boot", 28.00, "ea"),
        ("Return air grille", 32.00, "ea"),
        ("Smart thermostat", 240.00, "ea"),
        ("MERV 11 filter cabinet", 165.00, "ea"),
        ("Condensate pump", 95.00, "ea"),
        ("Disconnect + whip", 85.00, "ea"),
        ("Concrete equipment pad", 75.00, "ea"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceTable"] = DefinedName("PriceTable", attr_text=f"PRICING!$A${pricing_start}:$C${r-1}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Install labor (system swap, lead tech)", 95.00, "per hour"),
        ("Helper labor", 55.00, "per hour"),
        ("System swap labor hours (typical)", 16.00, "hours"),
        ("New ductwork labor (per supply run)", 185.00, "per run"),
        ("Refrigerant charge + commissioning", 285.00, "flat"),
        ("Permit + inspection", 250.00, "flat"),
        ("Equipment haul-away", 175.00, "flat"),
        ("Profit margin (decimal)", 0.25, "0.25 = 25%"),
        ("Sales tax (decimal)", 0.07, "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit or "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LeadRate"] = DefinedName("LeadRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["HelperRate"] = DefinedName("HelperRate", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["SwapHours"] = DefinedName("SwapHours", attr_text=f"PRICING!$B${labor_start+2}")
    wb.defined_names["DuctRunCost"] = DefinedName("DuctRunCost", attr_text=f"PRICING!$B${labor_start+3}")
    wb.defined_names["ChargeLabor"] = DefinedName("ChargeLabor", attr_text=f"PRICING!$B${labor_start+4}")
    wb.defined_names["Permit"] = DefinedName("Permit", attr_text=f"PRICING!$B${labor_start+5}")
    wb.defined_names["HaulAway"] = DefinedName("HaulAway", attr_text=f"PRICING!$B${labor_start+6}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+7}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+8}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "System", "=INPUTS!B39"); r += 1
    label_value(ws, r, "Design heating", "=DesignHeatBTU&\" BTU/hr\""); r += 1
    label_value(ws, r, "Design cooling", "=DesignCoolBTU&\" BTU/hr (\"&CEILING(DesignCoolBTU/12000,0.5)&\" tons)\""); r += 1

    r += 1
    section(ws, r, "EQUIPMENT", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    eq_start = r
    eq_items = [
        ("Heat pump condenser (sized to design)", 1, "Heat pump condenser, 3 ton (15 SEER2)"),
        ("Air handler", 1, "Air handler with electric heat strips"),
        ("Refrigerant line set", "=LineSet", "Refrigerant line set (3/8 + 7/8 insulated)"),
        ("Smart thermostat", 1, "Smart thermostat"),
        ("Filter cabinet (MERV 11)", 1, "MERV 11 filter cabinet"),
        ("Condensate pump", 1, "Condensate pump"),
        ("Equipment pad", 1, "Concrete equipment pad"),
        ("Disconnect + whip", 1, "Disconnect + whip"),
    ]
    for desc, qty, price_label in eq_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        p = ws.cell(r, 3, f'=VLOOKUP("{price_label}",PriceTable,2,FALSE)')
        p.font = MONO; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    eq_end = r - 1

    ws.cell(r, 1, "Equipment subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{eq_start}:D{eq_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    eq_subtotal_row = r; r += 1

    r += 1
    section(ws, r, "LABOR & FEES", span=4); r += 1
    labor_lines = [
        ("Install labor (lead, swap hours)",         "=SwapHours*LeadRate"),
        ("Install labor (helper, swap hours)",       "=SwapHours*HelperRate"),
        ("Refrigerant charge + commissioning",       "=ChargeLabor"),
        ("Permit + inspection",                      "=Permit"),
        ("Equipment haul-away",                      "=HaulAway"),
    ]
    labor_start = r
    for desc, formula in labor_lines:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        for col in (2, 3): ws.cell(r, col, "").border = BOX
        t = ws.cell(r, 4, formula); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    labor_end = r - 1

    ws.cell(r, 1, "Labor subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{labor_start}:D{labor_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    labor_subtotal_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (equipment + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{eq_subtotal_row}+D{labor_subtotal_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Equipment availability and refrigerant pricing "
        "subject to change. Install scheduled within 2-3 weeks of signed acceptance.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_load(wb)
    sheet_duct(wb)
    sheet_ventilation(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "LOAD", "DUCT", "VENTILATION", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "HVAC-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
