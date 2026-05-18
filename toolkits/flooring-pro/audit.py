"""Audit the Flooring Pro Toolkit workbook."""

import math
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"
XLSX = Path(__file__).parent / "dist" / "Flooring-Pro-Toolkit-v1.xlsx"
EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def recalc_via_libreoffice(src):
    tmp = Path(tempfile.mkdtemp(prefix="flooring_audit_"))
    cmd = [SOFFICE, "--headless", "--calc", "--norestore",
           "--convert-to", "xlsx", "--outdir", str(tmp), str(src)]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if res.returncode != 0:
        raise RuntimeError(f"soffice convert failed: {res.stderr or res.stdout}")
    out = tmp / src.name
    if not out.exists():
        raise RuntimeError(f"soffice did not produce {out}")
    return out


def scan_for_errors(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    found = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(e in v for e in EXCEL_ERRORS):
                    found.append((sheet, cell.coordinate, v))
    wb.close()
    return found


def expected_results():
    # INPUTS defaults
    mat_type = "LVP"
    waste = 0.10
    box_cov = 23.45
    tear_out = "Y"
    under_inc = "Y"

    # ROOMS rooms (length, width, transitions)
    rooms = [
        (18, 14, 1),
        (14, 12, 3),
        (12, 12, 2),
        (14, 13, 2),
        (12, 11, 1),
        (11, 10, 1),
        (12, 4, 3),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
    ]
    total_sf = sum(l * w for l, w, _ in rooms)
    total_trans_lf = sum(t for _, _, t in rooms)

    # MATERIALS
    mat_sf = total_sf * (1 + waste)
    boxes = math.ceil(mat_sf / box_cov)
    under_sf = total_sf if under_inc == "Y" else 0
    thinset = math.ceil(mat_sf / 40) if mat_type == "Tile" else 0
    grout = math.ceil(mat_sf / 75) if mat_type == "Tile" else 0
    adh = math.ceil(mat_sf / 200) if mat_type == "Vinyl" else 0
    nail = math.ceil(mat_sf / 100) if mat_type == "Hardwood" else 0
    pad = total_sf if mat_type == "Carpet" else 0
    trans = total_trans_lf

    # LABOR
    install_rates = {"LVP": 0.04, "Tile": 0.12, "Hardwood": 0.10, "Carpet": 0.03, "Vinyl": 0.05}
    install_rate = install_rates.get(mat_type, 0.05)
    tearout_h = (total_sf if tear_out == "Y" else 0) * 0.025
    prep_h = total_sf * 0.010
    install_h = total_sf * install_rate
    trans_h = total_trans_lf * 0.25
    setup_h = 4.0
    total_hours = tearout_h + prep_h + install_h + trans_h + setup_h

    # PRICING
    prices = {
        "LVP": 3.50, "Tile": 4.50, "Hardwood": 7.50, "Carpet": 3.20, "Vinyl": 2.20,
    }
    price_under = 0.45
    price_thinset = 22.00
    price_grout = 18.00
    price_adh = 35.00
    price_nail = 8.00
    price_pad = 0.55
    price_trans = 18.00
    price_misc = 50.00
    labor_rate = 55.00
    margin = 0.25
    tax_rate = 0.07

    mat_price = prices[mat_type]
    trans_sticks = math.ceil(trans / 4)
    materials = (
        mat_sf * mat_price +
        under_sf * price_under +
        thinset * price_thinset +
        grout * price_grout +
        adh * price_adh +
        nail * price_nail +
        pad * price_pad +
        trans_sticks * price_trans +
        1 * price_misc
    )
    labor_cost = total_hours * labor_rate
    cost_sub = materials + labor_cost
    margin_amt = cost_sub * margin
    pretax = cost_sub + margin_amt
    tax = pretax * tax_rate
    total_due = pretax + tax

    return {
        "TotalSF": total_sf,
        "TotalTransLF": total_trans_lf,
        "MatSF": mat_sf,
        "Boxes": boxes,
        "UnderSF": under_sf,
        "ThinsetBags": thinset,
        "GroutBags": grout,
        "AdhGal": adh,
        "NailLb": nail,
        "PadSF": pad,
        "TransLF": trans,
        "TotalHours": total_hours,
        "Materials": materials,
        "Labor": labor_cost,
        "CostSub": cost_sub,
        "Margin": margin_amt,
        "Pretax": pretax,
        "Tax": tax,
        "TotalDue": total_due,
    }


def approx(a, b, tol_pct=0.005, tol_abs=0.01):
    if a is None or b is None:
        return False
    if a == b:
        return True
    if abs(a - b) <= tol_abs:
        return True
    denom = max(abs(a), abs(b))
    return denom > 0 and abs(a - b) / denom <= tol_pct


def main():
    if not XLSX.exists():
        print(f"FAIL: {XLSX} does not exist — run build.py first.")
        sys.exit(1)

    print(f"Source: {XLSX}")
    print("Recalculating via LibreOffice headless...")
    recalced = recalc_via_libreoffice(XLSX)
    print(f"  -> {recalced}")

    errors = scan_for_errors(recalced)
    if errors:
        print(f"\nFAIL: {len(errors)} Excel error cell(s):")
        for sheet, coord, val in errors[:20]:
            print(f"  {sheet}!{coord} = {val}")
        sys.exit(1)
    print("OK: no #REF!/#DIV/0!/#VALUE!/#NAME?/#NULL!/#NUM!/#N/A in any cell")

    wb = load_workbook(recalced, data_only=True)
    wb_f = load_workbook(XLSX, data_only=False)
    exp = expected_results()

    def resolve(name):
        ref = wb_f.defined_names[name].value
        sheet, coord = ref.split("!")
        coord = coord.replace("$", "")
        return wb[sheet][coord].value

    quote = wb["QUOTE"]
    quote_vals = {}
    for row in quote.iter_rows():
        label_cell = row[0]
        total_cell = row[3] if len(row) >= 4 else None
        if isinstance(label_cell.value, str) and total_cell is not None:
            quote_vals[label_cell.value.strip()] = total_cell.value

    checks = [
        ("TotalSF",     resolve("TotalSF"),     exp["TotalSF"]),
        ("TotalTransLF",resolve("TotalTransLF"),exp["TotalTransLF"]),
        ("MatSF",       resolve("MatSF"),       exp["MatSF"]),
        ("Boxes",       resolve("Boxes"),       exp["Boxes"]),
        ("UnderSF",     resolve("UnderSF"),     exp["UnderSF"]),
        ("ThinsetBags", resolve("ThinsetBags"), exp["ThinsetBags"]),
        ("GroutBags",   resolve("GroutBags"),   exp["GroutBags"]),
        ("AdhGal",      resolve("AdhGal"),      exp["AdhGal"]),
        ("NailLb",      resolve("NailLb"),      exp["NailLb"]),
        ("PadSF",       resolve("PadSF"),       exp["PadSF"]),
        ("TransLF",     resolve("TransLF"),     exp["TransLF"]),
        ("TotalHours",  resolve("TotalHours"),  exp["TotalHours"]),
        ("Materials sub", quote_vals.get("Materials subtotal"),               exp["Materials"]),
        ("Cost sub",      quote_vals.get("Cost subtotal (materials + labor)"), exp["CostSub"]),
        ("Profit margin", quote_vals.get("Profit margin"),                     exp["Margin"]),
        ("Pre-tax total", quote_vals.get("Pre-tax total"),                     exp["Pretax"]),
        ("Sales tax",     quote_vals.get("Sales tax"),                         exp["Tax"]),
        ("TOTAL DUE",     quote_vals.get("TOTAL DUE"),                         exp["TotalDue"]),
    ]

    print("\nNumeric checks (workbook vs. pure-Python expected):")
    print(f"  {'Label':<22} {'Workbook':>14} {'Expected':>14}   Result")
    fails = 0
    for label, got, expv in checks:
        try:
            got_n = float(got) if got is not None else None
            exp_n = float(expv)
        except (TypeError, ValueError):
            print(f"  {label:<22} {'(not numeric)':>14} {'':>14}   FAIL")
            fails += 1
            continue
        ok = approx(got_n, exp_n)
        flag = "OK  " if ok else "FAIL"
        print(f"  {label:<22} {got_n:>14,.2f} {exp_n:>14,.2f}   {flag}")
        if not ok:
            fails += 1

    wb.close()
    wb_f.close()
    shutil.rmtree(recalced.parent, ignore_errors=True)

    if fails:
        print(f"\nFAIL: {fails} of {len(checks)} numeric checks failed")
        sys.exit(1)
    print(f"\nPASS: {len(checks)} checks, 0 Excel errors, {XLSX.name} is clean")


if __name__ == "__main__":
    main()
