"""Build the Flooring Pro Toolkit workbook.

Run: python build.py
Output: dist/Flooring-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# Sheet builders
# ==================================================================

def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Flooring Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)

    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn a list of rooms into flooring "
         "material boxes (LVP, tile, hardwood, carpet, or sheet vinyl), "
         "underlayment sq ft, supplemental materials (thinset, grout, "
         "adhesive, nails), transitions, labor hours by task, and a "
         "print-ready customer quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Pick the material type (LVP / Tile / "
         "Hardwood / Carpet / Vinyl). Yellow cells are inputs.\n"
         "2. ROOMS — enter L × W per room and transition LF per room.\n"
         "3. MATERIALS converts total sq ft + waste factor into boxes "
         "and ancillary materials.\n"
         "4. LABOR returns tear-out + prep + install + transitions hours.\n"
         "5. QUOTE is print-ready — edit prices on PRICING first. "
         "Material price is pulled by type automatically."),
        ("WASTE FACTORS (recommended)",
         "LVP/LVT click: 8-10% (straight runs, simple rooms).\n"
         "Ceramic/porcelain tile: 10-15% (diagonal layouts 15%+).\n"
         "Hardwood plank: 5-10%.\n"
         "Carpet roll: 10-15% (seam-dependent).\n"
         "Sheet vinyl: 10-15% (single-piece preferred)."),
        ("MATERIAL COVERAGE (typical box / roll)",
         "LVP: 22-30 sq ft per box (default 23.45).\n"
         "Ceramic 12x24 tile: 16 sq ft per box.\n"
         "Hardwood 3/4\" solid: 20-24 sq ft per box.\n"
         "Carpet broadloom: 12 ft × 15 ft roll = 180 sq ft.\n"
         "Sheet vinyl: 12 ft × any length roll."),
        ("SUPPLEMENTALS BY TYPE",
         "LVP click: underlayment + transitions only.\n"
         "Tile: thinset (1 bag / ~40 sf) + grout (1 bag / ~75 sf).\n"
         "Hardwood: underlayment + cleats/staples.\n"
         "Carpet: pad (1:1 sf) + tackstrip + seam tape.\n"
         "Vinyl: adhesive (~1 gal / 200 sf if glued)."),
        ("PRICING",
         "Material per-sf and supplemental costs on PRICING are 2026 "
         "ballpark — edit to your supplier. Margin compounds on "
         "(material + labor)."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label)
        a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top")
        a.fill = PANEL_FILL
        b = ws.cell(r, 2, body)
        b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Residence Flooring"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "MATERIAL", span=3); r += 1
    mat = [
        ("Material type (LVP/Tile/Hardwood/Carpet/Vinyl)", "LVP", None),
        ("Waste factor (decimal)", 0.10, "0.0%"),
        ("Coverage per box (sq ft)", 23.45, "0.00"),
        ("Tear-out existing? (Y/N)", "Y", None),
        ("Underlayment included? (Y/N)", "Y", None),
    ]
    mat_start = r
    for label, default, fmt in mat:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["MatType"] = DefinedName("MatType", attr_text=f"INPUTS!$B${mat_start}")
    wb.defined_names["Waste"] = DefinedName("Waste", attr_text=f"INPUTS!$B${mat_start+1}")
    wb.defined_names["BoxCov"] = DefinedName("BoxCov", attr_text=f"INPUTS!$B${mat_start+2}")
    wb.defined_names["TearOut"] = DefinedName("TearOut", attr_text=f"INPUTS!$B${mat_start+3}")
    wb.defined_names["UnderInc"] = DefinedName("UnderInc", attr_text=f"INPUTS!$B${mat_start+4}")

    r += 1
    n = ws.cell(r, 1, "Material options: LVP · Tile · Hardwood · Carpet · Vinyl — pricing pulls automatically"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Box coverage default = LVP (~23.45 sf). Tile boxes ~16 sf · Hardwood ~22 sf · Carpet roll ~180 sf"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Waste % default 10%; raise to 15% for diagonal tile or complex rooms"); n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_rooms(wb):
    ws = wb.create_sheet("ROOMS")
    auto_width(ws, [22, 10, 10, 14, 18, 24])
    banner(ws, "Rooms", "Enter L × W per room + transition LF", width=6)

    r = 4
    section(ws, r, "ROOM TAKEOFF (yellow cells = enter values)", span=6); r += 1
    headers = ["Room", "Length", "Width", "Sq ft", "Transitions LF", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    rooms = [
        ("Living Room", 18, 14, 1, "main entry"),
        ("Kitchen", 14, 12, 3, "to dining/hall/exterior"),
        ("Dining", 12, 12, 2, "to living/kitchen"),
        ("Master Bedroom", 14, 13, 2, "closet + hall"),
        ("Bedroom 2", 12, 11, 1, ""),
        ("Bedroom 3", 11, 10, 1, ""),
        ("Hallway", 12, 4, 3, "branches to bedrooms"),
        ("", 0, 0, 0, ""),
        ("", 0, 0, 0, ""),
        ("", 0, 0, 0, ""),
    ]
    room_start = r
    for name, length, width, trans, note in rooms:
        ws.cell(r, 1, name).font = LBL; ws.cell(r, 1).border = BOX
        for col, val, fmt in [(2, length, "0.0"), (3, width, "0.0"), (5, trans, "0")]:
            c = ws.cell(r, col, val); c.font = MONO; c.alignment = RIGHT
            c.number_format = fmt; c.fill = YELLOW; c.border = BOX
        sf = ws.cell(r, 4, f"=B{r}*C{r}")
        sf.font = MONO_BOLD; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
        n = ws.cell(r, 6, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1
    room_end = r - 1

    ws.cell(r, 1, "TOTAL").font = LBL_BOLD; ws.cell(r, 1).fill = PANEL_FILL; ws.cell(r, 1).border = BOX
    for col in (2, 3):
        ws.cell(r, col, "").fill = PANEL_FILL; ws.cell(r, col).border = BOX
    ts = ws.cell(r, 4, f"=SUM(D{room_start}:D{room_end})")
    ts.font = MONO_BOLD; ts.alignment = RIGHT; ts.number_format = "#,##0"; ts.fill = YELLOW; ts.border = BOX
    tt = ws.cell(r, 5, f"=SUM(E{room_start}:E{room_end})")
    tt.font = MONO_BOLD; tt.alignment = RIGHT; tt.number_format = "0"; tt.fill = YELLOW; tt.border = BOX
    ws.cell(r, 6, "").border = BOX
    total_row = r
    wb.defined_names["TotalSF"] = DefinedName("TotalSF", attr_text=f"ROOMS!$D${total_row}")
    wb.defined_names["TotalTransLF"] = DefinedName("TotalTransLF", attr_text=f"ROOMS!$E${total_row}")
    r += 2

    note = ws.cell(r, 1,
        "Sq ft is gross room area. Subtract islands, fixed cabinets, "
        "and built-ins from the room dimensions if those areas won't "
        "receive flooring. Transitions LF includes thresholds between "
        "rooms + at exterior doors.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    return ws


def sheet_materials(wb):
    ws = wb.create_sheet("MATERIALS")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Materials Takeoff", "Boxes + supplementals by material type", width=5)

    r = 4
    section(ws, r, "FLOORING MATERIAL", span=5); r += 1
    headers = ["Item", "Formula", "Quantity", "Unit", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Material sf needed (with waste)
    ws.cell(r, 1, "Material sq ft (with waste)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "TotalSF × (1 + waste)"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=TotalSF*(1+Waste)")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "sf").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Order at least this much").font = SMALL; ws.cell(r, 5).border = BOX
    matsf_row = r; r += 1

    # Boxes needed
    ws.cell(r, 1, "Boxes / cartons needed").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(material sf / box coverage)"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f"=ROUNDUP(C{matsf_row}/BoxCov,0)")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX; q.fill = YELLOW
    ws.cell(r, 4, "ea").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Always round up — boxes don't split").font = SMALL; ws.cell(r, 5).border = BOX
    boxes_row = r; r += 1

    r += 1
    section(ws, r, "SUPPLEMENTALS (material-specific)", span=5); r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Underlayment
    ws.cell(r, 1, "Underlayment").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "TotalSF if included else 0"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, '=IF(UPPER(UnderInc)="Y",TotalSF,0)')
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "#,##0"; q.border = BOX
    ws.cell(r, 4, "sf").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Foam or felt — LVP/HW typical").font = SMALL; ws.cell(r, 5).border = BOX
    under_row = r; r += 1

    # Thinset bags (tile only): 1 bag / 40 sf, 50lb bag
    ws.cell(r, 1, "Thinset bags (tile only)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(material sf / 40) if MatType=Tile"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f'=IF(MatType="Tile",ROUNDUP(C{matsf_row}/40,0),0)')
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "50lb bag").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Modified, 12x24 tile @ 1/4 trowel").font = SMALL; ws.cell(r, 5).border = BOX
    thinset_row = r; r += 1

    # Grout bags (tile only): 1 bag / 75 sf, 10lb
    ws.cell(r, 1, "Grout bags (tile only)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(material sf / 75) if MatType=Tile"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f'=IF(MatType="Tile",ROUNDUP(C{matsf_row}/75,0),0)')
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "10lb bag").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Sanded for joints ≥1/8\"").font = SMALL; ws.cell(r, 5).border = BOX
    grout_row = r; r += 1

    # Adhesive gallons (vinyl/glue-down LVP): 1 gal / 200 sf
    ws.cell(r, 1, "Adhesive gallons (Vinyl)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(material sf / 200) if MatType=Vinyl"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f'=IF(MatType="Vinyl",ROUNDUP(C{matsf_row}/200,0),0)')
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "gal").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Pressure-sensitive for sheet"); ws.cell(r, 5).font = SMALL; ws.cell(r, 5).border = BOX
    adh_row = r; r += 1

    # Cleats / nails (hardwood): 1 lb / 100 sf
    ws.cell(r, 1, "Cleats / nails (Hardwood)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "ROUNDUP(material sf / 100) if MatType=Hardwood"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, f'=IF(MatType="Hardwood",ROUNDUP(C{matsf_row}/100,0),0)')
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "lb").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "L or T cleat at 16\" OC").font = SMALL; ws.cell(r, 5).border = BOX
    nails_row = r; r += 1

    # Carpet pad sf (1:1)
    ws.cell(r, 1, "Carpet pad").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "TotalSF if MatType=Carpet"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, '=IF(MatType="Carpet",TotalSF,0)')
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "#,##0"; q.border = BOX
    ws.cell(r, 4, "sf").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "8 lb rebond typical"); ws.cell(r, 5).font = SMALL; ws.cell(r, 5).border = BOX
    pad_row = r; r += 1

    # Transitions
    ws.cell(r, 1, "Transitions (T-mold/reducer)").font = LBL; ws.cell(r, 1).border = BOX
    f = ws.cell(r, 2, "TotalTransLF"); f.font = MONO; f.alignment = LEFT; f.border = BOX
    q = ws.cell(r, 3, "=TotalTransLF")
    q.font = MONO_BOLD; q.alignment = RIGHT; q.number_format = "0"; q.border = BOX
    ws.cell(r, 4, "LF").font = LBL; ws.cell(r, 4).border = BOX
    ws.cell(r, 5, "Cut from 4 ft sticks typically"); ws.cell(r, 5).font = SMALL; ws.cell(r, 5).border = BOX
    trans_row = r; r += 1

    # Named ranges
    wb.defined_names["MatSF"] = DefinedName("MatSF", attr_text=f"MATERIALS!$C${matsf_row}")
    wb.defined_names["Boxes"] = DefinedName("Boxes", attr_text=f"MATERIALS!$C${boxes_row}")
    wb.defined_names["UnderSF"] = DefinedName("UnderSF", attr_text=f"MATERIALS!$C${under_row}")
    wb.defined_names["ThinsetBags"] = DefinedName("ThinsetBags", attr_text=f"MATERIALS!$C${thinset_row}")
    wb.defined_names["GroutBags"] = DefinedName("GroutBags", attr_text=f"MATERIALS!$C${grout_row}")
    wb.defined_names["AdhGal"] = DefinedName("AdhGal", attr_text=f"MATERIALS!$C${adh_row}")
    wb.defined_names["NailLb"] = DefinedName("NailLb", attr_text=f"MATERIALS!$C${nails_row}")
    wb.defined_names["PadSF"] = DefinedName("PadSF", attr_text=f"MATERIALS!$C${pad_row}")
    wb.defined_names["TransLF"] = DefinedName("TransLF", attr_text=f"MATERIALS!$C${trans_row}")
    return ws


def sheet_labor(wb):
    ws = wb.create_sheet("LABOR")
    auto_width(ws, [34, 14, 14, 14, 28])
    banner(ws, "Labor Hours", "Tear-out + prep + install + transitions", width=5)

    r = 4
    section(ws, r, "INSTALL RATE BY MATERIAL", span=5); r += 1
    label_value(ws, r, "Install hr/sf (lookup by MatType)",
        '=IF(MatType="LVP",0.04,IF(MatType="Tile",0.12,IF(MatType="Hardwood",0.10,IF(MatType="Carpet",0.03,IF(MatType="Vinyl",0.05,0.05)))))',
        fmt="0.000").font = MONO_BOLD
    rate_row = r; r += 1

    r += 1
    section(ws, r, "BASE HOURS PER TASK", span=5); r += 1
    headers = ["Task", "Sq ft / unit", "Hr per unit", "Hours", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    # Tear-out (only if TearOut=Y)
    ws.cell(r, 1, "Tear-out existing flooring").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, '=IF(UPPER(TearOut)="Y",TotalSF,0)')
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.025)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Pry, scrape, haul to dumpster"); n.font = SMALL; n.border = BOX
    tearout_row = r; r += 1

    # Subfloor prep
    ws.cell(r, 1, "Subfloor prep + level").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalSF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.010)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Sweep, fill, sand, skim if needed"); n.font = SMALL; n.border = BOX
    prep_row = r; r += 1

    # Install (uses rate by MatType)
    ws.cell(r, 1, "Install flooring").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalSF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "#,##0"; sf.border = BOX
    hr = ws.cell(r, 3, f"=B{rate_row}")
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.000"; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Rate auto-set by material type"); n.font = SMALL; n.border = BOX
    install_row = r; r += 1

    # Transitions: 0.25 hr / LF
    ws.cell(r, 1, "Transitions install").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, "=TotalTransLF")
    sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 0.25)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.00"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Measure, cut, fasten"); n.font = SMALL; n.border = BOX
    trans_hours_row = r; r += 1

    # Setup
    ws.cell(r, 1, "Setup + cleanup").font = LBL; ws.cell(r, 1).border = BOX
    sf = ws.cell(r, 2, 1); sf.font = MONO; sf.alignment = RIGHT; sf.number_format = "0"; sf.border = BOX
    hr = ws.cell(r, 3, 4.0)
    hr.font = MONO; hr.alignment = RIGHT; hr.number_format = "0.0"; hr.fill = YELLOW; hr.border = BOX
    h = ws.cell(r, 4, f"=B{r}*C{r}")
    h.font = MONO_BOLD; h.alignment = RIGHT; h.number_format = "0.0"; h.border = BOX
    n = ws.cell(r, 5, "Move furniture, protect, vacuum"); n.font = SMALL; n.border = BOX
    setup_row = r; r += 1

    ws.cell(r, 1, "TOTAL LABOR HOURS").font = LBL_BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).fill = PANEL_FILL
    for col in (2, 3): ws.cell(r, col, "").border = BOX; ws.cell(r, col).fill = PANEL_FILL
    s = ws.cell(r, 4, f"=SUM(D{tearout_row}:D{setup_row})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = "0.0"; s.fill = YELLOW; s.border = BOX
    ws.cell(r, 5, "").border = BOX; ws.cell(r, 5).fill = PANEL_FILL
    total_hours_row = r
    wb.defined_names["TotalHours"] = DefinedName("TotalHours", attr_text=f"LABOR!$D${total_hours_row}")
    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "FLOORING MATERIAL — $ PER SQ FT (2026)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("LVP / LVT — per sf",                  3.50, "sf"),
        ("Ceramic / Porcelain Tile — per sf",   4.50, "sf"),
        ("Hardwood (solid 3/4\") — per sf",     7.50, "sf"),
        ("Carpet (mid-grade) — per sf",         3.20, "sf"),
        ("Sheet Vinyl — per sf",                2.20, "sf"),
        ("Underlayment foam — per sf",          0.45, "sf"),
        ("Thinset (50lb bag)",                 22.00, "bag"),
        ("Grout (10lb bag)",                   18.00, "bag"),
        ("Vinyl adhesive (1 gal)",             35.00, "gal"),
        ("Hardwood cleats / nails (per lb)",    8.00, "lb"),
        ("Carpet pad — per sf",                 0.55, "sf"),
        ("Transition (4 ft T-mold or reducer)", 18.00, "stick"),
        ("Misc fasteners + trim (flat)",       50.00, "job"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceLVP"] = DefinedName("PriceLVP", attr_text=f"PRICING!$B${pricing_start}")
    wb.defined_names["PriceTile"] = DefinedName("PriceTile", attr_text=f"PRICING!$B${pricing_start+1}")
    wb.defined_names["PriceHW"] = DefinedName("PriceHW", attr_text=f"PRICING!$B${pricing_start+2}")
    wb.defined_names["PriceCarpet"] = DefinedName("PriceCarpet", attr_text=f"PRICING!$B${pricing_start+3}")
    wb.defined_names["PriceVinyl"] = DefinedName("PriceVinyl", attr_text=f"PRICING!$B${pricing_start+4}")
    wb.defined_names["PriceUnder"] = DefinedName("PriceUnder", attr_text=f"PRICING!$B${pricing_start+5}")
    wb.defined_names["PriceThinset"] = DefinedName("PriceThinset", attr_text=f"PRICING!$B${pricing_start+6}")
    wb.defined_names["PriceGrout"] = DefinedName("PriceGrout", attr_text=f"PRICING!$B${pricing_start+7}")
    wb.defined_names["PriceAdh"] = DefinedName("PriceAdh", attr_text=f"PRICING!$B${pricing_start+8}")
    wb.defined_names["PriceNail"] = DefinedName("PriceNail", attr_text=f"PRICING!$B${pricing_start+9}")
    wb.defined_names["PricePad"] = DefinedName("PricePad", attr_text=f"PRICING!$B${pricing_start+10}")
    wb.defined_names["PriceTrans"] = DefinedName("PriceTrans", attr_text=f"PRICING!$B${pricing_start+11}")
    wb.defined_names["PriceMisc"] = DefinedName("PriceMisc", attr_text=f"PRICING!$B${pricing_start+12}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Labor rate (lead installer)",   55.00, "per hour"),
        ("Profit margin (decimal)",        0.25, "0.25 = 25%"),
        ("Sales tax (decimal)",            0.07, "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit or "margin" in label.lower() or "tax" in label.lower():
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["LaborRate"] = DefinedName("LaborRate", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+2}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Material type", "=MatType"); r += 1
    label_value(ws, r, "Total area", '=TotalSF&" sq ft"'); r += 1
    label_value(ws, r, "Transitions", '=TotalTransLF&" LF"'); r += 1
    label_value(ws, r, "Total labor hours", "=TotalHours", fmt="0.0"); r += 1

    r += 1
    section(ws, r, "MATERIALS", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    mat_start = r
    # Flooring material — per sf with waste, price selected by MatType
    mat_items = [
        ("Flooring material (with waste)", "=MatSF",
         '=IF(MatType="LVP",PriceLVP,IF(MatType="Tile",PriceTile,IF(MatType="Hardwood",PriceHW,IF(MatType="Carpet",PriceCarpet,IF(MatType="Vinyl",PriceVinyl,PriceLVP)))))'),
        ("Underlayment", "=UnderSF", "=PriceUnder"),
        ("Thinset", "=ThinsetBags", "=PriceThinset"),
        ("Grout", "=GroutBags", "=PriceGrout"),
        ("Adhesive", "=AdhGal", "=PriceAdh"),
        ("Cleats / nails", "=NailLb", "=PriceNail"),
        ("Carpet pad", "=PadSF", "=PricePad"),
        ("Transition pieces (sticks)", "=ROUNDUP(TransLF/4,0)", "=PriceTrans"),
        ("Misc fasteners + trim", 1, "=PriceMisc"),
    ]
    for desc, qty, price in mat_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        p = ws.cell(r, 3, price); p.font = MONO; p.alignment = RIGHT
        p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    mat_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{mat_start}:D{mat_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    mat_sub_row = r; r += 1

    r += 1
    section(ws, r, "LABOR", span=4); r += 1
    ws.cell(r, 1, "Labor (total hours × rate)").font = LBL; ws.cell(r, 1).border = BOX
    c = ws.cell(r, 2, "=TotalHours"); c.font = MONO; c.alignment = RIGHT; c.number_format = "0.0"; c.border = BOX
    p = ws.cell(r, 3, "=LaborRate"); p.font = MONO; p.alignment = RIGHT
    p.number_format = '"$"#,##0.00'; p.border = BOX
    t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
    t.number_format = '"$"#,##0.00'; t.border = BOX
    labor_total_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{mat_sub_row}+D{labor_total_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3):
        c = ws.cell(r, col, ""); c.fill = YELLOW; c.border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Custom patterns (herringbone, "
        "diagonal, inlays), pattern matching, floor leveling beyond "
        "1/8\" in 10 ft, and trim/baseboard work are billed separately. "
        "Subfloor moisture must be within manufacturer spec.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_rooms(wb)
    sheet_materials(wb)
    sheet_labor(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "ROOMS", "MATERIALS", "LABOR", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Flooring-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
