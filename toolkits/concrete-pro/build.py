"""Build the Concrete Pro Toolkit workbook.

Run: python build.py
Output: dist/Concrete-Pro-Toolkit-v1.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

HI_VIS = "FFFFD400"
INK = "FF0E0E0C"
INK_2 = "FF555555"
LINE = "FFCCCCCC"
PANEL = "FFF6F6F2"
PANEL_2 = "FFFAFAF7"

H1 = Font(name="Arial Black", size=18, color=INK, bold=True)
H2 = Font(name="Arial Black", size=12, color=INK, bold=True)
H3 = Font(name="Arial", size=11, color=INK, bold=True)
LBL = Font(name="Arial", size=10, color=INK)
LBL_BOLD = Font(name="Arial", size=10, color=INK, bold=True)
MONO = Font(name="Consolas", size=10, color=INK)
MONO_BOLD = Font(name="Consolas", size=10, color=INK, bold=True)
SMALL = Font(name="Arial", size=9, color=INK_2, italic=True)

YELLOW = PatternFill("solid", fgColor=HI_VIS)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
PANEL2_FILL = PatternFill("solid", fgColor=PANEL_2)

THIN = Side(border_style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def banner(ws, title, subtitle, width=8):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c1 = ws.cell(1, 1, "PROJECTCALC  ◆  " + title.upper())
    c1.fill = YELLOW
    c1.font = Font(name="Arial Black", size=16, color=INK, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2 = ws.cell(2, 1, subtitle)
    c2.fill = PatternFill("solid", fgColor=INK)
    c2.font = Font(name="Consolas", size=10, color="FFFFD400")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label_value(ws, row, label, value=None, fmt=None):
    a = ws.cell(row, 1, label)
    a.font = LBL_BOLD; a.alignment = LEFT; a.fill = PANEL_FILL; a.border = BOX
    b = ws.cell(row, 2, value)
    b.font = MONO; b.alignment = LEFT; b.fill = PANEL2_FILL; b.border = BOX
    if fmt:
        b.number_format = fmt
    return b


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = H2; c.fill = YELLOW; c.alignment = LEFT
    return c


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    auto_width(ws, [22, 70])
    banner(ws, "Concrete Pro Toolkit", "v1 · 2026 · projectcalc.app", width=2)
    rows = [
        ("WHAT'S INSIDE",
         "Seven linked tabs that turn slab, footing, and wall dimensions "
         "into a complete concrete + rebar takeoff — yardage, base, vapor "
         "barrier, rebar grid, ties, chairs, form lumber — plus a "
         "print-ready customer quote."),
        ("HOW TO USE",
         "1. Open the INPUTS tab. Yellow cells are inputs.\n"
         "2. SLAB returns slab yardage, base material, vapor barrier rolls, "
         "wire mesh, and curing compound.\n"
         "3. FOOTINGS-WALLS handles continuous footings + stem walls.\n"
         "4. REBAR runs the grid math for slab + wall reinforcement.\n"
         "5. PRICING — set unit costs once.\n"
         "6. Print QUOTE."),
        ("YARDAGE FORMULA",
         "Cubic yards = (length × width × thickness_inches / 12) / 27. "
         "Slab example: 30 × 40 × 4\" = (30 × 40 × 0.333) / 27 = 14.8 yd. "
         "Toolkit defaults to 8% waste — pump trucks return ~1 yd if you "
         "shorted the order, so order high."),
        ("BASE + VAPOR BARRIER",
         "Compacted base default 4\" of crushed stone or sand. Vapor barrier "
         "is 6-mil poly, sold in 200 sq ft rolls (e.g. 10×20 ft) — the "
         "SLAB tab counts rolls with 8\" overlap."),
        ("REBAR GRID (slab)",
         "Bars per direction = (perpendicular span / spacing) + 1. Default "
         "#4 (1/2\") at 18\" OC each way. Lap splice = 40 × bar diameter "
         "(20\" for #4). Tie wire roll covers ≈ 200 ties; 1 tie per "
         "intersection is standard."),
        ("FORM LUMBER",
         "Toolkit defaults to 2x4 SPF stakes every 4 ft + 2x10 SPF for the "
         "form face on slabs / footings. Walls use 3/4\" plywood with "
         "double-2x4 walers — toggle the wall thickness on INPUTS."),
        ("PRICING",
         "Concrete delivery rates are 2026 ballpark — edit PRICING to "
         "match your supplier. Pump truck billed per cubic yard or "
         "per-day flat. Add overtime + short-load fees if applicable."),
        ("LICENSE",
         "Single user. Use it on as many projects as you like; please "
         "don't redistribute the file. Bug reports & feature requests: "
         "bullbears21@gmail.com."),
    ]
    r = 4
    for label, body in rows:
        a = ws.cell(r, 1, label); a.font = H3
        a.alignment = Alignment(horizontal="left", vertical="top"); a.fill = PANEL_FILL
        b = ws.cell(r, 2, body); b.font = LBL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(36, body.count("\n") * 18 + 48)
        r += 1
    return ws


def sheet_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    auto_width(ws, [38, 18, 30])
    banner(ws, "Inputs", "Yellow cells = enter your project values", width=3)

    r = 4
    section(ws, r, "PROJECT", span=3); r += 1
    label_value(ws, r, "Project name", "Smith Garage Slab"); r += 1
    label_value(ws, r, "Customer name", "John Smith"); r += 1
    label_value(ws, r, "Address", "421 Maple Street"); r += 1
    label_value(ws, r, "Phone", "(555) 555-1234"); r += 1
    label_value(ws, r, "Estimator", "Your Name"); r += 1
    label_value(ws, r, "Date", "=TODAY()", fmt="mm/dd/yyyy"); r += 1

    r += 1
    section(ws, r, "SLAB", span=3); r += 1
    slab = [
        ("Slab length (ft)", 40, "0"),
        ("Slab width (ft)", 30, "0"),
        ("Slab thickness (inches)", 4, "0"),
        ("Compacted base depth (inches)", 4, "0"),
        ("Concrete strength (psi)", 4000, "0"),
        ("Slump target (inches)", 4, "0"),
        ("Air entrained (Y/N)", "Y", None),
        ("Wire mesh (Y/N)", "N", None),  # if N, rebar only
        ("Curing method", "Cure & seal", None),  # Cure & seal | Wet burlap | Plastic sheet
    ]
    slab_start = r
    for label, default, fmt in slab:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["SlabL"] = DefinedName("SlabL", attr_text=f"INPUTS!$B${slab_start}")
    wb.defined_names["SlabW"] = DefinedName("SlabW", attr_text=f"INPUTS!$B${slab_start+1}")
    wb.defined_names["SlabT"] = DefinedName("SlabT", attr_text=f"INPUTS!$B${slab_start+2}")
    wb.defined_names["BaseDepth"] = DefinedName("BaseDepth", attr_text=f"INPUTS!$B${slab_start+3}")
    wb.defined_names["PSI"] = DefinedName("PSI", attr_text=f"INPUTS!$B${slab_start+4}")
    wb.defined_names["Slump"] = DefinedName("Slump", attr_text=f"INPUTS!$B${slab_start+5}")
    wb.defined_names["AirEnt"] = DefinedName("AirEnt", attr_text=f"INPUTS!$B${slab_start+6}")
    wb.defined_names["WireMesh"] = DefinedName("WireMesh", attr_text=f"INPUTS!$B${slab_start+7}")
    wb.defined_names["CureMethod"] = DefinedName("CureMethod", attr_text=f"INPUTS!$B${slab_start+8}")

    r += 1
    section(ws, r, "FOOTINGS (continuous)", span=3); r += 1
    foot = [
        ("Footing length (lin ft)", 80, "0"),
        ("Footing width (inches)", 16, "0"),
        ("Footing depth (inches)", 8, "0"),
    ]
    foot_start = r
    for label, default, fmt in foot:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["FootLen"] = DefinedName("FootLen", attr_text=f"INPUTS!$B${foot_start}")
    wb.defined_names["FootW"] = DefinedName("FootW", attr_text=f"INPUTS!$B${foot_start+1}")
    wb.defined_names["FootD"] = DefinedName("FootD", attr_text=f"INPUTS!$B${foot_start+2}")

    r += 1
    section(ws, r, "STEM WALLS (optional)", span=3); r += 1
    wall = [
        ("Wall length (lin ft)", 0, "0"),
        ("Wall height (ft)", 4, "0"),
        ("Wall thickness (inches)", 8, "0"),
    ]
    wall_start = r
    for label, default, fmt in wall:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["WallLen"] = DefinedName("WallLen", attr_text=f"INPUTS!$B${wall_start}")
    wb.defined_names["WallH"] = DefinedName("WallH", attr_text=f"INPUTS!$B${wall_start+1}")
    wb.defined_names["WallTh"] = DefinedName("WallTh", attr_text=f"INPUTS!$B${wall_start+2}")

    r += 1
    section(ws, r, "REBAR + WASTE", span=3); r += 1
    rebar = [
        ("Slab rebar size (#3 / #4 / #5)", "#4", None),
        ("Slab rebar spacing (inches OC)", 18, "0"),
        ("Footing rebar size", "#4", None),
        ("Footing rebar count (longitudinal bars)", 2, "0"),
        ("Wall rebar size", "#4", None),
        ("Wall rebar spacing (inches OC, both directions)", 18, "0"),
        ("Concrete waste (%)", 8, "0"),
    ]
    rebar_start = r
    for label, default, fmt in rebar:
        c = label_value(ws, r, label, default, fmt=fmt); c.fill = YELLOW; r += 1
    wb.defined_names["SlabBar"] = DefinedName("SlabBar", attr_text=f"INPUTS!$B${rebar_start}")
    wb.defined_names["SlabSpacing"] = DefinedName("SlabSpacing", attr_text=f"INPUTS!$B${rebar_start+1}")
    wb.defined_names["FootBar"] = DefinedName("FootBar", attr_text=f"INPUTS!$B${rebar_start+2}")
    wb.defined_names["FootBarCount"] = DefinedName("FootBarCount", attr_text=f"INPUTS!$B${rebar_start+3}")
    wb.defined_names["WallBar"] = DefinedName("WallBar", attr_text=f"INPUTS!$B${rebar_start+4}")
    wb.defined_names["WallSpacing"] = DefinedName("WallSpacing", attr_text=f"INPUTS!$B${rebar_start+5}")
    wb.defined_names["WastePct"] = DefinedName("WastePct", attr_text=f"INPUTS!$B${rebar_start+6}")

    r += 1
    n = ws.cell(r, 1, "Rebar size options: #3 (3/8 in) | #4 (1/2 in) | #5 (5/8 in)")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1, "Curing options: Cure & seal | Wet burlap | Plastic sheet")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    n = ws.cell(r, 1,
        "Concrete strength: 3000 psi sidewalks · 3500 driveways · 4000 garage / structural slabs · 4500+ commercial.")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    return ws


def sheet_slab(wb):
    ws = wb.create_sheet("SLAB")
    auto_width(ws, [38, 14, 14, 36])
    banner(ws, "Slab Takeoff", "Yardage, base, vapor barrier, mesh, curing — pulled from INPUTS", width=4)

    r = 4
    section(ws, r, "SLAB GEOMETRY", span=4); r += 1
    label_value(ws, r, "Slab area (sq ft)", "=SlabL*SlabW", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Slab volume (cu ft)", "=SlabL*SlabW*(SlabT/12)", fmt="#,##0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Slab volume (cu yd)", "=SlabL*SlabW*(SlabT/12)/27", fmt="0.00").font = MONO_BOLD; r += 1
    label_value(ws, r, "Slab volume + waste (cu yd)",
        "=ROUNDUP(SlabL*SlabW*(SlabT/12)/27*(1+WastePct/100)*4,0)/4", fmt="0.00").font = Font(name="Arial Black", size=11, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    slab_yd_row = r; r += 1

    wb.defined_names["SlabYd"] = DefinedName("SlabYd", attr_text=f"SLAB!$B${slab_yd_row}")

    r += 1
    section(ws, r, "MATERIALS (with waste)", span=4); r += 1
    headers = ["Material", "Quantity", "Unit", "Notes / Coverage"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    items = [
        ("Compacted base material (cu yd)",
         "=ROUNDUP(SlabL*SlabW*(BaseDepth/12)/27*1.10,0)",
         "cu yd",
         "Crushed stone or sand; +10% compaction allowance"),
        ("Vapor barrier rolls (6-mil poly)",
         "=ROUNDUP(SlabL*SlabW/200,0)",
         "rolls",
         "1 roll = 200 sq ft (10×20); 8\" overlap covered"),
        ("Wire mesh sheets (4×7 ft)",
         '=IF(WireMesh="Y",ROUNDUP(SlabL*SlabW/28*1.05,0),0)',
         "sheets",
         "If using mesh instead of grid rebar; +5% overlap"),
        ("Curing compound (membrane forming)",
         "=ROUNDUP(SlabL*SlabW/200,0)",
         "gallons",
         "≈200 sq ft per gallon"),
        ("Slab edge form (2x10 SPF)",
         "=ROUNDUP((SlabL*2+SlabW*2)/12,0)",
         "12' pieces",
         "Perimeter form, 2x10 deep enough for 4-8\" slab"),
        ("Form stakes (2x4 SPF, 4 ft OC)",
         "=ROUNDUP((SlabL*2+SlabW*2)/4,0)",
         "stakes",
         "1 stake every 4 ft of perimeter"),
        ("Form release (oil)",
         "=ROUNDUP((SlabL*2+SlabW*2)*(SlabT/12)/100,0)",
         "gallons",
         "≈100 sq ft of form face per gallon"),
        ("Construction joints (sawcut, lin ft)",
         "=ROUNDUP(SlabL*SlabW/144,0)*12",
         "lin ft",
         "Joints in 12×12 ft grid (control cracking)"),
    ]
    for label, formula, unit, note in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    note = ws.cell(r, 1,
        "Pump truck recommended for any slab > 10 cu yd or anywhere a "
        "concrete truck can't reach the form. Order ½ yd over the calc "
        "to avoid running short — supplier usually allows return of unused.")
    note.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)
    return ws


def sheet_footings_walls(wb):
    ws = wb.create_sheet("FOOTINGS-WALLS")
    auto_width(ws, [38, 14, 14, 36])
    banner(ws, "Footings + Stem Walls", "Continuous footings + optional walls — pulled from INPUTS", width=4)

    r = 4
    section(ws, r, "FOOTINGS", span=4); r += 1
    label_value(ws, r, "Footing volume (cu ft)",
        "=FootLen*(FootW/12)*(FootD/12)", fmt="#,##0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Footing volume (cu yd, with waste)",
        "=ROUNDUP(FootLen*(FootW/12)*(FootD/12)/27*(1+WastePct/100)*4,0)/4",
        fmt="0.00").font = Font(name="Arial Black", size=11, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    foot_yd_row = r; r += 1

    wb.defined_names["FootYd"] = DefinedName("FootYd", attr_text=f"'FOOTINGS-WALLS'!$B${foot_yd_row}")

    headers = ["Material", "Quantity", "Unit", "Notes"]
    r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    foot_items = [
        ("Footing form (2x material, both sides)",
         "=ROUNDUP(FootLen*2/8,0)",
         "8' pieces",
         "2x material sized to footing depth"),
        ("Form stakes (4 ft OC)",
         "=ROUNDUP(FootLen*2/4,0)",
         "stakes",
         "Both sides of trench form"),
    ]
    for label, formula, unit, note in foot_items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 1
    section(ws, r, "STEM WALLS (skipped if WallLen = 0)", span=4); r += 1
    label_value(ws, r, "Wall volume (cu ft)",
        "=WallLen*WallH*(WallTh/12)", fmt="#,##0.0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Wall volume (cu yd, with waste)",
        "=ROUNDUP(WallLen*WallH*(WallTh/12)/27*(1+WastePct/100)*4,0)/4",
        fmt="0.00").font = Font(name="Arial Black", size=11, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    wall_yd_row = r; r += 1

    wb.defined_names["WallYd"] = DefinedName("WallYd", attr_text=f"'FOOTINGS-WALLS'!$B${wall_yd_row}")

    r += 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1

    wall_items = [
        ("Wall forms (3/4\" plywood 4×8 sheets)",
         "=ROUNDUP(WallLen*WallH*2/32,0)",
         "sheets",
         "Both sides; 4×8 sheet = 32 sq ft"),
        ("Walers (double 2x4, 2 ft OC vertically)",
         "=ROUNDUP(WallLen*(WallH/2)*4/8,0)",
         "8' pieces",
         "Walers + braces for both wall sides"),
        ("Form ties (snap ties, 2 ft OC each way)",
         "=ROUNDUP(WallLen*WallH/4,0)",
         "ties",
         "1 tie per 4 sq ft of wall face"),
        ("Form stakes / kickers (4 ft OC)",
         "=ROUNDUP(WallLen*2/4,0)",
         "stakes",
         "Both sides, kicked back to grade"),
    ]
    for label, formula, unit, note in wall_items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    r += 2
    section(ws, r, "TOTAL CONCRETE ORDER", span=4); r += 1
    label_value(ws, r, "Total cubic yards (slab + footing + wall)",
        "=SlabYd+FootYd+WallYd", fmt="0.00").font = Font(name="Arial Black", size=12, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    r += 1
    return ws


def sheet_rebar(wb):
    ws = wb.create_sheet("REBAR")
    auto_width(ws, [38, 14, 14, 36])
    banner(ws, "Rebar Takeoff", "Slab grid + footing + wall reinforcement", width=4)

    r = 4
    section(ws, r, "BAR DIAMETER REFERENCE", span=4); r += 1
    headers = ["Bar Size", "Diameter", "Lap Splice (40d)", "Weight (lb/ft)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    bars = [
        ("#3", "3/8 in", "15 in", 0.376),
        ("#4", "1/2 in", "20 in", 0.668),
        ("#5", "5/8 in", "25 in", 1.043),
        ("#6", "3/4 in", "30 in", 1.502),
    ]
    for size, dia, lap, wt in bars:
        ws.cell(r, 1, size).font = MONO_BOLD; ws.cell(r, 1).alignment = CENTER; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, dia).font = MONO; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BOX
        ws.cell(r, 3, lap).font = MONO; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BOX
        c = ws.cell(r, 4, wt); c.font = MONO; c.alignment = RIGHT; c.number_format = "0.000"; c.border = BOX
        r += 1

    r += 1
    section(ws, r, "SLAB GRID (each way)", span=4); r += 1
    # Bars in long direction (running along length, spaced across width)
    label_value(ws, r, "Bars across width (count)",
        "=ROUNDUP(SlabW*12/SlabSpacing,0)+1", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Bars across length (count)",
        "=ROUNDUP(SlabL*12/SlabSpacing,0)+1", fmt="#,##0").font = MONO_BOLD; r += 1
    label_value(ws, r, "Bar length needed each way (ft)",
        "=(ROUNDUP(SlabW*12/SlabSpacing,0)+1)*SlabL+(ROUNDUP(SlabL*12/SlabSpacing,0)+1)*SlabW",
        fmt="#,##0").font = MONO_BOLD
    slab_lf_row = r; r += 1
    label_value(ws, r, "Bar count (20-ft sticks, 5% waste)",
        f"=ROUNDUP(B{slab_lf_row}/20*1.05,0)", fmt="#,##0").font = Font(name="Arial Black", size=11, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    slab_count_row = r; r += 1

    wb.defined_names["SlabBars"] = DefinedName("SlabBars", attr_text=f"REBAR!$B${slab_count_row}")

    r += 1
    section(ws, r, "FOOTING (longitudinal)", span=4); r += 1
    label_value(ws, r, "Total footing bar length (ft)",
        "=FootLen*FootBarCount", fmt="#,##0").font = MONO_BOLD
    foot_lf_row = r; r += 1
    label_value(ws, r, "Bar count (20-ft sticks, 5% waste)",
        f"=ROUNDUP(B{foot_lf_row}/20*1.05,0)", fmt="#,##0").font = Font(name="Arial Black", size=11, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    foot_count_row = r; r += 1

    wb.defined_names["FootBars"] = DefinedName("FootBars", attr_text=f"REBAR!$B${foot_count_row}")

    r += 1
    section(ws, r, "WALL (each way grid)", span=4); r += 1
    label_value(ws, r, "Vertical bars (count)",
        "=IF(WallLen=0,0,ROUNDUP(WallLen*12/WallSpacing,0)+1)",
        fmt="#,##0").font = MONO_BOLD
    wall_v_row = r; r += 1
    label_value(ws, r, "Horizontal bars (rows)",
        "=IF(WallLen=0,0,ROUNDUP(WallH*12/WallSpacing,0)+1)",
        fmt="#,##0").font = MONO_BOLD
    wall_h_row = r; r += 1
    label_value(ws, r, "Total wall bar length (ft)",
        f"=B{wall_v_row}*WallH+B{wall_h_row}*WallLen",
        fmt="#,##0").font = MONO_BOLD
    wall_lf_row = r; r += 1
    label_value(ws, r, "Bar count (20-ft sticks, 5% waste)",
        f"=ROUNDUP(B{wall_lf_row}/20*1.05,0)", fmt="#,##0").font = Font(name="Arial Black", size=11, color=INK, bold=True)
    ws.cell(r, 2).fill = YELLOW
    wall_count_row = r; r += 1

    wb.defined_names["WallBars"] = DefinedName("WallBars", attr_text=f"REBAR!$B${wall_count_row}")

    r += 1
    section(ws, r, "ACCESSORIES", span=4); r += 1
    headers2 = ["Item", "Quantity", "Unit", "Notes"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    acc = [
        ("Tie wire (16 ga, 3.5 lb roll)",
         f"=ROUNDUP(((B{slab_count_row}+B{foot_count_row}+B{wall_count_row})*20*0.5)/200,0)",
         "rolls",
         "≈1 tie per 2 ft of bar; 200 ties per roll"),
        ("Rebar chairs (slab grid support)",
         "=ROUNDUP(SlabL*SlabW/16,0)",
         "ea",
         "1 chair per 16 sq ft of slab grid"),
        ("Bolster (continuous footing chair)",
         "=ROUNDUP(FootLen/4,0)",
         "ea",
         "1 every 4 lin ft of footing"),
        ("Total bars (all elements)",
         f"=B{slab_count_row}+B{foot_count_row}+B{wall_count_row}",
         "20' sticks",
         "Combined order quantity"),
    ]
    for label, formula, unit, note in acc:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.font = MONO_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        n = ws.cell(r, 4, note); n.font = SMALL; n.alignment = LEFT; n.border = BOX
        r += 1

    return ws


def sheet_pricing(wb):
    ws = wb.create_sheet("PRICING")
    auto_width(ws, [38, 14, 18])
    banner(ws, "Pricing", "Edit unit prices to match your supplier (yellow cells)", width=3)

    r = 4
    section(ws, r, "MATERIALS (2026 ballpark)", span=3); r += 1
    headers = ["Item", "Unit Cost", "Unit"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = PANEL_FILL; c.alignment = CENTER; c.border = BOX
    r += 1
    items = [
        ("Concrete delivered, 3000 psi",       195.00, "cu yd"),
        ("Concrete delivered, 3500 psi",       205.00, "cu yd"),
        ("Concrete delivered, 4000 psi",       215.00, "cu yd"),
        ("Concrete delivered, 4500 psi",       225.00, "cu yd"),
        ("Air-entrainment surcharge",          8.00,   "cu yd"),
        ("Short-load fee (< 5 yd)",            150.00, "load"),
        ("Saturday delivery surcharge",        125.00, "load"),
        ("Crushed stone base (3/4\" gravel)",  42.00,  "cu yd"),
        ("Sand base",                           38.00,  "cu yd"),
        ("Vapor barrier (6-mil 200 sf roll)",  28.00,  "roll"),
        ("Wire mesh 6×6 W2.9 (4×7 sheet)",     18.00,  "sheet"),
        ("Curing compound (5-gal pail)",       95.00,  "pail"),
        ("Curing compound (1-gal)",            22.00,  "gal"),
        ("Form release oil (5-gal)",           58.00,  "pail"),
        ("Rebar #3 (20-ft stick)",             7.50,   "ea"),
        ("Rebar #4 (20-ft stick)",             8.50,   "ea"),
        ("Rebar #5 (20-ft stick)",             14.00,  "ea"),
        ("Rebar tie wire (3.5 lb roll)",       12.00,  "roll"),
        ("Rebar chair, 3\" plastic",           0.45,   "ea"),
        ("Bolster (continuous chair, 5 ft)",   8.50,   "ea"),
        ("2x10 SPF 12-ft (form)",              22.00,  "stick"),
        ("2x4 SPF 8-ft (stake / brace)",       5.20,   "stick"),
        ("Plywood 3/4\" 4×8 (wall form)",      52.00,  "sheet"),
        ("Snap ties (5 lb box, ~50 ties)",     38.00,  "box"),
        ("Sawcut joint (per lin ft)",          1.85,   "lin ft"),
        ("Pump truck (small, < 30 yd)",        850.00, "day"),
        ("Pump truck (line pump, per yd)",     5.50,   "cu yd"),
    ]
    pricing_start = r
    for label, price, unit in items:
        ws.cell(r, 1, label).font = LBL; ws.cell(r, 1).border = BOX
        p = ws.cell(r, 2, price); p.font = MONO_BOLD; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.fill = YELLOW; p.border = BOX
        ws.cell(r, 3, unit).font = LBL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["PriceTable"] = DefinedName("PriceTable", attr_text=f"PRICING!$A${pricing_start}:$C${r-1}")

    r += 1
    section(ws, r, "LABOR & OVERHEAD", span=3); r += 1
    labor = [
        ("Foreman labor",                  85.00,  "per hour"),
        ("Finisher labor",                 65.00,  "per hour"),
        ("Helper labor",                   45.00,  "per hour"),
        ("Form + place + finish (per sq ft slab)", 4.50, "per sq ft"),
        ("Footing labor (per lin ft)",     12.00,  "per lin ft"),
        ("Wall labor (per sq ft of wall)", 6.50,   "per sq ft"),
        ("Permit + inspection",            225.00, "flat"),
        ("Excavation / haul-off (flat)",   650.00, "flat"),
        ("Profit margin (decimal)",        0.20,   "0.20 = 20%"),
        ("Sales tax (decimal)",            0.07,   "0.07 = 7%"),
    ]
    labor_start = r
    for label, val, unit in labor:
        ws.cell(r, 1, label).font = LBL_BOLD; ws.cell(r, 1).border = BOX
        v = ws.cell(r, 2, val); v.font = MONO_BOLD; v.alignment = RIGHT; v.fill = YELLOW; v.border = BOX
        if "decimal" in unit:
            v.number_format = "0.0%"
        else:
            v.number_format = '"$"#,##0.00'
        ws.cell(r, 3, unit).font = SMALL; ws.cell(r, 3).border = BOX
        r += 1
    wb.defined_names["Foreman"] = DefinedName("Foreman", attr_text=f"PRICING!$B${labor_start}")
    wb.defined_names["Finisher"] = DefinedName("Finisher", attr_text=f"PRICING!$B${labor_start+1}")
    wb.defined_names["Helper"] = DefinedName("Helper", attr_text=f"PRICING!$B${labor_start+2}")
    wb.defined_names["SlabLab"] = DefinedName("SlabLab", attr_text=f"PRICING!$B${labor_start+3}")
    wb.defined_names["FootLab"] = DefinedName("FootLab", attr_text=f"PRICING!$B${labor_start+4}")
    wb.defined_names["WallLab"] = DefinedName("WallLab", attr_text=f"PRICING!$B${labor_start+5}")
    wb.defined_names["Permit"] = DefinedName("Permit", attr_text=f"PRICING!$B${labor_start+6}")
    wb.defined_names["ExcavFee"] = DefinedName("ExcavFee", attr_text=f"PRICING!$B${labor_start+7}")
    wb.defined_names["Margin"] = DefinedName("Margin", attr_text=f"PRICING!$B${labor_start+8}")
    wb.defined_names["TaxRate"] = DefinedName("TaxRate", attr_text=f"PRICING!$B${labor_start+9}")
    return ws


def sheet_quote(wb):
    ws = wb.create_sheet("QUOTE")
    auto_width(ws, [44, 12, 14, 16])
    banner(ws, "Customer Quote", "Print-ready — review prices in PRICING first", width=4)

    r = 4
    section(ws, r, "PROJECT", span=4); r += 1
    label_value(ws, r, "Customer", "=INPUTS!B6"); r += 1
    label_value(ws, r, "Address", "=INPUTS!B7"); r += 1
    label_value(ws, r, "Date", "=INPUTS!B10", fmt="mm/dd/yyyy"); r += 1
    label_value(ws, r, "Slab", "=SlabL&\" × \"&SlabW&\" × \"&SlabT&\" in (\"&SlabL*SlabW&\" sf)\""); r += 1
    label_value(ws, r, "Footings", "=FootLen&\" lin ft × \"&FootW&\" × \"&FootD&\" in\""); r += 1
    label_value(ws, r, "Total concrete", "=ROUND(SlabYd+FootYd+WallYd,2)&\" cu yd (\"&PSI&\" psi)\""); r += 1

    r += 1
    section(ws, r, "MATERIALS (typical garage slab + perimeter footing)", span=4); r += 1
    headers = ["Description", "Qty", "Unit Price", "Total"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.font = H3; c.fill = YELLOW; c.alignment = CENTER; c.border = BOX
    r += 1
    eq_start = r
    eq_items = [
        ("Concrete delivered, 4000 psi (slab + footing)", "=ROUND(SlabYd+FootYd+WallYd,2)",
         "Concrete delivered, 4000 psi"),
        ("Crushed stone base", "=ROUNDUP(SlabL*SlabW*(BaseDepth/12)/27*1.10,0)",
         'Crushed stone base (3/4" gravel)'),
        ("Vapor barrier rolls (6-mil)", "=ROUNDUP(SlabL*SlabW/200,0)",
         "Vapor barrier (6-mil 200 sf roll)"),
        ("Curing compound (5-gal)", "=ROUNDUP(SlabL*SlabW/200/5,0)",
         "Curing compound (5-gal pail)"),
        ("Rebar #4 (slab grid)", "=SlabBars",
         "Rebar #4 (20-ft stick)"),
        ("Rebar #4 (footing)", "=FootBars",
         "Rebar #4 (20-ft stick)"),
        ("Tie wire rolls", "=ROUNDUP((SlabBars+FootBars+WallBars)*20*0.5/200,0)",
         "Rebar tie wire (3.5 lb roll)"),
        ("Rebar chairs (slab)", "=ROUNDUP(SlabL*SlabW/16,0)",
         'Rebar chair, 3" plastic'),
        ("Continuous bolster (footing)", "=ROUNDUP(FootLen/4,0)",
         "Bolster (continuous chair, 5 ft)"),
        ("2x10 SPF (slab edge form)", "=ROUNDUP((SlabL*2+SlabW*2)/12,0)",
         "2x10 SPF 12-ft (form)"),
        ("2x4 SPF stakes", "=ROUNDUP((SlabL*2+SlabW*2)/4,0)+ROUNDUP(FootLen*2/4,0)",
         "2x4 SPF 8-ft (stake / brace)"),
        ("Form release oil", "=ROUNDUP((SlabL*2+SlabW*2)*(SlabT/12)/100/5,0)",
         "Form release oil (5-gal)"),
        ("Sawcut control joints (lin ft)", "=ROUNDUP(SlabL*SlabW/144,0)*12",
         "Sawcut joint (per lin ft)"),
    ]
    for desc, qty, price_label in eq_items:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, qty); c.font = MONO; c.alignment = RIGHT; c.number_format = "0"; c.border = BOX
        escaped = price_label.replace('"', '""')
        p = ws.cell(r, 3, f'=VLOOKUP("{escaped}",PriceTable,2,FALSE)')
        p.font = MONO; p.alignment = RIGHT; p.number_format = '"$"#,##0.00'; p.border = BOX
        t = ws.cell(r, 4, f"=B{r}*C{r}"); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    eq_end = r - 1

    ws.cell(r, 1, "Materials subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{eq_start}:D{eq_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    eq_subtotal_row = r; r += 1

    r += 1
    section(ws, r, "LABOR & FEES", span=4); r += 1
    labor_lines = [
        ("Slab labor (form + place + finish)",  "=SlabL*SlabW*SlabLab"),
        ("Footing labor (per lin ft)",          "=FootLen*FootLab"),
        ("Wall labor (per sq ft of wall)",      "=WallLen*WallH*WallLab"),
        ("Excavation / haul-off",               "=ExcavFee"),
        ("Permit + inspection",                 "=Permit"),
    ]
    labor_start = r
    for desc, formula in labor_lines:
        ws.cell(r, 1, desc).font = LBL; ws.cell(r, 1).border = BOX
        for col in (2, 3): ws.cell(r, col, "").border = BOX
        t = ws.cell(r, 4, formula); t.font = MONO_BOLD; t.alignment = RIGHT
        t.number_format = '"$"#,##0.00'; t.border = BOX
        r += 1
    labor_end = r - 1

    ws.cell(r, 1, "Labor subtotal").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    s = ws.cell(r, 4, f"=SUM(D{labor_start}:D{labor_end})")
    s.font = MONO_BOLD; s.alignment = RIGHT; s.number_format = '"$"#,##0.00'; s.border = BOX; s.fill = PANEL_FILL
    labor_subtotal_row = r; r += 1

    r += 1
    ws.cell(r, 1, "Cost subtotal (materials + labor)").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    cs = ws.cell(r, 4, f"=D{eq_subtotal_row}+D{labor_subtotal_row}")
    cs.font = MONO_BOLD; cs.alignment = RIGHT; cs.number_format = '"$"#,##0.00'; cs.border = BOX
    cost_row = r; r += 1

    ws.cell(r, 1, "Profit margin").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    m = ws.cell(r, 4, f"=D{cost_row}*Margin")
    m.font = MONO_BOLD; m.alignment = RIGHT; m.number_format = '"$"#,##0.00'; m.border = BOX
    margin_row = r; r += 1

    ws.cell(r, 1, "Pre-tax total").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    pt = ws.cell(r, 4, f"=D{cost_row}+D{margin_row}")
    pt.font = MONO_BOLD; pt.alignment = RIGHT; pt.number_format = '"$"#,##0.00'; pt.border = BOX
    pretax_row = r; r += 1

    ws.cell(r, 1, "Sales tax").font = LBL_BOLD; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").border = BOX
    tx = ws.cell(r, 4, f"=D{pretax_row}*TaxRate")
    tx.font = MONO_BOLD; tx.alignment = RIGHT; tx.number_format = '"$"#,##0.00'; tx.border = BOX
    tax_row = r; r += 1

    ws.cell(r, 1, "TOTAL DUE").font = H2; ws.cell(r, 1).fill = YELLOW; ws.cell(r, 1).border = BOX
    for col in (2, 3): ws.cell(r, col, "").fill = YELLOW; ws.cell(r, col, "").border = BOX
    gt = ws.cell(r, 4, f"=D{pretax_row}+D{tax_row}")
    gt.font = Font(name="Arial Black", size=14, color=INK, bold=True)
    gt.alignment = RIGHT; gt.number_format = '"$"#,##0.00'; gt.fill = YELLOW; gt.border = BOX
    ws.row_dimensions[r].height = 28
    r += 2

    section(ws, r, "ACCEPTANCE", span=4); r += 1
    a = ws.cell(r, 1, "Customer signature: __________________________________  Date: ____________")
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 2
    a = ws.cell(r, 1, '=CONCATENATE("Estimator: ",INPUTS!B9)')
    a.font = LBL; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    n = ws.cell(r, 1,
        "Quote valid for 30 days. Concrete pricing subject to fuel + cement "
        "surcharges. Pour scheduled within 1-2 weeks of signed acceptance, "
        "weather permitting (no pour below 40 °F or above 90 °F without admixtures).")
    n.font = SMALL
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:D{r+2}"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_inputs(wb)
    sheet_slab(wb)
    sheet_footings_walls(wb)
    sheet_rebar(wb)
    sheet_pricing(wb)
    sheet_quote(wb)

    order = ["README", "INPUTS", "SLAB", "FOOTINGS-WALLS", "REBAR", "PRICING", "QUOTE"]
    wb._sheets = [wb[name] for name in order]

    out = Path(__file__).parent / "dist" / "Concrete-Pro-Toolkit-v1.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"Built {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
